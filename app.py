import os
from pathlib import Path
import sys
import csv
import json
import base64
import uuid
import secrets
import shutil
import zipfile
import sqlite3
from io import StringIO, BytesIO
from urllib.parse import quote_plus
import struct
import html
import re
import unicodedata
import logging
import traceback
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, Response, abort, has_request_context, jsonify
from reportlab.lib.pagesizes import landscape, A6, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from PIL import Image, ImageDraw, ImageFont, ImageEnhance, ImageFilter
import qrcode

try:
    from webauthn import generate_registration_options, verify_registration_response, generate_authentication_options, verify_authentication_response, options_to_json, base64url_to_bytes
    from webauthn.helpers.structs import PublicKeyCredentialDescriptor, AuthenticatorSelectionCriteria, AuthenticatorAttachment, ResidentKeyRequirement, UserVerificationRequirement
    WEBAUTHN_AVAILABLE = True
except Exception:
    WEBAUTHN_AVAILABLE = False

SOURCE_DIR = os.path.abspath(os.path.dirname(__file__))
if getattr(sys, "frozen", False):
    # Quand l'application est transformée en .exe, la base de données et les uploads
    # doivent rester à côté de l'exécutable pour ne pas être perdus.
    BASE_DIR = os.path.dirname(sys.executable)
    RESOURCE_DIR = getattr(sys, "_MEIPASS", BASE_DIR)
else:
    BASE_DIR = SOURCE_DIR
    RESOURCE_DIR = SOURCE_DIR

TEMPLATE_DIR = os.path.join(RESOURCE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
BUNDLED_STATIC_DIR = os.path.join(RESOURCE_DIR, "static")

# En mode .exe, les fichiers statiques fournis avec l'application sont copiés
# une fois à côté de l'exécutable. Les logos/photos ajoutés ensuite restent persistants.
if getattr(sys, "frozen", False) and os.path.isdir(BUNDLED_STATIC_DIR):
    shutil.copytree(BUNDLED_STATIC_DIR, STATIC_DIR, dirs_exist_ok=True)

DB_PATH = os.environ.get("DATABASE_PATH", os.path.join(BASE_DIR, "asbl.db"))
UPLOAD_ROOT = os.environ.get("UPLOAD_ROOT", os.path.join(STATIC_DIR, "uploads"))
RDC_FLAG_REL = "img/drapeau_rdc.jpg"
RDC_FLAG_ABS = os.path.join(STATIC_DIR, RDC_FLAG_REL)
ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp", "pdf", "doc", "docx", "xls", "xlsx"}
APP_VERSION = "47.0.0"
APP_RELEASE_NAME = "FOBAK Manager Pro — Menu mobile et déconnexion V47"
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
logging.basicConfig(filename=os.path.join(LOG_DIR, "application.log"), level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR, static_url_path="/static")

def _persistent_secret_key():
    configured = os.environ.get("SECRET_KEY", "").strip()
    if configured:
        return configured
    secret_file = os.path.join(BASE_DIR, ".secret_key")
    try:
        if os.path.exists(secret_file):
            value = Path(secret_file).read_text(encoding="utf-8").strip()
            if len(value) >= 32:
                return value
        value = secrets.token_hex(32)
        Path(secret_file).write_text(value, encoding="utf-8")
        return value
    except OSError:
        return secrets.token_hex(32)

app.secret_key = _persistent_secret_key()
app.config.update(
    MAX_CONTENT_LENGTH=128 * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("COOKIE_SECURE", "0") == "1",
)

ROLE_LABELS = {
    "super_admin": "Super administrateur",
    "president": "Président National",
    "secretary": "Secrétaire National / Général",
    "national_secretary": "Secrétaire National",
    "national_treasurer": "Trésorier national",
    "provincial_president": "Président provincial",
    "provincial_admin": "Coordonnateur / Administrateur provincial",
    "provincial_secretary": "Secrétaire provincial",
    "provincial_treasurer": "Trésorier provincial",
    "local_admin": "Responsable local",
    "registration_agent": "Agent d'enregistrement",
    "member": "Membre",
    "visitor": "Visiteur / Invité",
}

NATIONAL_ROLES = {"super_admin", "president", "secretary", "national_secretary", "national_treasurer"}
PROVINCIAL_ROLES = {"provincial_president", "provincial_admin", "provincial_secretary", "provincial_treasurer", "registration_agent"}
ADMIN_ROLES = {"super_admin", "president", "secretary", "national_secretary", "national_treasurer", "provincial_president", "provincial_admin", "provincial_secretary", "provincial_treasurer", "local_admin", "registration_agent"}
DESTRUCTIVE_ROLES = {"super_admin", "secretary", "national_secretary"}

PROVINCES = [
    "Bas-Uele", "Équateur", "Haut-Katanga", "Haut-Lomami", "Haut-Uele", "Ituri", "Kasaï",
    "Kasaï-Central", "Kasaï-Oriental", "Kinshasa", "Kongo-Central", "Kwango", "Kwilu", "Lomami",
    "Lualaba", "Mai-Ndombe", "Maniema", "Mongala", "Nord-Kivu", "Nord-Ubangi", "Sankuru",
    "Sud-Kivu", "Sud-Ubangi", "Tanganyika", "Tshopo", "Tshuapa"
]

TERRITOIRES = {
    "Bas-Uele": ["Aketi", "Ango", "Bambesa", "Bondo", "Buta", "Poko"],
    "Équateur": ["Basankusu", "Bikoro", "Bolomba", "Bomongo", "Ingende", "Lukolela", "Makanza"],
    "Haut-Katanga": ["Kambove", "Kasenga", "Kipushi", "Mitwaba", "Pweto", "Sakania"],
    "Haut-Lomami": ["Bukama", "Kabongo", "Kamina", "Kaniama", "Malemba-Nkulu"],
    "Haut-Uele": ["Dungu", "Faradje", "Niangara", "Rungu", "Wamba", "Watsa"],
    "Ituri": ["Aru", "Djugu", "Irumu", "Mahagi", "Mambasa"],
    "Kasaï": ["Dekese", "Ilebo", "Kamonia", "Luebo", "Mweka"],
    "Kasaï-Central": ["Demba", "Dibaya", "Dimbelenge", "Kazumba", "Luiza"],
    "Kasaï-Oriental": ["Kabeya-Kamwanga", "Katanda", "Lupatapata", "Miabi", "Tshilenge"],
    "Kinshasa": ["Bandalungwa", "Barumbu", "Bumbu", "Gombe", "Kalamu", "Kasa-Vubu", "Kimbanseke", "Kinshasa", "Kintambo", "Kisenso", "Lemba", "Limete", "Lingwala", "Makala", "Maluku", "Masina", "Matete", "Mont-Ngafula", "N'Djili", "N'Sele", "Ngaba", "Ngaliema", "Ngiri-Ngiri", "Selembao"],
    "Kongo-Central": ["Kasangulu", "Kimvula", "Lukula", "Luozi", "Madimba", "Mbanza-Ngungu", "Muanda", "Sekebanza", "Songololo", "Tshela"],
    "Kwango": ["Feshi", "Kahemba", "Kasongo-Lunda", "Kenge", "Popokabaka"],
    "Kwilu": ["Bagata", "Bulungu", "Gungu", "Idiofa", "Masi-Manimba"],
    "Lomami": ["Kabinda", "Kamiji", "Lubao", "Luilu", "Ngandajika"],
    "Lualaba": ["Dilolo", "Kapanga", "Lubudi", "Mutshatsha", "Sandoa"],
    "Mai-Ndombe": ["Bolobo", "Inongo", "Kiri", "Kutu", "Kwamouth", "Mushie", "Oshwe", "Yumbi"],
    "Maniema": ["Kabambare", "Kailo", "Kasongo", "Kibombo", "Lubutu", "Pangi", "Punia"],
    "Mongala": ["Bongandanga", "Bumba", "Lisala"],
    "Nord-Kivu": ["Beni", "Lubero", "Masisi", "Nyiragongo", "Rutshuru", "Walikale"],
    "Nord-Ubangi": ["Bosobolo", "Businga", "Mobayi-Mbongo", "Yakoma"],
    "Sankuru": ["Katako-Kombe", "Kole", "Lodja", "Lomela", "Lubefu", "Lusambo"],
    "Sud-Kivu": ["Fizi", "Idjwi", "Kabare", "Kalehe", "Mwenga", "Shabunda", "Uvira", "Walungu"],
    "Sud-Ubangi": ["Budjala", "Gemena", "Kungu", "Libenge"],
    "Tanganyika": ["Kabalo", "Kalemie", "Kongolo", "Manono", "Moba", "Nyunzu"],
    "Tshopo": ["Bafwasende", "Banalia", "Basoko", "Isangi", "Opala", "Ubundu", "Yahuma"],
    "Tshuapa": ["Befale", "Boende", "Bokungu", "Djolu", "Ikela", "Monkoto"],
}

# Indicatifs internationaux. La liste principale est générée avec phonenumbers/pycountry si disponibles.
FALLBACK_COUNTRY_CODES = [
    ("CD", "+243", "République démocratique du Congo"), ("CG", "+242", "Congo"),
    ("AO", "+244", "Angola"), ("BI", "+257", "Burundi"), ("RW", "+250", "Rwanda"),
    ("UG", "+256", "Ouganda"), ("TZ", "+255", "Tanzanie"), ("ZM", "+260", "Zambie"),
    ("ZA", "+27", "Afrique du Sud"), ("BE", "+32", "Belgique"), ("FR", "+33", "France"),
    ("GB", "+44", "Royaume-Uni"), ("US", "+1", "États-Unis / Canada"), ("DE", "+49", "Allemagne"),
    ("IT", "+39", "Italie"), ("ES", "+34", "Espagne"), ("PT", "+351", "Portugal"),
    ("CN", "+86", "Chine"), ("IN", "+91", "Inde"), ("AE", "+971", "Émirats arabes unis")
]

def build_country_codes():
    try:
        import phonenumbers
        import pycountry
        rows=[]
        for region in sorted(phonenumbers.SUPPORTED_REGIONS):
            code=phonenumbers.country_code_for_region(region)
            country=pycountry.countries.get(alpha_2=region)
            name=country.name if country else region
            rows.append((region, f"+{code}", name))
        rows.sort(key=lambda x: x[2])
        return rows
    except Exception:
        return FALLBACK_COUNTRY_CODES

COUNTRY_CODES = build_country_codes()


EDUCATION_LEVELS = ["Primaire", "Secondaire", "Diplôme d'État", "Graduat", "Licence", "Master", "Doctorat", "Formation professionnelle", "Autre"]
EXPERIENCE_LEVELS = ["Aucune", "Moins d'un an", "1 à 3 ans", "3 à 5 ans", "5 à 10 ans", "Plus de 10 ans"]
STUDY_CHECKBOXES = ["Primaire", "Secondaire", "1er cycle", "2ème cycle", "3ème cycle"]
MARITAL_STATUS = ["Célibataire", "Marié(e)", "Veuf/Veuve", "Divorcé(e)", "Autre"]


AVAILABLE_LANGUAGES = {
    "fr": "Français",
    "ln": "Lingala",
    "en": "Anglais",
    "pt": "Portugais",
    "es": "Espagnol",
    "sw": "Swahili",
    "kg": "Kikongo",
    "ts": "Tshiluba",
}

UI_TRANSLATIONS = {
    "Accueil": {"ln":"Ebandeli", "en":"Home", "pt":"Início", "es":"Inicio", "sw":"Mwanzo", "kg":"Luyantiku", "ts":"Ntendekelu"},
    "Tableau de bord": {"ln":"Etando ya mosala", "en":"Dashboard", "pt":"Painel", "es":"Panel", "sw":"Dashibodi", "kg":"Kipangu kya kisalu", "ts":"Tshibangu tsha mudimu"},
    "Membres": {"ln":"Ba membɛ", "en":"Members", "pt":"Membros", "es":"Miembros", "sw":"Wanachama", "kg":"Bampangi", "ts":"Bantu ba kabidi"},
    "Adhésions": {"ln":"Kokoma membɛ", "en":"Membership", "pt":"Adesões", "es":"Adhesiones", "sw":"Usajili", "kg":"Kota", "ts":"Kudifundisha"},
    "Cotisations": {"ln":"Makabo", "en":"Cotisations", "pt":"Contribuições", "es":"Contribuciones", "sw":"Michango", "kg":"Makabu", "ts":"Mipatshila"},
    "Activités": {"ln":"Misala", "en":"Activities", "pt":"Atividades", "es":"Actividades", "sw":"Shughuli", "kg":"Bisalu", "ts":"Midimu"},
    "Notifications": {"ln":"Mayebisi", "en":"Notifications", "pt":"Notificações", "es":"Notificaciones", "sw":"Taarifa", "kg":"Bansangu", "ts":"Meyebisha"},
    "Paramètres": {"ln":"Bobongisi", "en":"Settings", "pt":"Definições", "es":"Configuración", "sw":"Mipangilio", "kg":"Masono", "ts":"Mipangilu"},
    "Connexion": {"ln":"Kokota", "en":"Login", "pt":"Entrar", "es":"Conexión", "sw":"Ingia", "kg":"Kota", "ts":"Kubuela"},
    "Déconnexion": {"ln":"Kobima", "en":"Logout", "pt":"Sair", "es":"Salir", "sw":"Ondoka", "kg":"Basika", "ts":"Kupatuka"},
    "Devenir membre": {"ln":"Kokoma membɛ", "en":"Become a member", "pt":"Tornar-se membro", "es":"Hacerse miembro", "sw":"Kuwa mwanachama", "kg":"Kuma mpangi", "ts":"Kukuma mudi wa kabidi"},
    "Recherche": {"ln":"Boluki", "en":"Search", "pt":"Pesquisar", "es":"Buscar", "sw":"Tafuta", "kg":"Sosa", "ts":"Longa"},
    "Aide intelligente": {"ln":"Lisungi ya mayele", "en":"Smart help", "pt":"Ajuda inteligente", "es":"Ayuda inteligente", "sw":"Msaada wa haraka", "kg":"Lusadisu", "ts":"Dikuatshisha"},
    "Signaler un problème": {"ln":"Yebisa mokakatano", "en":"Report a problem", "pt":"Reportar problema", "es":"Informar un problema", "sw":"Ripoti tatizo", "kg":"Zabisa mpasi", "ts":"Manya tshilumbu"},
    "Bienvenue": {"ln":"Boyei malamu", "en":"Welcome", "pt":"Bem-vindo", "es":"Bienvenido", "sw":"Karibu", "kg":"Mbote", "ts":"Moyo"},
}


def db():
    # WAL et délai d'attente réduisent les erreurs « database is locked »
    # lors des accès simultanés sur le serveur local.
    con = sqlite3.connect(DB_PATH, timeout=30)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA foreign_keys = ON")
    con.execute("PRAGMA busy_timeout = 30000")
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA synchronous = NORMAL")
    return con


def init_db():
    os.makedirs(UPLOAD_ROOT, exist_ok=True)
    os.makedirs(os.path.join(STATIC_DIR, "img"), exist_ok=True)
    for sub in ["carousel", "logos", "photos", "signatures", "cards", "activities", "support", "backups", "official_docs"]:
        os.makedirs(os.path.join(UPLOAD_ROOT, sub), exist_ok=True)
    os.makedirs(os.path.join(BASE_DIR, "backups"), exist_ok=True)
    con = db()
    cur = con.cursor()
    cur.executescript('''
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    );
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE,
        phone TEXT UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'member',
        province TEXT,
        localite TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS active_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_token TEXT UNIQUE NOT NULL,
        user_id INTEGER NOT NULL,
        login_at TEXT NOT NULL,
        last_seen TEXT NOT NULL,
        ip_address TEXT,
        user_agent TEXT,
        active INTEGER NOT NULL DEFAULT 1,
        logout_at TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id)
    );
    CREATE INDEX IF NOT EXISTS idx_active_sessions_user ON active_sessions(user_id);
    CREATE INDEX IF NOT EXISTS idx_active_sessions_active ON active_sessions(active,last_seen);
    CREATE TABLE IF NOT EXISTS passkey_credentials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        credential_id TEXT NOT NULL UNIQUE,
        public_key TEXT NOT NULL,
        sign_count INTEGER NOT NULL DEFAULT 0,
        transports TEXT,
        label TEXT,
        created_at TEXT NOT NULL,
        last_used_at TEXT,
        active INTEGER DEFAULT 1
    );
    CREATE INDEX IF NOT EXISTS idx_passkey_user ON passkey_credentials(user_id,active);
    CREATE TABLE IF NOT EXISTS member_applications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        gender TEXT,
        email TEXT NOT NULL,
        phone TEXT NOT NULL,
        nationality TEXT DEFAULT 'Congolaise',
        province TEXT,
        territory TEXT,
        commune TEXT,
        localite TEXT,
        physical_address TEXT,
        birth_date TEXT,
        birth_place TEXT,
        marital_status TEXT,
        profession TEXT,
        education TEXT,
        studies_done TEXT,
        experience TEXT,
        motivation TEXT,
        photo_path TEXT,
        status TEXT NOT NULL DEFAULT 'pending',
        created_at TEXT NOT NULL,
        reviewed_at TEXT,
        reviewed_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS members (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        code TEXT UNIQUE NOT NULL,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        gender TEXT,
        email TEXT NOT NULL,
        phone TEXT NOT NULL,
        nationality TEXT DEFAULT 'Congolaise',
        province TEXT,
        territory TEXT,
        commune TEXT,
        localite TEXT,
        physical_address TEXT,
        birth_date TEXT,
        birth_place TEXT,
        marital_status TEXT,
        profession TEXT,
        education TEXT,
        studies_done TEXT,
        experience TEXT,
        photo_path TEXT,
        adhesion_number TEXT,
        joined_at TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        approved_by INTEGER,
        deleted_at TEXT
    );
    CREATE TABLE IF NOT EXISTS card_renewals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        old_expiry TEXT,
        new_expiry TEXT NOT NULL,
        renewed_at TEXT NOT NULL,
        renewed_by INTEGER,
        notes TEXT,
        FOREIGN KEY(member_id) REFERENCES members(id)
    );
    CREATE INDEX IF NOT EXISTS idx_card_renewals_member ON card_renewals(member_id);
    CREATE TABLE IF NOT EXISTS activities (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        subtitle TEXT,
        body TEXT,
        image_path TEXT,
        youtube_url TEXT,
        published_at TEXT NOT NULL,
        author_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS videos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        youtube_url TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER,
        amount REAL NOT NULL,
        currency TEXT DEFAULT 'CDF',
        method TEXT,
        reference TEXT,
        status TEXT DEFAULT 'pending',
        paid_at TEXT,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        target_scope TEXT NOT NULL,
        province TEXT,
        email TEXT,
        phone TEXT,
        message TEXT NOT NULL,
        created_at TEXT NOT NULL,
        sent INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS carousel_images (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        image_path TEXT NOT NULL,
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS adhesion_fields (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        section TEXT NOT NULL DEFAULT 'Autres informations',
        label TEXT NOT NULL,
        field_type TEXT NOT NULL DEFAULT 'text',
        options TEXT,
        required INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1,
        sort_order INTEGER DEFAULT 100,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS audit_logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        action TEXT NOT NULL,
        target_type TEXT,
        target_id INTEGER,
        details TEXT,
        created_at TEXT NOT NULL
    );
    ''')
    # Migration douce : ajoute les nouvelles colonnes si la base existe déjà.
    def ensure_column(table, column, definition):
        """Ajoute une colonne uniquement si la table existe déjà.

        Certaines installations démarrent avec une base vide ou une ancienne base
        partielle. La migration ne doit jamais échouer simplement parce qu'une table
        sera créée quelques lignes plus loin.
        """
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", table or ""):
            raise ValueError(f"Nom de table invalide: {table!r}")
        if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", column or ""):
            raise ValueError(f"Nom de colonne invalide: {column!r}")
        table_exists = cur.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
        ).fetchone()
        if not table_exists:
            return False
        existing = [r[1] for r in cur.execute(f'PRAGMA table_info("{table}")').fetchall()]
        if column not in existing:
            cur.execute(f'ALTER TABLE "{table}" ADD COLUMN "{column}" {definition}')
        return True

    for table in ["member_applications", "members"]:
        ensure_column(table, "birth_place", "TEXT")
        ensure_column(table, "marital_status", "TEXT")
        ensure_column(table, "studies_done", "TEXT")
        ensure_column(table, "physical_address", "TEXT")
        ensure_column(table, "custom_fields", "TEXT")
    ensure_column("members", "adhesion_number", "TEXT")
    ensure_column("members", "status", "TEXT DEFAULT 'active'")
    ensure_column("members", "updated_at", "TEXT")
    ensure_column("users", "deleted_at", "TEXT")
    ensure_column("users", "last_login", "TEXT")
    ensure_column("users", "force_password_change", "INTEGER DEFAULT 0")
    ensure_column("users", "failed_login_count", "INTEGER DEFAULT 0")
    ensure_column("users", "locked_until", "TEXT")
    ensure_column("users", "password_changed_at", "TEXT")
    ensure_column("users", "first_name", "TEXT")
    ensure_column("users", "last_name", "TEXT")
    ensure_column("users", "photo_path", "TEXT")
    ensure_column("users", "created_by", "INTEGER")
    ensure_column("users", "preferred_language", "TEXT DEFAULT 'fr'")
    ensure_column("members", "created_by", "INTEGER")
    ensure_column("members", "is_administrative", "INTEGER DEFAULT 0")
    ensure_column("members", "role_label", "TEXT")
    ensure_column("activities", "image_fit", "TEXT DEFAULT 'cover'")
    ensure_column("activities", "status", "TEXT DEFAULT 'approved'")
    ensure_column("activities", "province", "TEXT")
    ensure_column("activities", "localite", "TEXT")
    ensure_column("activities", "approved_at", "TEXT")
    ensure_column("activities", "approved_by", "INTEGER")
    ensure_column("activities", "category", "TEXT DEFAULT 'Activité'")
    ensure_column("carousel_images", "message", "TEXT")
    ensure_column("carousel_images", "image_fit", "TEXT DEFAULT 'cover'")
    ensure_column("payments", "contribution_type", "TEXT DEFAULT 'Cotisation'")
    ensure_column("payments", "created_by", "INTEGER")
    ensure_column("audit_logs", "ip_address", "TEXT")
    ensure_column("audit_logs", "user_agent", "TEXT")
    ensure_column("audit_logs", "route", "TEXT")
    ensure_column("audit_logs", "status", "TEXT DEFAULT 'success'")
    ensure_column("users", "pin_hash", "TEXT")
    ensure_column("users", "pin_length", "INTEGER")
    ensure_column("users", "pin_failed_count", "INTEGER DEFAULT 0")
    ensure_column("users", "pin_locked_until", "TEXT")
    ensure_column("users", "passkey_enabled", "INTEGER DEFAULT 0")

    cur.executescript("""
    CREATE TABLE IF NOT EXISTS support_tickets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tracking_code TEXT UNIQUE,
        user_id INTEGER,
        member_id INTEGER,
        full_name TEXT,
        email TEXT,
        phone TEXT,
        province TEXT,
        localite TEXT,
        category TEXT DEFAULT 'Général',
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        attachment_path TEXT,
        status TEXT DEFAULT 'new',
        priority TEXT DEFAULT 'normal',
        assigned_to INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT,
        closed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS support_ticket_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_id INTEGER NOT NULL,
        user_id INTEGER,
        author_name TEXT,
        message TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS internal_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        role TEXT,
        province TEXT,
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        link TEXT,
        read_at TEXT,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS fiscal_years (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT UNIQUE NOT NULL,
        start_date TEXT,
        end_date TEXT,
        status TEXT DEFAULT 'open',
        closed_at TEXT,
        closed_by INTEGER,
        note TEXT
    );
    CREATE TABLE IF NOT EXISTS db_migrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        applied_at TEXT NOT NULL,
        details TEXT
    );
    """)
    for col, definition in {
        "tracking_code": "TEXT", "user_id": "INTEGER", "member_id": "INTEGER", "full_name": "TEXT",
        "email": "TEXT", "phone": "TEXT", "province": "TEXT", "localite": "TEXT",
        "category": "TEXT DEFAULT 'Général'", "title": "TEXT", "message": "TEXT", "attachment_path": "TEXT",
        "status": "TEXT DEFAULT 'new'", "priority": "TEXT DEFAULT 'normal'", "assigned_to": "INTEGER",
        "created_at": "TEXT", "updated_at": "TEXT", "closed_at": "TEXT"
    }.items():
        ensure_column("support_tickets", col, definition)
    for col, definition in {
        "user_id":"INTEGER", "role":"TEXT", "province":"TEXT", "title":"TEXT", "message":"TEXT",
        "link":"TEXT", "read_at":"TEXT", "created_at":"TEXT",
        "category":"TEXT DEFAULT 'general'", "sound_enabled":"INTEGER DEFAULT 0"
    }.items():
        ensure_column("internal_notifications", col, definition)
    cur.execute("UPDATE members SET created_by=approved_by WHERE (created_by IS NULL OR created_by='') AND approved_by IS NOT NULL")
    cur.execute("UPDATE members SET is_administrative=0 WHERE is_administrative IS NULL")
    cur.execute("UPDATE activities SET status='approved' WHERE status IS NULL OR status=''")
    cur.execute("INSERT OR IGNORE INTO db_migrations(name, applied_at, details) VALUES(?,?,?)", ("2026_support_stability_mobile", now(), "Pages légales, centre stabilité, tickets support, notifications internes, PWA, clôture exercice"))
    cur.execute("INSERT OR IGNORE INTO db_migrations(name, applied_at, details) VALUES(?,?,?)", ("2026_production_security_backup", now(), "Sécurité production, changement mot de passe obligatoire, verrouillage connexion, sauvegarde et centre production"))
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS sync_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entity_type TEXT NOT NULL,
        entity_id TEXT,
        action TEXT NOT NULL,
        payload TEXT,
        status TEXT DEFAULT 'pending',
        attempts INTEGER DEFAULT 0,
        last_error TEXT,
        created_at TEXT NOT NULL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS sync_conflicts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        entity_type TEXT NOT NULL,
        entity_id TEXT,
        local_payload TEXT,
        remote_payload TEXT,
        status TEXT DEFAULT 'open',
        resolution TEXT,
        created_at TEXT NOT NULL,
        resolved_at TEXT,
        resolved_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS app_updates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        version TEXT NOT NULL,
        title TEXT,
        notes TEXT,
        package_url TEXT,
        published_at TEXT,
        installed_at TEXT,
        status TEXT DEFAULT 'available'
    );
    """)
    cur.execute("INSERT OR IGNORE INTO db_migrations(name, applied_at, details) VALUES(?,?,?)", ("2026_sync_backup_trash_updates", now(), "Centre de synchronisation, conflits, sauvegarde, corbeille et mises à jour"))
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        province TEXT,
        status TEXT DEFAULT 'en cours',
        description TEXT,
        budget REAL DEFAULT 0,
        created_at TEXT NOT NULL,
        author_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        file_path TEXT,
        public INTEGER DEFAULT 0,
        province TEXT,
        created_at TEXT NOT NULL,
        author_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS official_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        document_type TEXT NOT NULL DEFAULT 'statuts',
        title TEXT NOT NULL,
        description TEXT,
        file_path TEXT,
        version_label TEXT,
        adoption_date TEXT,
        effective_date TEXT,
        public INTEGER DEFAULT 1,
        active INTEGER DEFAULT 1,
        deleted_at TEXT,
        created_at TEXT NOT NULL,
        created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS statute_sections (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        section_key TEXT UNIQUE NOT NULL,
        title TEXT NOT NULL,
        content TEXT,
        sort_order INTEGER DEFAULT 100,
        public INTEGER DEFAULT 1,
        active INTEGER DEFAULT 1,
        updated_at TEXT,
        updated_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS support_tickets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tracking_code TEXT UNIQUE,
        user_id INTEGER,
        member_id INTEGER,
        full_name TEXT,
        email TEXT,
        phone TEXT,
        province TEXT,
        localite TEXT,
        category TEXT DEFAULT 'Général',
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        attachment_path TEXT,
        status TEXT DEFAULT 'new',
        priority TEXT DEFAULT 'normal',
        assigned_to INTEGER,
        created_at TEXT NOT NULL,
        updated_at TEXT,
        closed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS support_ticket_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_id INTEGER NOT NULL,
        user_id INTEGER,
        author_name TEXT,
        message TEXT NOT NULL,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS internal_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        role TEXT,
        province TEXT,
        title TEXT NOT NULL,
        message TEXT NOT NULL,
        link TEXT,
        read_at TEXT,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS fiscal_years (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        label TEXT UNIQUE NOT NULL,
        start_date TEXT,
        end_date TEXT,
        status TEXT DEFAULT 'open',
        closed_at TEXT,
        closed_by INTEGER,
        note TEXT
    );
    CREATE TABLE IF NOT EXISTS db_migrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        applied_at TEXT NOT NULL,
        details TEXT
    );
    """)


    # Modules ajoutés avant publication : rôles personnalisables, trésorerie, première configuration.
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS role_permissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        role_key TEXT NOT NULL,
        role_label TEXT NOT NULL,
        permission_key TEXT NOT NULL,
        allowed INTEGER DEFAULT 0,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS treasury_entries (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL DEFAULT 'entrée',
        category TEXT,
        description TEXT,
        amount REAL NOT NULL DEFAULT 0,
        currency TEXT DEFAULT 'CDF',
        method TEXT,
        reference TEXT,
        province TEXT,
        localite TEXT,
        entry_date TEXT,
        created_by INTEGER,
        created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS app_setup (
        id INTEGER PRIMARY KEY CHECK (id=1),
        completed INTEGER DEFAULT 0,
        completed_at TEXT,
        completed_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS provincial_offices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        province TEXT UNIQUE NOT NULL,
        president_name TEXT,
        president_function TEXT DEFAULT 'Président provincial',
        president_signature_path TEXT,
        secretary_name TEXT,
        secretary_function TEXT DEFAULT 'Secrétaire provincial',
        secretary_signature_path TEXT,
        office_address TEXT,
        office_phones TEXT,
        updated_at TEXT,
        updated_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS monthly_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        province TEXT NOT NULL,
        report_month TEXT NOT NULL,
        status TEXT DEFAULT 'generated',
        generated_at TEXT NOT NULL,
        generated_by INTEGER,
        submitted_at TEXT,
        submitted_by INTEGER,
        UNIQUE(province, report_month)
    );
    """)
    for col, definition in {
        "type": "TEXT NOT NULL DEFAULT 'entrée'", "category": "TEXT", "description": "TEXT", "amount": "REAL NOT NULL DEFAULT 0",
        "currency": "TEXT DEFAULT 'CDF'", "method": "TEXT", "reference": "TEXT", "province": "TEXT", "localite": "TEXT",
        "entry_date": "TEXT", "created_by": "INTEGER", "created_at": "TEXT"
    }.items():
        ensure_column("treasury_entries", col, definition)

    # ===== V35 : Services nationaux et Centre de santé =====
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS foundation_services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL UNIQUE,
        name TEXT NOT NULL,
        description TEXT,
        scope TEXT DEFAULT 'national',
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS service_structures (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER NOT NULL,
        parent_id INTEGER,
        code TEXT NOT NULL UNIQUE,
        name TEXT NOT NULL,
        structure_type TEXT DEFAULT 'antenne',
        province TEXT,
        territory TEXT,
        commune TEXT,
        address TEXT,
        phone TEXT,
        email TEXT,
        manager_member_id INTEGER,
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT,
        updated_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS service_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER,
        name TEXT NOT NULL,
        description TEXT,
        level TEXT DEFAULT 'structure',
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        UNIQUE(service_id,name)
    );
    CREATE TABLE IF NOT EXISTS service_staff_assignments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        service_id INTEGER NOT NULL,
        structure_id INTEGER,
        role_id INTEGER NOT NULL,
        start_date TEXT,
        end_date TEXT,
        status TEXT DEFAULT 'active',
        notes TEXT,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        UNIQUE(member_id,service_id,structure_id,role_id)
    );
    CREATE TABLE IF NOT EXISTS service_modules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER NOT NULL,
        code TEXT NOT NULL,
        name TEXT NOT NULL,
        description TEXT,
        active INTEGER DEFAULT 1,
        config_json TEXT DEFAULT '{}',
        created_at TEXT NOT NULL,
        created_by INTEGER,
        UNIQUE(service_id,code)
    );
    CREATE TABLE IF NOT EXISTS health_centers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        code TEXT NOT NULL UNIQUE,
        name TEXT NOT NULL,
        province TEXT NOT NULL,
        territory TEXT,
        commune TEXT,
        address TEXT,
        phone TEXT,
        email TEXT,
        manager_member_id INTEGER,
        opening_date TEXT,
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS health_roles (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        description TEXT,
        level TEXT DEFAULT 'centre',
        active INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_staff (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        member_id INTEGER NOT NULL,
        center_id INTEGER NOT NULL,
        role_id INTEGER NOT NULL,
        specialty TEXT,
        professional_number TEXT,
        diploma TEXT,
        engagement_date TEXT,
        status TEXT DEFAULT 'active',
        created_at TEXT NOT NULL,
        created_by INTEGER,
        UNIQUE(member_id, center_id, role_id)
    );
    CREATE TABLE IF NOT EXISTS patients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        patient_code TEXT NOT NULL UNIQUE,
        center_id INTEGER NOT NULL,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        gender TEXT,
        birth_date TEXT,
        phone TEXT,
        address TEXT,
        province TEXT,
        emergency_contact TEXT,
        blood_group TEXT,
        allergies TEXT,
        medical_history TEXT,
        photo_path TEXT,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS consultations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        consultation_code TEXT NOT NULL UNIQUE,
        patient_id INTEGER NOT NULL,
        center_id INTEGER NOT NULL,
        staff_id INTEGER,
        consultation_date TEXT NOT NULL,
        reason TEXT,
        complaints TEXT,
        disease_history TEXT,
        temperature TEXT,
        blood_pressure TEXT,
        heart_rate TEXT,
        respiratory_rate TEXT,
        weight TEXT,
        height TEXT,
        clinical_exam TEXT,
        diagnosis TEXT,
        requested_tests TEXT,
        treatment TEXT,
        orientation TEXT,
        next_appointment TEXT,
        notes TEXT,
        status TEXT DEFAULT 'draft',
        created_at TEXT NOT NULL,
        created_by INTEGER,
        validated_at TEXT,
        validated_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS medical_form_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        form_type TEXT NOT NULL,
        scope TEXT DEFAULT 'national',
        province TEXT,
        center_id INTEGER,
        fields_json TEXT NOT NULL DEFAULT '[]',
        source_file TEXT,
        import_status TEXT DEFAULT 'ready',
        active INTEGER DEFAULT 1,
        version INTEGER DEFAULT 1,
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS executive_numbers (id INTEGER PRIMARY KEY AUTOINCREMENT,number TEXT NOT NULL UNIQUE,position_name TEXT NOT NULL,member_id INTEGER,status TEXT DEFAULT 'reserved',assigned_at TEXT,assigned_by INTEGER,released_at TEXT,notes TEXT,created_at TEXT NOT NULL,created_by INTEGER,updated_at TEXT,deleted_at TEXT,deleted_by INTEGER);
    CREATE TABLE IF NOT EXISTS health_card_requests (id INTEGER PRIMARY KEY AUTOINCREMENT,request_code TEXT NOT NULL UNIQUE,staff_id INTEGER NOT NULL,center_id INTEGER NOT NULL,province TEXT,reason TEXT DEFAULT 'creation',status TEXT DEFAULT 'requested',requested_at TEXT NOT NULL,requested_by INTEGER,provincial_review_at TEXT,provincial_review_by INTEGER,national_approval_at TEXT,national_approval_by INTEGER,printed_at TEXT,printed_by INTEGER,delivered_at TEXT,delivered_by INTEGER,card_number TEXT UNIQUE,expires_at TEXT,notes TEXT,deleted_at TEXT,deleted_by INTEGER);
    """)
    for col, definition in {
        "icon": "TEXT DEFAULT '🏛️'", "color": "TEXT DEFAULT '#0f766e'", "manager_member_id": "INTEGER",
        "updated_by": "INTEGER", "settings_json": "TEXT DEFAULT '{}'"
    }.items():
        ensure_column("foundation_services", col, definition)
    cur.execute("INSERT OR IGNORE INTO foundation_services(code,name,description,scope,active,created_at,created_by) VALUES(?,?,?,?,?,?,?)",
                ('SANTE','Centre de santé','Réseau national des structures sanitaires de la Fondation','national',1,now(),1))
    health_service = cur.execute("SELECT id FROM foundation_services WHERE code='SANTE'").fetchone()
    if health_service:
        hs_id = health_service['id']
        for mcode,mname in [('PATIENTS','Patients'),('CONSULTATIONS','Consultations'),('LABORATOIRE','Laboratoire'),('PHARMACIE','Pharmacie'),('STOCKS','Stocks'),('FACTURATION','Facturation'),('HOSPITALISATION','Hospitalisation'),('RAPPORTS','Rapports')]:
            cur.execute("INSERT OR IGNORE INTO service_modules(service_id,code,name,description,active,created_at,created_by) VALUES(?,?,?,?,1,?,?)",(hs_id,mcode,mname,'Module du service Centre de santé',now(),1))
        for rname,rlevel in [('Responsable national','national'),('Coordonnateur provincial','province'),('Responsable de structure','structure')]:
            cur.execute("INSERT OR IGNORE INTO service_roles(service_id,name,description,level,active,created_at,created_by) VALUES(?,?,?,?,1,?,?)",(hs_id,rname,'Rôle générique du service',rlevel,now(),1))

    for col, definition in {
        'header_title': 'TEXT', 'header_subtitle': 'TEXT', 'header_address': 'TEXT', 'header_contact': 'TEXT',
        'logo_path': 'TEXT', 'footer_note': 'TEXT'
    }.items():
        ensure_column('health_centers', col, definition)
    for table in ['health_centers','health_roles','health_staff','patients','consultations','medical_form_templates',
                  'health_appointments','health_visits','health_lab_orders','health_products','health_stock_batches',
                  'health_admissions','health_invoices','health_social_support','health_equipment','health_suppliers',
                  'health_referrals','health_losses','health_monthly_reports','foundation_services','service_structures','service_roles',
                  'service_staff_assignments','service_modules']:
        ensure_column(table, 'deleted_at', 'TEXT')
        ensure_column(table, 'deleted_by', 'INTEGER')

    ensure_column('members', 'executive_number', 'TEXT')
    ensure_column('members', 'executive_position', 'TEXT')

    for role_name, level in [('Médecin directeur','centre'),('Docteur / Médecin','centre'),('Infirmier','centre'),('Sage-femme','centre'),('Laborantin','centre'),('Pharmacien','centre'),('Réceptionniste','centre'),('Caissier','centre'),('Gestionnaire de stock','centre'),('Coordonnateur provincial de santé','province'),('Administrateur national de santé','national')]:
        cur.execute("INSERT OR IGNORE INTO health_roles(name,description,level,active,created_at,created_by) VALUES(?,?,?,?,?,?)", (role_name,'Rôle sanitaire configurable',level,1,now(),1))
    default_patient_fields = json.dumps([
        {'key':'first_name','label':'Prénom','type':'text','required':True},
        {'key':'last_name','label':'Nom et postnom','type':'text','required':True},
        {'key':'gender','label':'Sexe','type':'select','required':False},
        {'key':'birth_date','label':'Date de naissance','type':'date','required':False},
        {'key':'phone','label':'Téléphone','type':'tel','required':False},
        {'key':'address','label':'Adresse','type':'textarea','required':False},
        {'key':'blood_group','label':'Groupe sanguin','type':'text','required':False},
        {'key':'allergies','label':'Allergies','type':'textarea','required':False},
        {'key':'medical_history','label':'Antécédents','type':'textarea','required':False}
    ], ensure_ascii=False)
    default_consult_fields = json.dumps([
        {'key':'reason','label':'Motif de consultation','type':'textarea','required':True},
        {'key':'complaints','label':'Plaintes principales','type':'textarea','required':False},
        {'key':'temperature','label':'Température','type':'text','required':False},
        {'key':'blood_pressure','label':'Tension artérielle','type':'text','required':False},
        {'key':'weight','label':'Poids','type':'text','required':False},
        {'key':'clinical_exam','label':'Examen clinique','type':'textarea','required':False},
        {'key':'diagnosis','label':'Diagnostic','type':'textarea','required':True},
        {'key':'treatment','label':'Traitement / prescription','type':'textarea','required':False}
    ], ensure_ascii=False)
    if not cur.execute("SELECT id FROM medical_form_templates WHERE form_type='patient' AND scope='national' LIMIT 1").fetchone():
        cur.execute("INSERT INTO medical_form_templates(name,form_type,scope,fields_json,active,version,created_at,created_by) VALUES(?,?,?,?,1,1,?,1)", ('Fiche nationale de renseignements du patient','patient','national',default_patient_fields,now()))
    if not cur.execute("SELECT id FROM medical_form_templates WHERE form_type='consultation' AND scope='national' LIMIT 1").fetchone():
        cur.execute("INSERT INTO medical_form_templates(name,form_type,scope,fields_json,active,version,created_at,created_by) VALUES(?,?,?,?,1,1,?,1)", ('Fiche nationale de consultation','consultation','national',default_consult_fields,now()))

    # ===== V36 : Circuit clinique, pharmacie, laboratoire, hospitalisation et gestion =====
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS health_appointments (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL,
        service_name TEXT, professional_id INTEGER, appointment_date TEXT NOT NULL, appointment_time TEXT, reason TEXT,
        status TEXT DEFAULT 'scheduled', reminder_status TEXT DEFAULT 'pending', created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_visits (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL,
        visit_date TEXT NOT NULL, arrival_time TEXT, service_name TEXT, priority TEXT DEFAULT 'normal', queue_number INTEGER,
        status TEXT DEFAULT 'waiting', started_at TEXT, completed_at TEXT, created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_triage (
        id INTEGER PRIMARY KEY AUTOINCREMENT, visit_id INTEGER NOT NULL UNIQUE, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL,
        temperature REAL, systolic INTEGER, diastolic INTEGER, heart_rate INTEGER, respiratory_rate INTEGER, oxygen_saturation REAL,
        weight REAL, height REAL, glucose REAL, pain_score INTEGER, consciousness TEXT, danger_signs TEXT, pregnancy_status TEXT,
        triage_level TEXT DEFAULT 'normal', notes TEXT, recorded_at TEXT NOT NULL, recorded_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_lab_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, consultation_id INTEGER, center_id INTEGER NOT NULL,
        test_name TEXT NOT NULL, sample_type TEXT, priority TEXT DEFAULT 'normal', status TEXT DEFAULT 'ordered',
        requested_at TEXT NOT NULL, requested_by INTEGER, collected_at TEXT, received_at TEXT, result_text TEXT, reference_range TEXT,
        validated_at TEXT, validated_by INTEGER, attachment_path TEXT
    );
    CREATE TABLE IF NOT EXISTS health_products (
        id INTEGER PRIMARY KEY AUTOINCREMENT, center_id INTEGER NOT NULL, code TEXT NOT NULL, name TEXT NOT NULL, category TEXT DEFAULT 'medicament',
        unit TEXT, minimum_stock REAL DEFAULT 0, sale_price REAL DEFAULT 0, active INTEGER DEFAULT 1, UNIQUE(center_id,code)
    );
    CREATE TABLE IF NOT EXISTS health_stock_batches (
        id INTEGER PRIMARY KEY AUTOINCREMENT, product_id INTEGER NOT NULL, batch_number TEXT, expiry_date TEXT, quantity REAL DEFAULT 0,
        purchase_price REAL DEFAULT 0, supplier_id INTEGER, received_at TEXT, created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_prescriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, consultation_id INTEGER, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL,
        prescribed_at TEXT NOT NULL, prescribed_by INTEGER, status TEXT DEFAULT 'pending', notes TEXT
    );
    CREATE TABLE IF NOT EXISTS health_prescription_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, prescription_id INTEGER NOT NULL, product_id INTEGER, medication_name TEXT NOT NULL, dosage TEXT,
        frequency TEXT, duration TEXT, quantity REAL DEFAULT 0, dispensed_quantity REAL DEFAULT 0, substitution_note TEXT
    );
    CREATE TABLE IF NOT EXISTS health_admissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL, ward TEXT, room TEXT, bed TEXT,
        admitted_at TEXT NOT NULL, diagnosis TEXT, responsible_staff_id INTEGER, status TEXT DEFAULT 'admitted', discharged_at TEXT,
        discharge_summary TEXT, outcome TEXT, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL, invoice_date TEXT NOT NULL,
        subtotal REAL DEFAULT 0, discount REAL DEFAULT 0, social_support REAL DEFAULT 0, total REAL DEFAULT 0, paid REAL DEFAULT 0,
        status TEXT DEFAULT 'unpaid', payment_method TEXT, notes TEXT, created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_invoice_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, invoice_id INTEGER NOT NULL, item_type TEXT, description TEXT NOT NULL, quantity REAL DEFAULT 1,
        unit_price REAL DEFAULT 0, total REAL DEFAULT 0, source_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_social_support (
        id INTEGER PRIMARY KEY AUTOINCREMENT, patient_id INTEGER NOT NULL, center_id INTEGER NOT NULL, category TEXT, reason TEXT, requested_amount REAL DEFAULT 0,
        approved_amount REAL DEFAULT 0, status TEXT DEFAULT 'pending', valid_until TEXT, approved_by INTEGER, created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, phone TEXT, email TEXT, address TEXT, province TEXT, active INTEGER DEFAULT 1, created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS health_equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT, center_id INTEGER NOT NULL, inventory_code TEXT NOT NULL UNIQUE, name TEXT NOT NULL, category TEXT,
        location TEXT, state TEXT DEFAULT 'operational', acquisition_date TEXT, warranty_end TEXT, next_maintenance TEXT, supplier_id INTEGER,
        notes TEXT, created_at TEXT NOT NULL, created_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_referrals (
        id INTEGER PRIMARY KEY AUTOINCREMENT, code TEXT NOT NULL UNIQUE, patient_id INTEGER NOT NULL, from_center_id INTEGER NOT NULL, to_center_id INTEGER,
        external_destination TEXT, urgency TEXT DEFAULT 'normal', reason TEXT NOT NULL, clinical_summary TEXT, treatment_given TEXT,
        status TEXT DEFAULT 'sent', sent_at TEXT NOT NULL, received_at TEXT, feedback TEXT, created_by INTEGER
    );
    """)


    # ===== V40 : Rapports sanitaires hiérarchiques et pertes =====
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS health_losses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        center_id INTEGER NOT NULL,
        loss_date TEXT NOT NULL,
        category TEXT NOT NULL,
        description TEXT NOT NULL,
        quantity REAL DEFAULT 0,
        amount REAL DEFAULT 0,
        currency TEXT DEFAULT 'CDF',
        reason TEXT,
        responsible_member_id INTEGER,
        status TEXT DEFAULT 'declared',
        created_at TEXT NOT NULL,
        created_by INTEGER,
        validated_at TEXT,
        validated_by INTEGER
    );
    CREATE TABLE IF NOT EXISTS health_monthly_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_month TEXT NOT NULL,
        scope_type TEXT NOT NULL DEFAULT 'center',
        center_id INTEGER,
        province TEXT,
        status TEXT DEFAULT 'draft',
        narrative_activities TEXT,
        difficulties TEXT,
        needs TEXT,
        recommendations TEXT,
        important_events TEXT,
        snapshot_json TEXT DEFAULT '{}',
        generated_at TEXT NOT NULL,
        generated_by INTEGER,
        updated_at TEXT,
        updated_by INTEGER,
        submitted_at TEXT,
        submitted_by INTEGER,
        validated_at TEXT,
        validated_by INTEGER,
        UNIQUE(report_month, scope_type, center_id, province)
    );
    CREATE INDEX IF NOT EXISTS idx_health_reports_month ON health_monthly_reports(report_month);
    CREATE INDEX IF NOT EXISTS idx_health_reports_scope ON health_monthly_reports(scope_type,province,center_id);
    CREATE INDEX IF NOT EXISTS idx_health_losses_center_date ON health_losses(center_id,loss_date);
    """)

    default_permissions = ['voir','ajouter','modifier','supprimer','imprimer','exporter','valider','parametres','imprimer_cartes','telecharger_cartes']
    for role_key, role_label in ROLE_LABELS.items():
        for perm in default_permissions:
            allowed = 1 if role_key == 'super_admin' else 0
            if role_key in NATIONAL_ROLES and perm != 'parametres':
                allowed = 1
            if perm in ['imprimer_cartes','telecharger_cartes']:
                allowed = 1 if role_key in NATIONAL_ROLES else 0
            if role_key in PROVINCIAL_ROLES and perm in ['voir','ajouter','modifier','imprimer','exporter']:
                allowed = 1
            if role_key == 'member' and perm == 'voir':
                allowed = 1
            exists = cur.execute("SELECT id FROM role_permissions WHERE role_key=? AND permission_key=? LIMIT 1", (role_key, perm)).fetchone()
            if not exists:
                cur.execute("INSERT INTO role_permissions(role_key,role_label,permission_key,allowed,updated_at) VALUES(?,?,?,?,?)", (role_key, role_label, perm, allowed, now()))

    permission_modules = list(PERMISSION_MODULES.keys())
    permission_actions = ['voir','ajouter','modifier','supprimer','valider','imprimer','exporter','parametrer']
    for role_key, role_label in ROLE_LABELS.items():
        for module in permission_modules:
            for action in permission_actions:
                pkey = f'{module}.{action}'
                allowed = 1 if role_key == 'super_admin' else 0
                if role_key in NATIONAL_ROLES and action in ['voir','ajouter','modifier','valider','imprimer','exporter']:
                    allowed = 1
                if role_key in PROVINCIAL_ROLES and action in ['voir','ajouter','modifier','imprimer','exporter'] and module not in ['parametres','roles']:
                    allowed = 1
                if role_key == 'member' and module == 'dashboard' and action == 'voir':
                    allowed = 1
                cur.execute("INSERT OR IGNORE INTO role_permissions(role_key,role_label,permission_key,allowed,updated_at) VALUES(?,?,?,?,?)", (role_key,role_label,pkey,allowed,now()))

    cur.execute("INSERT OR IGNORE INTO app_setup(id,completed) VALUES(1,0)")
    cur.execute("INSERT OR IGNORE INTO db_migrations(name, applied_at, details) VALUES(?,?,?)", ("2026_publication_plus", now(), "Sauvegarde/restauration, reçus PDF, import Excel, rôles personnalisables, guide rapide, trésorerie, alertes"))

    defaults = {
        "structure_name": "FONDATION BAKITANI",
        "structure_motto": "« FOBAK »",
        "structure_header": "République Démocratique du Congo",
        "structure_foundation": "Association sans but lucratif",
        "structure_legal": "Organisation associative à vocation sociale et communautaire",
        "secretariat_label": "Bureau National",
        "headquarters": "96, Av. Yauma, Quartier SAIO, Commune de Kasa-Vubu, Kinshasa - RDC",
        "contact_phones": "+243 81 45 70 392 ; +243 81 44 00 233",
        "payment_mpesa_number": "+243 81 45 70 392",
        "payment_mpesa_name": "Fondation Bakitani",
        "payment_vodacom_api_url": "",
        "payment_vodacom_api_key": "",
        "payment_vodacom_enabled": "0",
        "payment_confirmation_instructions": "Après paiement, saisissez la référence de transaction afin que le Bureau compétent puisse vérifier et valider la cotisation.",
        "history": "La Fondation Bakitani œuvre pour le développement intégral, la solidarité, l'organisation des membres et la promotion des initiatives communautaires.",
        "president_name": "Présidence nationale",
        "secretary_name": "Secrétaire du Bureau National",
        "logo_path": "img/fobak_logo_official_v25.png",
        "president_signature_path": "",
        "secretary_signature_path": "",
        "official_stamp_path": "",
        "stamp_application_mode": "validated",
        "facebook": "#",
        "youtube": "#",
        "whatsapp": "#",
        "instagram": "#",
        "tiktok": "#",
        "x_twitter": "#",
        "linkedin": "#",
        "payment_info": "Paiement possible par banque, Airtel Money, M-Pesa, Orange Money ou autre Mobile Money. Configurez ici les numéros et comptes officiels.",
        "card_notice": "Les autorités civiles, militaires que policières sont priées d'apporter leur assistance en cas de nécessité",
        "public_base_url": "http://127.0.0.1:5000",
        "mission": "Servir la population, promouvoir l'unité et soutenir le développement intégral.",
        "vision": "Une structure organisée, transparente et proche de ses membres dans toutes les provinces.",
        "values": "Unité, discipline, transparence, service, responsabilité.",
        "objectives": "Former les membres, organiser les activités, accompagner les communautés et défendre les intérêts de la structure.",
        "advantages": "Carte de membre, informations officielles, participation aux activités, suivi des cotisations et accès à l'espace membre.",
        "partners": "Partenaires institutionnels, sociaux et techniques à compléter.",
        "public_communiques": "Communiqué : les inscriptions et adhésions FOBAK sont ouvertes. Veuillez remplir votre fiche, joindre votre photo et suivre l'état de votre demande en ligne.",
        "dashboard_message": "Bienvenue dans le tableau de bord. Les informations affichées respectent les droits de votre rôle.",
        "footer_note": "Plateforme professionnelle de gestion des adhésions, cartes, cotisations, activités et services de la Fondation Bakitani.",
        "initiator": "Fondation Bakitani",
        "stability_center_text": "Notre centre de stabilité présente l'état de la plateforme, les mesures de sécurité, les canaux d'assistance et les engagements de disponibilité.",
        "privacy_policy": "Cette plateforme collecte uniquement les informations nécessaires à l'identification des membres, à la gestion des adhésions, des cartes, des cotisations, des activités et du support. Les accès sont protégés selon les rôles : national, provincial, local et membre. Les données ne doivent pas être partagées sans autorisation de la structure.",
        "terms_of_use": "L'utilisateur s'engage à fournir des informations exactes, à protéger ses identifiants, à utiliser la plateforme dans le cadre officiel de la structure et à ne pas falsifier les cartes, QR codes, reçus ou rapports.",
        "support_intro": "Utilisez ce formulaire pour signaler une erreur, un problème de connexion, une difficulté de paiement, un souci de carte membre ou toute autre demande d'assistance.",
        "demo_mode_enabled": "0",
        "mobile_app_name": "FOBAK Manager Pro",
        "windows_client_download_url": "/static/downloads/FOBAK_Client_Windows.zip",
        "windows_server_download_url": "",
        "android_download_url": "",
        "download_section_enabled": "1",
        "structure_address": "Adresse officielle modifiable de la Fondation / ASBL",
        "default_language": "fr",
        "global_search_placeholder": "Rechercher un membre, cotisation, activité, ticket...",
        "ai_help_intro": "Posez une question sur l’utilisation de l’application : membre, carte, cotisation, impression, support, mot de passe ou paramètres.",
        "current_exercise_label": str(datetime.now().year),
        "sync_remote_url": "",
        "sync_api_key": "",
        "sync_enabled": "0",
        "backup_retention_days": "30",
        "update_feed_url": "",
        "statute_intro": "Consultez les Statuts et le Règlement intérieur de la Fondation Bakitani. Les informations synthétiques restent modifiables par les personnes autorisées.",
    }
    for k, v in defaults.items():
        cur.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", (k, v))
    cur.execute("SELECT id FROM users WHERE email=?", ("admin@asbl.local",))
    if not cur.fetchone():
        cur.execute("INSERT INTO users(email, phone, password_hash, role, province, localite, active, created_at, force_password_change) VALUES(?,?,?,?,?,?,?,?,?)",
                    ("admin@asbl.local", "0990000000", generate_password_hash("admin123"), "super_admin", "National", "National", 1, now(), 1))
    else:
        cur.execute("UPDATE users SET force_password_change=1 WHERE email='admin@asbl.local' AND (password_changed_at IS NULL OR password_changed_at='')")
    # V45 : le compte administrateur technique n'est plus automatiquement un membre.
    admin_row = cur.execute("SELECT * FROM users WHERE email=?", ("admin@asbl.local",)).fetchone()
    if admin_row:
        cur.execute("UPDATE members SET deleted_at=COALESCE(deleted_at, ?), status='archived', updated_at=? WHERE user_id=? AND COALESCE(is_administrative,0)=1 AND (code='TEMP' OR adhesion_number='TEMP' OR lower(COALESCE(role_label,'')) LIKE 'administrateur%')", (now(), now(), admin_row['id']))

    cur.execute("SELECT id FROM activities LIMIT 1")
    if not cur.fetchone():
        cur.execute("INSERT INTO activities(title, subtitle, body, published_at, author_id) VALUES(?,?,?,?,?)",
                    ("Bienvenue dans notre ASBL", "Activité récente", "Publiez ici les activités, communiqués et messages importants de la structure.", now(), 1))
    official_path = "uploads/official_docs/statut_bakitani_original.pdf"
    if not cur.execute("SELECT id FROM official_documents WHERE document_type='statuts' AND active=1 AND deleted_at IS NULL LIMIT 1").fetchone():
        cur.execute("INSERT INTO official_documents(document_type,title,description,file_path,version_label,adoption_date,effective_date,public,active,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                    ("statuts", "Statuts et Règlement intérieur FOBAK", "Document officiel authentifié de la Fondation Bakitani.", official_path, "Mars 2026", "2026-03-28", "2026-03-28", 1, 1, now(), 1))
    statute_defaults = [
        ("denomination", "Dénomination et sigle", "Fondation BAKITANI, en sigle FOBAK ASBL.", 10),
        ("headquarters", "Siège social", "96, Avenue Yauma, Quartier SAIO, Commune de Kasa-Vubu, Kinshasa, République Démocratique du Congo.", 20),
        ("duration", "Durée", "La Fondation est créée pour une durée indéterminée.", 30),
        ("motto", "Devise", "Unité – Travail – Développement.", 40),
        ("vision", "Vision", "Former un noyau de citoyens de bonne qualité, caractérisés par le respect des lois, le civisme, l’intégrité et l’engagement pour le développement du pays.", 50),
        ("sectors", "Secteurs d’intervention", "Santé; Éducation; Agriculture; Environnement et développement durable; Affaires sociales et actions humanitaires; Éducation civique et éveil patriotique.", 60),
        ("member_categories", "Catégories de membres", "Membres fondateurs; membres effectifs; membres sympathisants; membres d’honneur.", 70),
        ("membership", "Adhésion", "L’adhésion est libre et volontaire, sous réserve du respect des conditions statutaires et du règlement intérieur.", 80),
        ("organs", "Organes de la Fondation", "Assemblée Générale; Comité Exécutif National; Collège des Conseillers; Commissariat aux Comptes.", 90),
        ("resources", "Ressources", "Cotisations des membres; dons et legs; subventions; produits des activités; autres contributions autorisées.", 100),
        ("discipline", "Discipline et sanctions", "Avertissement; blâme; suspension; exclusion, selon la gravité des faits et la procédure prévue.", 110),
        ("modification", "Modification et entrée en vigueur", "Les statuts sont modifiables conformément aux règles prévues et entrent en vigueur à compter de leur adoption.", 120),
    ]
    for skey, stitle, scontent, sorder in statute_defaults:
        cur.execute("INSERT OR IGNORE INTO statute_sections(section_key,title,content,sort_order,public,active,updated_at,updated_by) VALUES(?,?,?,?,1,1,?,1)", (skey, stitle, scontent, sorder, now()))
    con.commit()
    con.close()


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today():
    return datetime.now().strftime("%Y-%m-%d")


def get_settings():
    con = db()
    rows = con.execute("SELECT key,value FROM settings").fetchall()
    con.close()
    data = {r["key"]: r["value"] for r in rows}
    # Le logo choisi dans Paramètres reste prioritaire et se répercute partout.
    data.setdefault("logo_path", "img/fobak_logo_official_v25.png")
    if not data.get("logo_path"):
        data["logo_path"] = "img/fobak_logo_official_v25.png"
    data.setdefault("logo_dark_path", data["logo_path"])
    data.setdefault("logo_watermark_path", "img/fobak_logo_watermark_v25.png")
    data.setdefault("logo_print_path", "img/fobak_logo_print_v25.png")
    data.setdefault("favicon_path", "img/fobak_app_icon_v21.png")
    if not data.get("structure_name") or data.get("structure_name") == "NOM DE L'ASBL":
        data["structure_name"] = "FONDATION BAKITANI"
    if not data.get("structure_header") or data.get("structure_header","").startswith("Association sans but lucratif"):
        data["structure_header"] = "République Démocratique du Congo"
    if not data.get("structure_motto") or data.get("structure_motto") == "Unité • Service • Développement":
        data["structure_motto"] = "« FOBAK »"
    if not data.get("headquarters") or "adresse officielle" in data.get("headquarters", ""):
        data["headquarters"] = "96, Av. Yauma, Quartier SAIO, Commune de Kasa-Vubu, Kinshasa - RDC"
    if not data.get("contact_phones") or data.get("contact_phones") == "Tél. : +243 ...":
        data["contact_phones"] = "+243 81 45 70 392 ; +243 81 44 00 233"
    if not data.get("footer_note") or data.get("footer_note") == "Gestion professionnelle des adhésions, cartes, cotisations et activités de la structure.":
        data["footer_note"] = "Plateforme professionnelle de gestion des adhésions, cartes, cotisations, activités et services de la Fondation Bakitani."
    return data


def get_custom_fields(active_only=True):
    con = db()
    q = "SELECT * FROM adhesion_fields"
    if active_only:
        q += " WHERE active=1"
    q += " ORDER BY sort_order ASC, id ASC"
    rows = con.execute(q).fetchall()
    con.close()
    return rows


def collect_custom_field_values():
    values = {}
    for field in get_custom_fields(active_only=True):
        key = f"custom_{field['id']}"
        if field['field_type'] == 'checkbox':
            value = ", ".join(request.form.getlist(key))
        else:
            value = request.form.get(key, "").strip()
        values[str(field['id'])] = value
    values["member_type"] = request.form.get("member_type", "").strip()
    values["phone_country_code"] = request.form.get("phone_country_code", "+243").strip()
    return json.dumps(values, ensure_ascii=False)


def custom_values_dict(row):
    if not row:
        return {}
    try:
        raw = row['custom_fields'] or '{}'
        data = json.loads(raw)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def draw_actual_rdc_flag(c, x, y, w, h):
    """Dessine le vrai drapeau RDC fourni, sans débordement."""
    if os.path.exists(RDC_FLAG_ABS):
        try:
            c.drawImage(ImageReader(RDC_FLAG_ABS), x, y, w, h, preserveAspectRatio=True, anchor='c', mask='auto')
            c.setStrokeColor(colors.white)
            c.setLineWidth(0.35)
            c.rect(x, y, w, h, fill=0, stroke=1)
            return True
        except Exception:
            return False
    return False


@app.before_request
def enforce_idle_timeout_and_heartbeat():
    if request.endpoint == "static":
        return
    user_id = session.get("user_id")
    token = session.get("session_token")
    now_ts = int(datetime.now().timestamp())
    last_ts = int(session.get("last_activity_ts") or now_ts)
    if user_id and now_ts - last_ts >= 600:
        try:
            con = db()
            if token:
                con.execute("UPDATE active_sessions SET active=0, logout_at=? WHERE session_token=?", (now(), token))
                con.commit()
            con.close()
        except Exception:
            pass
        session.clear()
        if request.path.startswith("/api/") or request.is_json:
            return jsonify({"ok": False, "error": "session_expired"}), 401
        flash("Votre session a été fermée après 10 minutes d’inactivité.", "warning")
        return redirect(url_for("login"))
    if user_id:
        session["last_activity_ts"] = now_ts
        if token:
            try:
                con = db()
                con.execute("UPDATE active_sessions SET last_seen=?, active=1 WHERE session_token=? AND user_id=?", (now(), token, user_id))
                con.execute("UPDATE active_sessions SET active=0 WHERE active=1 AND datetime(last_seen) < datetime('now','-10 minutes')")
                con.commit(); con.close()
            except Exception:
                pass


def get_latest_login_sound_alert(user):
    if not user or user['role'] not in ('super_admin','president'):
        return None
    try:
        con=db()
        row=con.execute("""SELECT id,title,message,created_at FROM internal_notifications
                           WHERE user_id=? AND category='login_alert' AND sound_enabled=1 AND read_at IS NULL
                           ORDER BY id DESC LIMIT 1""", (user['id'],)).fetchone()
        con.close()
        return dict(row) if row else None
    except Exception:
        return None


@app.context_processor
def inject_globals():
    return dict(settings=get_settings(), role_labels=ROLE_LABELS, provinces=PROVINCES, territoires=TERRITOIRES,
                education_levels=EDUCATION_LEVELS, experience_levels=EXPERIENCE_LEVELS, study_checkboxes=STUDY_CHECKBOXES,
                marital_status_options=MARITAL_STATUS, adhesion_fields=get_custom_fields(), rdc_flag_rel=RDC_FLAG_REL, country_codes=COUNTRY_CODES,
                current_user=current_user(), current_year=datetime.now().year, demo_mode=demo_mode_enabled(),
                unread_notification_count=unread_notification_count(current_user()), lang_code=current_lang(),
                lang_options=AVAILABLE_LANGUAGES, _=translate_label, current_user_display_name=current_user_display_name(),
                maps_url="https://www.google.com/maps/search/?api=1&query=" + quote_plus(get_settings().get("headquarters", "Kinshasa RDC")),
                voice_first_name=(current_user_display_name().split()[0] if current_user_display_name() else ""),
                voice_role=ROLE_LABELS.get(session.get("role"), session.get("role", "utilisateur")),
                voice_welcome_pending=bool(session.pop("voice_welcome_pending", 0)),
                login_sound_alert=get_latest_login_sound_alert(current_user()),
                app_version=APP_VERSION, app_release_name=APP_RELEASE_NAME, module_allowed=module_allowed)


def current_user():
    if "user_id" not in session:
        return None
    con = db()
    user = con.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
    con.close()
    return user


def current_lang():
    lang = session.get("lang") or (current_user()["preferred_language"] if current_user() and "preferred_language" in current_user().keys() and current_user()["preferred_language"] else None) or get_settings().get("default_language", "fr")
    return lang if lang in AVAILABLE_LANGUAGES else "fr"


def translate_label(label, lang=None):
    lang = lang or current_lang()
    if lang == "fr":
        return label
    return UI_TRANSLATIONS.get(label, {}).get(lang, label)


def user_member_profile(user_id):
    if not user_id:
        return None
    con = db()
    row = con.execute("SELECT * FROM members WHERE user_id=? AND deleted_at IS NULL ORDER BY is_administrative DESC, id DESC LIMIT 1", (user_id,)).fetchone()
    con.close()
    return row


def current_user_display_name():
    user = current_user()
    if not user:
        return ""
    first = user["first_name"] if "first_name" in user.keys() and user["first_name"] else ""
    last = user["last_name"] if "last_name" in user.keys() and user["last_name"] else ""
    if first or last:
        return f"{first} {last}".strip()
    member = user_member_profile(user["id"])
    if member:
        return f"{member['first_name']} {member['last_name']}".strip()
    return user["email"] or user["phone"] or "Utilisateur"


def create_person_member_for_user(con, user_id, first_name, last_name, email, phone, role, province, localite, photo_path="", created_by=None):
    """Crée/complète le dossier membre d'un utilisateur, même administratif, afin qu'il ait profil + carte."""
    existing = con.execute("SELECT * FROM members WHERE user_id=? AND deleted_at IS NULL LIMIT 1", (user_id,)).fetchone()
    if existing:
        con.execute("""UPDATE members SET first_name=?, last_name=?, email=?, phone=?, province=?, localite=?, photo_path=COALESCE(NULLIF(?,''), photo_path), role_label=?, is_administrative=?, updated_at=? WHERE id=?""",
                    (first_name or existing["first_name"], last_name or existing["last_name"], email or existing["email"], phone or existing["phone"], province or existing["province"], localite or existing["localite"], photo_path, ROLE_LABELS.get(role, role), 1 if role != 'member' else 0, now(), existing["id"]))
        return existing["id"]
    joined = today()
    expires = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
    is_admin = 1 if role != "member" else 0
    con.execute("""INSERT INTO members(user_id, code, first_name, last_name, gender, email, phone, nationality, province, territory, commune, localite, physical_address, birth_date, birth_place, marital_status, profession, education, studies_done, experience, photo_path, custom_fields, adhesion_number, joined_at, expires_at, approved_by, created_by, status, updated_at, is_administrative, role_label)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (user_id, "TEMP", first_name or "Utilisateur", last_name or ROLE_LABELS.get(role, role), "", email or f"{phone}@fondation.local", phone or "", "Congolaise", province or "National", "", "", localite or "", "", "", "", "", ROLE_LABELS.get(role, role), "", "", "", photo_path or "", "{}", "TEMP", joined, expires, created_by, created_by, "active", now(), is_admin, ROLE_LABELS.get(role, role)))
    member_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    code = create_member_code(member_id, province or "NAT")
    adhesion_number = create_adhesion_number(member_id, joined)
    con.execute("UPDATE members SET code=?, adhesion_number=? WHERE id=?", (code, adhesion_number, member_id))
    return member_id


def demo_mode_enabled():
    try:
        return session.get("demo_mode") == 1 or get_settings().get("demo_mode_enabled") == "1"
    except Exception:
        return False


def unread_notification_count(user):
    if not user:
        return 0
    try:
        con = db()
        if user["role"] in NATIONAL_ROLES:
            count = con.execute("""
                SELECT COUNT(*) AS n FROM internal_notifications
                WHERE read_at IS NULL
                  AND (user_id=? OR user_id IS NULL)
                  AND (role IS NULL OR role=? OR role='all' OR province IS NOT NULL)
            """, (user["id"], user["role"])).fetchone()["n"]
        else:
            count = con.execute("""
                SELECT COUNT(*) AS n FROM internal_notifications
                WHERE read_at IS NULL
                  AND (user_id=? OR user_id IS NULL)
                  AND (role IS NULL OR role=? OR role='all')
                  AND (province IS NULL OR province='' OR province=?)
            """, (user["id"], user["role"], user["province"] or "")).fetchone()["n"]
        con.close()
        return count
    except Exception:
        return 0


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        user = current_user()
        if not user:
            flash("Veuillez vous connecter.", "warning")
            return redirect(url_for("login"))
        if should_force_password_change(user) and request.endpoint not in {"change_password", "logout", "static"}:
            flash("Pour sécuriser la production, changez d'abord votre mot de passe.", "warning")
            return redirect(url_for("change_password"))
        return fn(*args, **kwargs)
    return wrapper


def role_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user or user["role"] not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXT


def save_data_url_image(data_url, subfolder="photos"):
    """Enregistre une photo capturée par caméra (data URL) dans les uploads."""
    if not data_url or not data_url.startswith("data:image/"):
        return ""
    try:
        header, encoded = data_url.split(",", 1)
        ext = header.split("/")[1].split(";")[0].lower()
        if ext == "jpeg": ext = "jpg"
        if ext not in {"png", "jpg", "webp"}: ext = "jpg"
        raw = base64.b64decode(encoded)
        if len(raw) > app.config["MAX_CONTENT_LENGTH"]:
            return ""
        folder = os.path.join(UPLOAD_ROOT, subfolder)
        os.makedirs(folder, exist_ok=True)
        name = f"camera_{uuid.uuid4().hex}.{ext}"
        file_path = os.path.join(folder, name)
        with open(file_path, "wb") as f:
            f.write(raw)
        if subfolder == 'photos':
            _normalize_passport_image(file_path)
        return f"uploads/{subfolder}/{name}"
    except Exception:
        return ""

def _normalize_passport_image(abs_path, target_size=(600, 800)):
    """Normalise une photo au format passeport 3:4 sans déformer le visage."""
    try:
        with Image.open(abs_path) as im:
            im = im.convert('RGB')
            w, h = im.size
            if not w or not h:
                return
            target_ratio = 3 / 4
            current_ratio = w / h
            if current_ratio > target_ratio:
                crop_w = int(h * target_ratio)
                crop_h = h
            else:
                crop_w = w
                crop_h = int(w / target_ratio)
            left = max(0, (w - crop_w) // 2)
            top = max(0, (h - crop_h) // 2)
            im = im.crop((left, top, left + crop_w, top + crop_h))
            if target_size:
                im = im.resize(target_size, Image.Resampling.LANCZOS)
            im.save(abs_path, quality=92, optimize=True)
    except Exception:
        pass


def normalized_phone(prefix, number):
    prefix=(prefix or "+243").strip()
    number=''.join(ch for ch in (number or '') if ch.isdigit())
    number=number.lstrip('0')
    return f"{prefix}{number}" if number else ""

def save_upload(file, folder):
    if not file or not file.filename or not allowed_file(file.filename):
        return ""
    safe = secure_filename(file.filename)
    stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    name = f"{stamp}_{safe}"
    target_dir = os.path.join(UPLOAD_ROOT, folder)
    os.makedirs(target_dir, exist_ok=True)
    abs_path = os.path.join(target_dir, name)
    file.save(abs_path)
    if folder == 'photos':
        _normalize_passport_image(abs_path)
    return f"uploads/{folder}/{name}"


def save_logo_upload(file):
    """Enregistre le logo choisi et crée automatiquement ses variantes sans modifier sa forme."""
    if not file or not file.filename or not allowed_file(file.filename):
        return {}
    try:
        folder = os.path.join(UPLOAD_ROOT, "logos")
        os.makedirs(folder, exist_ok=True)
        img = Image.open(file.stream).convert("RGBA")
        # Pour les JPG sur fond blanc, supprimer automatiquement seulement le fond proche du blanc.
        # Le dessin et les proportions du logo restent inchangés.
        corners = [img.getpixel((0,0)), img.getpixel((img.width-1,0)), img.getpixel((0,img.height-1)), img.getpixel((img.width-1,img.height-1))]
        if all(px[0] > 235 and px[1] > 235 and px[2] > 235 for px in corners):
            cleaned = []
            for r, g, b, a in img.getdata():
                if r > 248 and g > 248 and b > 248:
                    cleaned.append((255,255,255,0))
                elif r > 235 and g > 235 and b > 235 and max(r,g,b)-min(r,g,b) < 12:
                    cleaned.append((r,g,b,max(0,min(255,int((255-r)*13)))))
                else:
                    cleaned.append((r,g,b,a))
            img.putdata(cleaned)
        bbox = img.getbbox()
        if bbox:
            img = img.crop(bbox)
        w, h = img.size
        if not w or not h:
            return {}
        # Accepter le logo de l'utilisateur tel qu'il est, sans imposer une forme ronde ou carrée.
        max_w = 2200
        if w > max_w:
            nh = max(1, round(h * max_w / w))
            img = img.resize((max_w, nh), Image.Resampling.LANCZOS)
        stamp = datetime.now().strftime('%Y%m%d%H%M%S%f')

        def variant(base, scale=1, color=1.0, contrast=1.0, sharpness=1.0, alpha_factor=1.0):
            out = base.resize((base.width * scale, base.height * scale), Image.Resampling.LANCZOS) if scale > 1 else base.copy()
            alpha = out.getchannel('A')
            rgb = out.convert('RGB')
            rgb = ImageEnhance.Color(rgb).enhance(color)
            rgb = ImageEnhance.Contrast(rgb).enhance(contrast)
            rgb = ImageEnhance.Sharpness(rgb).enhance(sharpness)
            result = rgb.convert('RGBA')
            result.putalpha(alpha.point(lambda a: int(a * alpha_factor)))
            return result

        main = variant(img, 2, 1.06, 1.07, 1.3)
        dark = variant(img, 2, 1.08, 1.08, 1.4)
        alpha = dark.getchannel('A')
        halo_mask = alpha.filter(ImageFilter.GaussianBlur(4))
        halo = Image.new('RGBA', dark.size, (255,255,255,0))
        halo.putalpha(halo_mask.point(lambda a: min(130, int(a * .65))))
        dark = Image.alpha_composite(halo, dark)
        print_logo = variant(img, 3, 1.04, 1.06, 1.5)
        watermark = variant(img, 2, 1.0, 1.0, 1.0, .16)

        paths = {}
        for key, picture in {
            'logo_path': main,
            'logo_dark_path': dark,
            'logo_print_path': print_logo,
            'logo_watermark_path': watermark,
        }.items():
            name = f"{key}_{stamp}.png"
            picture.save(os.path.join(folder, name), 'PNG', optimize=True)
            paths[key] = f"uploads/logos/{name}"
        return paths
    except Exception:
        return {}


def youtube_embed(url):
    if not url:
        return ""
    if "youtu.be/" in url:
        vid = url.split("youtu.be/")[-1].split("?")[0]
        return f"https://www.youtube.com/embed/{vid}"
    if "watch?v=" in url:
        vid = url.split("watch?v=")[-1].split("&")[0]
        return f"https://www.youtube.com/embed/{vid}"
    if "embed/" in url:
        return url
    return url


def is_national_user(user):
    return user and user["role"] in NATIONAL_ROLES


def is_treasury_editor(user):
    """Les trésoriers provinciaux gèrent uniquement leur province; le trésorier national gère uniquement les opérations nationales."""
    return bool(user and user["role"] in {"super_admin", "provincial_treasurer", "national_treasurer"})


def can_edit_treasury_entry(user, entry=None):
    if not user:
        return False
    if user["role"] == "super_admin":
        return True
    if user["role"] == "provincial_treasurer":
        return entry is None or (entry["province"] or "") == (user["province"] or "")
    if user["role"] == "national_treasurer":
        return entry is None or not (entry["province"] or "")
    return False


def role_permission_allowed(user, permission_key, default=False):
    if not user:
        return False
    if user["role"] == "super_admin":
        return True
    try:
        con = db()
        row = con.execute("SELECT allowed FROM role_permissions WHERE role_key=? AND permission_key=? ORDER BY id DESC LIMIT 1", (user["role"], permission_key)).fetchone()
        con.close()
        return bool(row["allowed"]) if row else bool(default)
    except Exception:
        return bool(default)


def can_output_member_card(user):
    return bool(user and user["role"] in NATIONAL_ROLES and role_permission_allowed(user, "imprimer_cartes", True)) or role_permission_allowed(user, "imprimer_cartes", False)


def member_scope_query(user, base="WHERE deleted_at IS NULL"):
    params = []
    where = base
    if user and user["role"] in PROVINCIAL_ROLES:
        where += " AND province=?"
        params.append(user["province"])
    if user and user["role"] == "local_admin":
        where += " AND province=? AND localite=?"
        params.extend([user["province"], user["localite"]])
    return where, params


def can_manage_member(user, member):
    if not user or not member:
        return False
    if user["role"] in NATIONAL_ROLES:
        return True
    if user["role"] in PROVINCIAL_ROLES:
        return (member["province"] or "") == (user["province"] or "")
    if user["role"] == "local_admin":
        return (member["province"] or "") == (user["province"] or "") and (member["localite"] or "") == (user["localite"] or "")
    return False


def log_action(user_id, action, target_type="", target_id=None, details="", link="", status="success"):
    """Journalise qui a fait quoi, où et depuis quel appareil."""
    try:
        con = db()
        ip = request.remote_addr if has_request_context() else ""
        agent = (request.user_agent.string or "")[:500] if has_request_context() else ""
        route = (request.path or "")[:255] if has_request_context() else ""
        con.execute("""INSERT INTO audit_logs(user_id,action,target_type,target_id,details,created_at,ip_address,user_agent,route,status)
                       VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (user_id, action, target_type, target_id, details, now(), ip, agent, route, status))
        if user_id:
            message = action + (f" — {details}" if details else "")
            con.execute("""INSERT INTO internal_notifications(user_id,title,message,link,category,sound_enabled,created_at)
                           VALUES(?,?,?,?,?,?,?)""",
                        (user_id, "Action enregistrée", message[:500], link or "", "audit", 0, now()))
        con.commit(); con.close()
    except Exception:
        pass

def notify_national_login(user):
    """Avertit par notification sonore les administrateurs nationaux et le Président national."""
    try:
        con = db()
        full_name = (f"{user['first_name'] or ''} {user['last_name'] or ''}".strip()
                     or user['email'] or user['phone'] or f"Utilisateur #{user['id']}")
        role_name = ROLE_LABELS.get(user['role'], user['role'])
        message = f"{full_name} ({role_name}) vient de se connecter à {now()}."
        recipients = con.execute("""SELECT id FROM users WHERE active=1 AND deleted_at IS NULL
                                  AND role IN ('super_admin','president')""").fetchall()
        for recipient in recipients:
            if recipient['id'] == user['id']:
                continue
            con.execute("""INSERT INTO internal_notifications(user_id,title,message,link,category,sound_enabled,created_at)
                           VALUES(?,?,?,?,?,?,?)""",
                        (recipient['id'], "Connexion détectée", message, url_for('audit_center'), 'login_alert', 1, now()))
        con.commit(); con.close()
    except Exception:
        pass


def parse_datetime(value):
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value)[:19], fmt)
        except Exception:
            continue
    return None


def is_account_locked(user):
    locked_until = parse_datetime(user["locked_until"] if "locked_until" in user.keys() else "")
    return bool(locked_until and locked_until > datetime.now())


def should_force_password_change(user):
    try:
        return bool(user and int(user["force_password_change"] or 0) == 1)
    except Exception:
        return False


def validate_new_password(new_password, confirm_password=""):
    weak = {"admin123", "123456", "12345678", "password", "motdepasse", "000000", "111111"}
    if not new_password or len(new_password) < 8:
        return "Le nouveau mot de passe doit contenir au moins 8 caractères."
    if confirm_password and new_password != confirm_password:
        return "Les deux mots de passe ne sont pas identiques."
    if new_password.lower() in weak:
        return "Ce mot de passe est trop faible pour la production. Choisissez un mot de passe plus sûr."
    return ""


def create_backup_archive(reason="manuel"):
    backup_dir = os.path.join(BASE_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_path = os.path.join(backup_dir, f"sauvegarde_asbl_{stamp}.zip")
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as zf:
        if os.path.exists(DB_PATH):
            zf.write(DB_PATH, "asbl.db")
        # Sauvegarder les fichiers utiles ajoutés par l'administration.
        upload_base = os.path.join(STATIC_DIR, "uploads")
        if os.path.isdir(upload_base):
            for root, _, files in os.walk(upload_base):
                for name in files:
                    full = os.path.join(root, name)
                    arc = os.path.relpath(full, BASE_DIR)
                    if "backups" not in arc.replace("\\", "/"):
                        zf.write(full, arc)
        for extra in [".env", ".env.example", "server_url.txt"]:
            full = os.path.join(BASE_DIR, extra)
            if os.path.exists(full):
                zf.write(full, extra)
        zf.writestr("INFO_SAUVEGARDE.txt", f"Sauvegarde ASBL créée le {now()}\nMotif : {reason}\nNe partagez pas cette archive publiquement : elle peut contenir des données personnelles.\n")
    return archive_path


def production_health_status():
    con = db()
    tables = [r["name"] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()]
    migrations = con.execute("SELECT * FROM db_migrations ORDER BY applied_at DESC, id DESC LIMIT 8").fetchall()
    audit_count = con.execute("SELECT COUNT(*) AS n FROM audit_logs").fetchone()["n"]
    ticket_open = con.execute("SELECT COUNT(*) AS n FROM support_tickets WHERE status NOT IN ('closed','résolu','resolu')").fetchone()["n"]
    users_force = con.execute("SELECT COUNT(*) AS n FROM users WHERE deleted_at IS NULL AND COALESCE(force_password_change,0)=1").fetchone()["n"]
    con.close()
    backup_dir = os.path.join(BASE_DIR, "backups")
    backups = []
    if os.path.isdir(backup_dir):
        for name in sorted(os.listdir(backup_dir), reverse=True):
            if name.endswith(".zip"):
                full = os.path.join(backup_dir, name)
                backups.append({"name": name, "size": os.path.getsize(full), "created": datetime.fromtimestamp(os.path.getmtime(full)).strftime("%Y-%m-%d %H:%M:%S")})
    checks = [
        {"label": "Base de données", "ok": os.path.exists(DB_PATH), "detail": "Base SQLite présente" if os.path.exists(DB_PATH) else "Base absente"},
        {"label": "Migrations automatiques", "ok": "db_migrations" in tables, "detail": "Tables contrôlées au lancement"},
        {"label": "Journal d'audit", "ok": "audit_logs" in tables, "detail": f"{audit_count} action(s) enregistrée(s)"},
        {"label": "Support", "ok": "support_tickets" in tables, "detail": f"{ticket_open} ticket(s) ouvert(s)"},
        {"label": "Sauvegardes", "ok": len(backups) > 0, "detail": f"{len(backups)} sauvegarde(s) disponible(s)"},
        {"label": "Changement mot de passe", "ok": users_force == 0, "detail": f"{users_force} compte(s) doivent changer le mot de passe"},
    ]
    return {"tables": tables, "migrations": migrations, "backups": backups[:10], "checks": checks}


def create_member_code(member_id, province=""):
    """Code national lisible : ASBL-FOBAK-PROVINCE-NUMERO-NAT.

    Le numéro est unique dans la base et le code provincial facilite le classement
    des cartes par province sans dépendre de l'année d'impression.
    """
    raw = (province or "NAT").strip().upper()
    normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    words = [w for w in re.split(r"[^A-Z0-9]+", normalized) if w]
    if not words:
        province_code = "NAT"
    elif len(words) == 1:
        province_code = words[0][:3]
    else:
        province_code = "".join(w[0] for w in words)[:4]
    province_code = province_code or "NAT"
    return f"ASBL-FOBAK-{province_code}-{int(member_id):05d}-NAT"


def create_adhesion_number(member_id, joined_date=None):
    year = (joined_date or today())[:4]
    return f"{member_id:03d}/{year}"


def checked_studies(value):
    raw = value or ""
    return {x.strip() for x in raw.replace(";", ",").split(",") if x.strip()}


def draw_rdc_flag(c, x, y, w, h):
    """Dessine le vrai drapeau RDC fourni, avec secours vectoriel."""
    if draw_actual_rdc_flag(c, x, y, w, h):
        return
    c.setFillColor(colors.HexColor("#00A3E0"))
    c.rect(x, y, w, h, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#F7D618"))
    c.setLineWidth(max(2.2, h * 0.22))
    c.line(x + w * 0.03, y + h * 0.10, x + w * 0.97, y + h * 0.90)
    c.setStrokeColor(colors.HexColor("#CE1021"))
    c.setLineWidth(max(1.1, h * 0.10))
    c.line(x + w * 0.05, y + h * 0.12, x + w * 0.95, y + h * 0.88)
    c.setFillColor(colors.HexColor("#F7D618"))
    c.circle(x + w * 0.21, y + h * 0.74, h * 0.10, fill=1, stroke=0)
    c.setStrokeColor(colors.white)
    c.setLineWidth(0.4)
    c.rect(x, y, w, h, fill=0, stroke=1)


def wrap_text(text, max_chars=90):
    text = str(text or "")
    lines, line = [], ""
    for word in text.split():
        if len(line) + len(word) + 1 <= max_chars:
            line = (line + " " + word).strip()
        else:
            if line:
                lines.append(line)
            line = word
    if line:
        lines.append(line)
    return lines or [""]


def verification_url(code):
    settings = get_settings()
    base = (settings.get("public_base_url") or "http://127.0.0.1:5000").rstrip("/")
    return f"{base}/verification/{code}"

def scoped_members_rows(user, include_deleted=False):
    base = "WHERE 1=1" if include_deleted else "WHERE deleted_at IS NULL"
    where, params = member_scope_query(user, base)
    con = db()
    rows = con.execute(f"SELECT * FROM members {where} ORDER BY province,last_name,first_name", params).fetchall()
    con.close()
    return rows


def payment_filter_sql(user, args):
    where = "WHERE m.deleted_at IS NULL"
    params = []
    if user and user["role"] == "provincial_president":
        where += " AND m.province=? AND (m.created_by=? OR m.approved_by=?)"; params.extend([user["province"], user["id"], user["id"]])
    elif user and user["role"] in PROVINCIAL_ROLES:
        where += " AND m.province=?"; params.append(user["province"])
    if user and user["role"] == "local_admin":
        where += " AND m.province=? AND m.localite=?"; params.extend([user["province"], user["localite"]])
    province = args.get("province", "")
    if province and user and user["role"] in NATIONAL_ROLES:
        where += " AND m.province=?"; params.append(province)
    member_id = args.get("member_id", "")
    if member_id:
        where += " AND p.member_id=?"; params.append(member_id)
    for col in ["status", "method", "currency", "contribution_type"]:
        val = args.get(col, "")
        if val:
            where += f" AND p.{col}=?"; params.append(val)
    year = args.get("year", "")
    if year:
        where += " AND strftime('%Y', COALESCE(p.paid_at,p.created_at))=?"; params.append(year)
    month = args.get("month", "")
    if month:
        where += " AND strftime('%m', COALESCE(p.paid_at,p.created_at))=?"; params.append(month.zfill(2))
    date_start = args.get("date_start", "")
    if date_start:
        where += " AND date(COALESCE(p.paid_at,p.created_at))>=date(?)"; params.append(date_start)
    date_end = args.get("date_end", "")
    if date_end:
        where += " AND date(COALESCE(p.paid_at,p.created_at))<=date(?)"; params.append(date_end)
    return where, params


def get_payment_rows(user, args):
    con = db()
    where, params = payment_filter_sql(user, args)
    rows = con.execute(f"""
        SELECT p.*, m.code, m.first_name, m.last_name, m.phone, m.province, m.territory, m.commune
        FROM payments p
        LEFT JOIN members m ON m.id=p.member_id
        {where}
        ORDER BY COALESCE(p.paid_at,p.created_at) DESC, p.id DESC
    """, params).fetchall()
    totals = con.execute(f"""
        SELECT p.currency, p.status, SUM(p.amount) AS total, COUNT(*) AS n
        FROM payments p
        LEFT JOIN members m ON m.id=p.member_id
        {where}
        GROUP BY p.currency, p.status
    """, params).fetchall()
    con.close()
    return rows, totals


def upcoming_anniversaries(user, days=30):
    rows = scoped_members_rows(user)
    today_dt = datetime.now().date()
    results = []
    for r in rows:
        raw = (r["joined_at"] or "")[:10]
        try:
            d = datetime.strptime(raw, "%Y-%m-%d").date()
        except Exception:
            continue
        try:
            next_date = d.replace(year=today_dt.year)
        except ValueError:
            next_date = d.replace(year=today_dt.year, day=28)
        if next_date < today_dt:
            try:
                next_date = d.replace(year=today_dt.year + 1)
            except ValueError:
                next_date = d.replace(year=today_dt.year + 1, day=28)
        delta = (next_date - today_dt).days
        if 0 <= delta <= days:
            years = next_date.year - d.year
            results.append({"member": r, "next_date": next_date.strftime("%Y-%m-%d"), "days": delta, "years": years})
    results.sort(key=lambda x: x["days"])
    return results


def can_approve_publication(user):
    return bool(user and user["role"] in NATIONAL_ROLES)


def support_scope_sql(user, base="WHERE 1=1"):
    where = base
    params = []
    if user and user["role"] in PROVINCIAL_ROLES:
        where += " AND province=?"
        params.append(user["province"] or "")
    if user and user["role"] == "local_admin":
        where += " AND province=? AND localite=?"
        params.extend([user["province"] or "", user["localite"] or ""])
    return where, params


def create_internal_notification(title, message, link="", user_id=None, role=None, province=None):
    try:
        con = db()
        con.execute("""INSERT INTO internal_notifications(user_id,role,province,title,message,link,created_at)
                       VALUES(?,?,?,?,?,?,?)""", (user_id, role, province, title, message, link, now()))
        con.commit(); con.close()
    except Exception:
        pass


def ticket_code(ticket_id):
    return f"SUP-{datetime.now().year}-{ticket_id:05d}"


def demo_numbers():
    return {
        "total_members": 1248, "active_members": 1187, "pending_count": 23,
        "contribution_total": 12450, "contribution_currency": "USD",
        "activities_count": 28
    }


def generate_demo_lists():
    Row = lambda **kw: kw
    activities = [
        Row(title="Assemblée générale ordinaire", province="National", published_at="2026-05-20", status="approved"),
        Row(title="Campagne de sensibilisation", province="Kinshasa", published_at="2026-05-18", status="approved"),
        Row(title="Visite terrain", province="Nord-Kivu", published_at="2026-05-12", status="approved"),
    ]
    notifications = [
        Row(message="5 nouvelles demandes d'adhésion à vérifier", province="National", target_scope="all", created_at=now()),
        Row(message="Paiement en attente de confirmation", province="Kinshasa", target_scope="province", created_at=now()),
        Row(message="Nouvelle publication disponible", province="National", target_scope="members", created_at=now()),
    ]
    pending = [Row(last_name="KABONGO", first_name="Marie", province="Kinshasa", created_at=now()), Row(last_name="MUTOMBO", first_name="Jean", province="Mongala", created_at=now())]
    anniversaries = [{"member": Row(last_name="MULUMBA", first_name="Grâce"), "next_date": today(), "days": 0, "years": 1}]
    return activities, notifications, pending, anniversaries


def _card_font(size, bold=False):
    candidates = []
    if os.name == "nt":
        candidates += [r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf"]
    candidates += [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default()


def _open_contained(path, box, transparent=True):
    if not path or not os.path.exists(path):
        return None
    try:
        im = Image.open(path).convert("RGBA")
        im.thumbnail(box, Image.Resampling.LANCZOS)
        bg = (255,255,255,0) if transparent else (255,255,255,255)
        holder = Image.new("RGBA", box, bg)
        holder.alpha_composite(im, ((box[0]-im.width)//2, (box[1]-im.height)//2))
        return holder
    except Exception:
        return None


def _fit_text(draw, text, font, max_width, max_chars=60):
    value = str(text or "-").strip()
    if len(value) > max_chars:
        value = value[:max_chars-1] + "…"
    while len(value) > 2 and draw.textbbox((0,0), value, font=font)[2] > max_width:
        value = value[:-2] + "…"
    return value


def _card_status(member):
    try:
        expiry = datetime.strptime((member["expires_at"] or "")[:10], "%Y-%m-%d").date()
        days = (expiry - datetime.now().date()).days
        if days < 0:
            return "EXPIRÉE", days
        if days <= 30:
            return "BIENTÔT EXPIRÉE", days
        return "ACTIVE", days
    except Exception:
        return "À VÉRIFIER", 0


def _save_flat_psd(image, output_path):
    # PSD RGB aplati lisible dans Photoshop; les SVG fournis dans le ZIP restent entièrement modifiables.
    rgb = image.convert("RGB")
    w, h = rgb.size
    header = b"8BPS" + struct.pack(">H", 1) + b"\0"*6 + struct.pack(">HIIHH", 3, h, w, 8, 3)
    body = struct.pack(">I",0) + struct.pack(">I",0) + struct.pack(">I",0) + struct.pack(">H",0)
    with open(output_path, "wb") as f:
        f.write(header + body + b"".join(rgb.getchannel(i).tobytes() for i in range(3)))


def _card_svg(member, side, settings):
    esc = lambda v: html.escape(str(v or ""))
    logo_href = "/static/" + settings.get("logo_path", "")
    flag_href = "/static/" + RDC_FLAG_REL
    status, _ = _card_status(member)
    card_code = create_member_code(member["id"], member["province"] or "NAT")
    common = f'''<defs><linearGradient id="g" x2="1" y2="1"><stop stop-color="#f8fcff"/><stop offset=".55" stop-color="#e7f4fb"/><stop offset="1" stop-color="#c7e8f7"/></linearGradient><linearGradient id="band" x2="1"><stop stop-color="#073b66"/><stop offset="1" stop-color="#0879b9"/></linearGradient></defs>
<rect width="1011" height="638" rx="28" fill="url(#g)"/><rect x="8" y="8" width="995" height="622" rx="25" fill="none" stroke="#0b5f96" stroke-width="3"/><path d="M0 500 Q240 390 505 510 T1011 445 V638 H0Z" fill="#19a5d8" opacity=".18"/><path d="M0 545 Q280 445 540 548 T1011 500" fill="none" stroke="#f4b51e" stroke-width="5" opacity=".65"/>
<image href="{esc(logo_href)}" x="55" y="35" width="390" height="105" preserveAspectRatio="xMinYMid meet"/>
<image href="{esc(logo_href)}" x="235" y="180" width="530" height="300" opacity=".07" preserveAspectRatio="xMidYMid meet"/>
<image href="{esc(flag_href)}" x="745" y="35" width="92" height="58" preserveAspectRatio="xMidYMid meet"/>
<text x="850" y="57" font-family="Arial" font-weight="700" font-size="14" fill="#0b5594">RÉPUBLIQUE</text><text x="850" y="77" font-family="Arial" font-weight="700" font-size="14" fill="#0b5594">DÉMOCRATIQUE DU CONGO</text>
<rect x="35" y="150" width="600" height="65" rx="8" fill="url(#band)"/><text x="78" y="195" font-family="Arial" font-weight="700" font-size="35" fill="white">CARTE DE MEMBRE</text>'''
    if side == "front":
        fields = [("Nom",member["last_name"]),("Prénom",member["first_name"]),("Sexe",member["gender"]),("Nationalité",member["nationality"]),("Fonction",member["role_label"] if "role_label" in member.keys() else member["profession"]),("Adresse",member["physical_address"]),("N° Carte",card_code)]
        rows = "".join(f'<text x="75" y="{245+i*42}" font-family="Arial" font-weight="700" font-size="21" fill="#0b5594">{esc(k)}</text><text x="240" y="{245+i*42}" font-family="Arial" font-weight="600" font-size="20" fill="#172536">: {esc(v)}</text>' for i,(k,v) in enumerate(fields))
        body = f'''{rows}<rect x="735" y="160" width="215" height="260" rx="20" fill="white" fill-opacity=".82" stroke="#0872bb" stroke-width="4"/><text x="810" y="300" font-family="Arial" font-weight="700" font-size="20" fill="#6c7680">PHOTO</text>
<rect x="520" y="462" width="185" height="118" rx="16" fill="white" fill-opacity=".72" stroke="#0872bb" stroke-width="3"/><text x="565" y="493" font-family="Arial" font-weight="700" font-size="18" fill="#0872bb">SIGNATURE</text>
<rect x="748" y="440" width="207" height="140" rx="16" fill="white" fill-opacity=".72" stroke="#0872bb" stroke-width="3"/><text x="808" y="473" font-family="Arial" font-weight="700" font-size="18" fill="#0872bb">CACHET</text>'''
    else:
        body = f'''<text x="75" y="270" font-family="Arial" font-weight="700" font-size="23" fill="#0b5594">N° CARTE :</text><text x="300" y="270" font-family="Arial" font-weight="700" font-size="23">{esc(card_code)}</text>
<text x="75" y="323" font-family="Arial" font-weight="700" font-size="23" fill="#0b5594">DATE D’ÉMISSION :</text><text x="345" y="323" font-family="Arial" font-size="23">{esc((member['joined_at'] or '')[:10])}</text>
<text x="75" y="376" font-family="Arial" font-weight="700" font-size="23" fill="#0b5594">DATE D’EXPIRATION :</text><text x="365" y="376" font-family="Arial" font-size="23">{esc((member['expires_at'] or '')[:10])}</text>
<text x="75" y="429" font-family="Arial" font-weight="700" font-size="22" fill="#0b5594">STATUT :</text><text x="225" y="429" font-family="Arial" font-weight="700" font-size="22" fill="#c9252d">{esc(status)}</text>
<rect x="70" y="455" width="515" height="95" rx="15" fill="white" fill-opacity=".78" stroke="#0872bb" stroke-width="2"/><text x="100" y="490" font-family="Arial" font-size="17" fill="#163c61">Cette carte est personnelle et incessible.</text><text x="100" y="520" font-family="Arial" font-size="17" fill="#163c61">Elle atteste de la qualité de membre FOBAK ASBL.</text>
<rect x="705" y="160" width="230" height="245" rx="18" fill="white" fill-opacity=".82" stroke="#0872bb" stroke-width="3"/><text x="775" y="195" font-family="Arial" font-weight="700" font-size="18" fill="#0872bb">QR CODE</text>
<rect x="35" y="570" width="941" height="50" rx="18" fill="#075a9e"/><text x="65" y="600" font-family="Arial" font-size="15" fill="white">{esc(settings.get('headquarters',''))} | {esc(settings.get('contact_phones',''))}</text>'''
    return '<svg xmlns="http://www.w3.org/2000/svg" width="1011" height="638" viewBox="0 0 1011 638">' + common + body + '</svg>'


def generate_member_card_assets(member):
    settings = get_settings()
    cards_dir = os.path.join(UPLOAD_ROOT, "cards", member["code"])
    os.makedirs(cards_dir, exist_ok=True)
    W,H=1011,638; blue=(9,82,137); dark=(18,42,65); gold=(242,181,30)
    logo_abs=os.path.join(BASE_DIR,"static",settings.get("logo_path", "")); photo_abs=os.path.join(BASE_DIR,"static",member["photo_path"]) if member["photo_path"] else ""
    status,_=_card_status(member)
    card_code = create_member_code(member["id"], member["province"] or "NAT")
    def base_card():
        im=Image.new("RGB",(W,H),(248,252,255)); d=ImageDraw.Draw(im)
        for y in range(H):
            t=y/H
            d.line((0,y,W,y),fill=(int(250-34*t),int(253-18*t),int(255-5*t)))
        d.rounded_rectangle((8,8,W-8,H-8),radius=26,outline=(9,95,150),width=3)
        d.arc((-140,270,W+160,600),190,345,fill=(52,181,226),width=3)
        d.arc((-120,310,W+190,635),190,345,fill=gold,width=4)
        logo=_open_contained(logo_abs,(390,105)); wm_abs=os.path.join(BASE_DIR,"static",settings.get("logo_watermark_path", settings.get("logo_path", ""))); wm=_open_contained(wm_abs,(530,300))
        if logo: im.paste(logo,(45,30),logo)
        if wm:
            wm.putalpha(wm.getchannel("A").point(lambda a:int(a*.07))); im.paste(wm,(235,185),wm)
        flag=_open_contained(RDC_FLAG_ABS,(92,58),False)
        if flag: im.paste(flag.convert("RGB"),(742,38))
        d.text((846,48),"RÉPUBLIQUE",font=_card_font(13,True),fill=blue)
        d.text((846,70),"DÉMOCRATIQUE DU CONGO",font=_card_font(13,True),fill=blue)
        d.rounded_rectangle((30,150,635,215),radius=8,fill=(6,61,105))
        d.rectangle((30,205,635,215),fill=(17,145,194))
        d.text((75,164),"CARTE DE MEMBRE",font=_card_font(36,True),fill="white")
        return im,d
    front,d=base_card()
    labels=[("Nom",member["last_name"]),("Prénom",member["first_name"]),("Sexe",member["gender"]),("Nationalité",member["nationality"]),("Fonction",member["role_label"] if "role_label" in member.keys() and member["role_label"] else member["profession"]),("Adresse",member["physical_address"]),("N° Carte",card_code)]
    y=232
    for lab,val in labels:
        label_font=_card_font(18,True)
        value_font=_card_font(17,True)
        d.text((60,y),lab,font=label_font,fill=blue)
        max_width=390 if lab not in ("Adresse","Fonction") else 350
        value=_fit_text(d,val,value_font,max_width,75)
        d.text((215,y),": "+value,font=value_font,fill=dark)
        y+=38
    d.rounded_rectangle((735,160,950,408),radius=20,fill="white",outline=(8,114,187),width=4)
    photo=_open_contained(photo_abs,(195,226),False)
    if photo: front.paste(photo.convert("RGB"),(745,171))
    else: d.text((810,275),"PHOTO",font=_card_font(22,True),fill=(100,110,120))
    # Les cadres signature et cachet restent uniquement au recto et dans une zone de sécurité basse.
    d.rounded_rectangle((515,450,705,570),radius=16,fill="white",outline=(8,114,187),width=3)
    d.text((560,463),"SIGNATURE",font=_card_font(16,True),fill=(8,114,187))
    d.rounded_rectangle((735,430,955,570),radius=16,fill="white",outline=(8,114,187),width=3)
    d.text((808,443),"CACHET",font=_card_font(16,True),fill=(8,114,187))
    sig_rel=settings.get("president_signature_path",""); stamp_rel=settings.get("official_stamp_path","")
    sig=_open_contained(os.path.join(BASE_DIR,"static",sig_rel) if sig_rel else "",(165,78))
    stamp=_open_contained(os.path.join(BASE_DIR,"static",stamp_rel) if stamp_rel else "",(185,105))
    if sig: front.paste(sig,(528,482),sig)
    if stamp: front.paste(stamp,(752,458),stamp)
    back,d=base_card()
    rows=[("N° CARTE",card_code),("DATE D’ÉMISSION",(member["joined_at"] or "")[:10]),("DATE D’EXPIRATION",(member["expires_at"] or "")[:10]),("STATUT",status)]
    y=245
    for lab,val in rows:
        d.text((75,y),lab,font=_card_font(22,True),fill=blue); d.text((350 if "DATE" in lab else 265,y),": "+val,font=_card_font(22,True if lab in ("N° CARTE","STATUT") else False),fill=(201,37,45) if lab=="STATUT" else dark); y+=53
    d.rounded_rectangle((70,455,585,548),radius=16,fill="white",outline=(8,114,187),width=2); d.text((100,470),"Cette carte est personnelle et incessible.",font=_card_font(16,True),fill=(22,60,97)); d.text((100,500),"Elle atteste de la qualité de membre FOBAK ASBL.",font=_card_font(16),fill=(22,60,97))
    qrimg=qrcode.make(verification_url(member["code"])).convert("RGB").resize((180,180)); d.rounded_rectangle((705,160,935,405),radius=18,fill="white",outline=(8,114,187),width=3); back.paste(qrimg,(730,195)); d.text((775,168),"QR CODE",font=_card_font(17,True),fill=(8,114,187))
    # Le verso est réservé aux informations de validité, au QR code et à la vérification.
    d.rounded_rectangle((30,570,980,625),radius=18,fill=(5,90,158)); footer=_fit_text(d,settings.get("headquarters","")+" | "+settings.get("contact_phones",""),_card_font(14),900,130); d.text((55,588),footer,font=_card_font(14),fill="white")
    paths={}
    for side,img in (("recto",front),("verso",back)):
        png=os.path.join(cards_dir,side+".png"); img.save(png,dpi=(300,300)); paths[side+"_png"]=png
        psd=os.path.join(cards_dir,side+".psd"); _save_flat_psd(img,psd); paths[side+"_psd"]=psd
        svg=os.path.join(cards_dir,side+".svg"); Path(svg).write_text(_card_svg(member,"front" if side=="recto" else "back",settings),encoding="utf-8"); paths[side+"_svg"]=svg
    pdf=os.path.join(cards_dir,"carte_recto_verso.pdf"); c=canvas.Canvas(pdf,pagesize=(85.6*mm,54*mm))
    for side in ("recto","verso"): c.drawImage(paths[side+"_png"],0,0,85.6*mm,54*mm,preserveAspectRatio=False,mask='auto'); c.showPage()
    c.save(); paths["pdf"]=pdf
    zip_path=os.path.join(cards_dir,"sources_photoshop.zip")
    with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as z:
        for k,pth in paths.items():
            if k!="pdf": z.write(pth,os.path.basename(pth))
        z.writestr("LISEZ_MOI.txt","PSD : compatible Photoshop, image aplatie. SVG : source entièrement modifiable dans Photoshop/Illustrator/Inkscape. Recto et verso séparés pour impression PVC.\n")
    paths["zip"]=zip_path
    return paths


def generate_member_card_pdf(member):
    return generate_member_card_assets(member)["pdf"]

def draw_logo_or_placeholder(c, rel_path, x, y, size, label="LOGO"):
    abs_path = os.path.join(BASE_DIR, "static", rel_path) if rel_path else ""
    if rel_path and os.path.exists(abs_path):
        try:
            c.drawImage(ImageReader(abs_path), x, y, size, size, preserveAspectRatio=True, anchor='c', mask='auto')
            return
        except Exception:
            pass
    c.setStrokeColor(colors.HexColor("#777777"))
    c.roundRect(x, y, size, size, 2*mm, fill=0, stroke=1)
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica-Bold", 7)
    c.drawCentredString(x + size/2, y + size/2 - 2, label)




def draw_logo_rect(c, rel_path, x, y, width, height, label="LOGO"):
    abs_path = os.path.join(BASE_DIR, "static", rel_path) if rel_path else ""
    if rel_path and os.path.exists(abs_path):
        try:
            c.drawImage(ImageReader(abs_path), x, y, width, height, preserveAspectRatio=True, anchor='c', mask='auto')
            return
        except Exception:
            pass
    c.setStrokeColor(colors.HexColor("#777777"))
    c.roundRect(x, y, width, height, 2*mm, fill=0, stroke=1)
    c.setFillColor(colors.HexColor("#555555"))
    c.setFont("Helvetica-Bold", 7)
    c.drawCentredString(x + width/2, y + height/2 - 2, label)

def draw_field_line(c, label, value, x_label, x_line, y, line_w, font_size=10):
    c.setFillColor(colors.black)
    c.setFont("Helvetica", font_size)
    c.drawString(x_label, y, label)
    c.drawString(x_line - 5, y, ":")
    c.setDash(1, 2)
    c.line(x_line + 3, y - 1, x_line + line_w, y - 1)
    c.setDash()
    if value:
        c.setFont("Helvetica-Oblique", font_size)
        c.drawString(x_line + 5, y + 1, str(value)[:62])


def draw_checkbox(c, x, y, checked=False):
    c.setStrokeColor(colors.black)
    c.rect(x, y, 5*mm, 5*mm, fill=0, stroke=1)
    if checked:
        c.setLineWidth(1.2)
        c.line(x + 1*mm, y + 2.5*mm, x + 2.2*mm, y + 1*mm)
        c.line(x + 2.2*mm, y + 1*mm, x + 4.5*mm, y + 4.2*mm)
        c.setLineWidth(1)


def generate_adhesion_form_pdf(member=None, blank=False):
    """Génère une fiche d'adhésion A4 avec en-tête et pied de page harmonisés selon le modèle FOBAK."""
    settings = get_settings()
    cards_dir = os.path.join(UPLOAD_ROOT, "cards")
    os.makedirs(cards_dir, exist_ok=True)
    if blank or not member:
        pdf_path = os.path.join(cards_dir, "fiche_adhesion_vierge.pdf")
    else:
        pdf_path = os.path.join(cards_dir, f"fiche_adhesion_{member['code']}.pdf")

    c = canvas.Canvas(pdf_path, pagesize=A4)
    page_w, page_h = A4
    margin_x = 20 * mm
    content_left = margin_x + 6 * mm
    content_right = page_w - margin_x - 6 * mm

    c.setFillColor(colors.white)
    c.rect(0, 0, page_w, page_h, fill=1, stroke=0)

    def val(key):
        if not member:
            return ""
        try:
            return member[key] or ""
        except Exception:
            return ""

    # Filigrane central très discret
    wm_rel = settings.get("logo_watermark_path", settings.get("logo_path", ""))
    wm_abs = os.path.join(BASE_DIR, "static", wm_rel) if wm_rel else ""
    if wm_abs and os.path.exists(wm_abs):
        try:
            c.saveState()
            c.setFillAlpha(0.035)
            c.drawImage(ImageReader(wm_abs), page_w/2 - 45*mm, page_h/2 - 45*mm, 90*mm, 90*mm, preserveAspectRatio=True, mask='auto')
            c.restoreState()
        except Exception:
            pass

    # ===== En-tête modèle utilisateur =====
    header_top = page_h - 24 * mm
    logo_x, logo_y = content_left, header_top - 18 * mm
    draw_logo_rect(c, settings.get("logo_print_path", settings.get("logo_path", "")), logo_x, logo_y, 40 * mm, 16 * mm, "LOGO")

    flag_w, flag_h = 32 * mm, 18 * mm
    flag_x = page_w - margin_x - flag_w - 34 * mm
    flag_y = header_top - 15 * mm
    draw_rdc_flag(c, flag_x, flag_y, flag_w, flag_h)

    photo_w, photo_h = 23 * mm, 32 * mm
    photo_x = page_w - margin_x - photo_w
    photo_y = header_top - 36 * mm
    c.setStrokeColor(colors.HexColor("#777777"))
    c.setLineWidth(0.7)
    c.roundRect(photo_x, photo_y, photo_w, photo_h, 4 * mm, fill=0, stroke=1)
    if member and member["photo_path"]:
        photo_abs = os.path.join(BASE_DIR, "static", member["photo_path"])
        if os.path.exists(photo_abs):
            try:
                c.drawImage(ImageReader(photo_abs), photo_x + 1*mm, photo_y + 1*mm, photo_w - 2*mm, photo_h - 2*mm, preserveAspectRatio=True, anchor='c', mask='auto')
            except Exception:
                pass
    else:
        c.setFillColor(colors.HexColor("#666666"))
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(photo_x + photo_w/2, photo_y + photo_h/2 - 2, "PHOTO")

    center_x = page_w / 2
    c.setFillColor(colors.HexColor("#2f2f2f"))
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(center_x, header_top - 24*mm, settings.get("structure_name", "FONDATION BAKITANI").upper()[:45])
    c.setFont("Helvetica", 9)
    c.drawCentredString(center_x, header_top - 36*mm, settings.get("structure_legal", "Association sans but lucratif")[:75])
    c.drawCentredString(center_x, header_top - 43*mm, settings.get("structure_foundation", "Organisation associative à vocation sociale et communautaire")[:85])
    c.setFont("Helvetica-Bold", 14)
    c.drawString(content_left + 52*mm, header_top - 56*mm, settings.get("secretariat_label", "Bureau National"))

    adhesion_number = ""
    if member:
        adhesion_number = member["adhesion_number"] or create_adhesion_number(member["id"], member["joined_at"])
    c.setFillColor(colors.HexColor("#202020"))
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(center_x, header_top - 79*mm, f"FICHE D’ADHÉSION N° {adhesion_number or '____/____'}")
    c.setStrokeColor(colors.HexColor("#404040"))
    c.setLineWidth(0.8)
    c.line(center_x - 48*mm, header_top - 82*mm, center_x + 48*mm, header_top - 82*mm)

    # ===== Corps =====
    y = header_top - 96 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(content_left, y, "I.  IDENTITÉ")
    c.line(content_left, y-1, content_left + 30*mm, y-1)
    y -= 12 * mm

    birth_text = ""
    if member:
        birth_text = f"{val('birth_place')}  {val('birth_date')}".strip()
    fields = [
        ("-    NOM ET POST NOM", val("last_name")),
        ("-    Prénom", val("first_name")),
        ("-    Lieu et date de naissance", birth_text),
        ("-    Etat-civil", val("marital_status")),
        ("-    Province", val("province")),
        ("-    Territoire", val("territory")),
        ("-    Adresse physique", val("physical_address") or val("localite")),
        ("-    Téléphone", val("phone")),
        ("-    Email", val("email")),
    ]
    for label, value in fields:
        draw_field_line(c, label, value, content_left + 5*mm, content_left + 58*mm, y, 92*mm, 9.4)
        y -= 7 * mm

    y -= 3 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(content_left, y, "II.  ÉTUDES FAITES")
    c.line(content_left, y-1, content_left + 39*mm, y-1)
    y -= 11 * mm
    selected = checked_studies(val("studies_done") or val("education"))
    study_x2 = content_left + 48 * mm
    row_gap = 8 * mm
    for idx, label in enumerate(STUDY_CHECKBOXES):
        col = 0 if idx < 3 else 1
        row = idx if idx < 3 else idx - 3
        base_x = content_left + (col * study_x2)
        yy = y - row * row_gap
        c.setFont("Helvetica", 10)
        c.drawString(base_x + 3*mm, yy + 1*mm, "-")
        draw_checkbox(c, base_x + 10*mm, yy - 1*mm, checked=(label in selected))
        c.drawString(base_x + 19*mm, yy + 0.5*mm, label)
    y -= 3 * row_gap + 6 * mm

    custom_values = custom_values_dict(member) if member else {}
    current_section = None
    for f in get_custom_fields(active_only=True):
        if y < 95 * mm:
            break
        section = f['section'] or 'Autres informations'
        if section != current_section:
            y -= 2 * mm
            c.setFont("Helvetica-Bold", 10.5)
            c.drawString(content_left, y, section.upper())
            c.line(content_left, y-1, content_left + 55*mm, y-1)
            y -= 8 * mm
            current_section = section
        value = custom_values.get(str(f['id']), '')
        draw_field_line(c, f"-    {f['label']}", value, content_left + 5*mm, content_left + 58*mm, y, 92*mm, 9)
        y -= 7 * mm

    y -= 3 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(content_left, y, "III.  DÉCLARATION DE L’ADHÉRENT")
    c.line(content_left, y-1, content_left + 58*mm, y-1)
    y -= 11 * mm
    member_type = custom_values.get("member_type", "") if member else ""
    c.setFont("Helvetica", 9.4)
    c.drawString(content_left, y, f"Qualité : {member_type or '☐ Fondateur   ☐ Effectif   ☐ Honneur   ☐ Sympathisant'}")
    y -= 7 * mm
    declaration = (
        "Par la présente, j’adhère librement à la FONDATION BAKITANI (FOBAK) et je m’engage à respecter ses "
        "statuts, son règlement intérieur et les décisions de ses instances dirigeantes."
    )
    for line in wrap_text(declaration, 110)[:3]:
        c.drawString(content_left, y, line)
        y -= 5 * mm

    y -= 10 * mm
    date_text = today() if member else "____/____/______"
    c.drawString(page_w - 102*mm, y, f"Fait à __________________, le {date_text}")

    # ===== Pied de page modèle utilisateur =====
    visa_y = 52 * mm
    sign_y = 52 * mm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(content_left + 22*mm, visa_y, "VISA")
    c.drawString(page_w - margin_x - 58*mm, sign_y, "Signature de l’adhérant")

    # éventuel sceau à gauche sans débordement
    stamp_rel = settings.get("official_stamp_path", "")
    stamp_mode = settings.get("stamp_application_mode", "validated")
    should_stamp = bool(member) and (stamp_mode == "registered" or val("status") in ("active", "accepted", "validated", ""))
    stamp_abs = os.path.join(BASE_DIR, "static", stamp_rel) if stamp_rel else ""
    if should_stamp and stamp_abs and os.path.exists(stamp_abs):
        try:
            c.saveState()
            c.setFillAlpha(0.92)
            c.drawImage(ImageReader(stamp_abs), content_left + 10*mm, 35*mm, 28*mm, 15*mm, preserveAspectRatio=True, anchor='c', mask='auto')
            c.restoreState()
        except Exception:
            pass

    line_y = 32 * mm
    c.setStrokeColor(colors.HexColor("#555555"))
    c.setLineWidth(0.8)
    c.line(content_left + 4*mm, line_y, page_w - margin_x - 4*mm, line_y)

    c.setFillColor(colors.HexColor("#222222"))
    c.setFont("Helvetica-Bold", 10)
    phones = settings.get("contact_phones", "+243 81 45 70 392 ; 81 44 00 233")
    c.drawCentredString(page_w/2, 22*mm, f"Contacts : {phones}"[:120])
    c.drawCentredString(page_w/2, 16*mm, settings.get("headquarters", "96, Av. Yauma, Quartier SAIO Commune de KASA VUBU/KINSHASA - RDC")[:120])

    c.showPage()
    c.save()
    return pdf_path


@app.route("/centre-stabilite")
def stability_center():
    return render_template("stability_center.html")


@app.route("/politique-confidentialite")
def privacy_policy():
    return render_template("legal_page.html", page_title="Politique de confidentialité", content=get_settings().get("privacy_policy", ""))


@app.route("/conditions-utilisation")
def terms_of_use():
    return render_template("legal_page.html", page_title="Conditions d’utilisation", content=get_settings().get("terms_of_use", ""))


@app.route("/mode-demo")
def public_demo_mode():
    session["demo_mode"] = 1
    nums = demo_numbers()
    activities, notifications, pending, anniversaries = generate_demo_lists()
    return render_template("demo_dashboard.html", nums=nums, activities=activities, notifications=notifications, pending=pending, anniversaries=anniversaries)


@app.route("/quitter-mode-demo")
def quit_demo_mode():
    session.pop("demo_mode", None)
    flash("Mode démonstration désactivé.", "info")
    return redirect(url_for("index"))


@app.route("/signaler-probleme", methods=["GET", "POST"])
def report_problem():
    user = current_user()
    member = None
    if user:
        con = db()
        member = con.execute("SELECT * FROM members WHERE user_id=? AND deleted_at IS NULL", (user["id"],)).fetchone()
        con.close()
    if request.method == "POST":
        attachment_path = save_upload(request.files.get("attachment"), "support")
        full_name = request.form.get("full_name", "").strip() or (f"{member['last_name']} {member['first_name']}" if member else "")
        email = request.form.get("email", "").strip() or (user["email"] if user else "")
        phone = request.form.get("phone", "").strip() or (user["phone"] if user else "")
        province = request.form.get("province", "").strip() or ((member["province"] if member else user["province"]) if user else "")
        localite = request.form.get("localite", "").strip() or ((member["localite"] if member else user["localite"]) if user else "")
        title = request.form.get("title", "").strip()
        message = request.form.get("message", "").strip()
        if not title or not message:
            flash("Le titre et la description du problème sont obligatoires.", "danger")
            return redirect(url_for("report_problem"))
        con = db()
        cur = con.cursor()
        cur.execute("""INSERT INTO support_tickets(user_id,member_id,full_name,email,phone,province,localite,category,title,message,attachment_path,status,priority,created_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (user["id"] if user else None, member["id"] if member else None, full_name, email, phone, province, localite, request.form.get("category", "Général"), title, message, attachment_path, "new", request.form.get("priority", "normal"), now(), now()))
        tid = cur.lastrowid
        code = ticket_code(tid)
        cur.execute("UPDATE support_tickets SET tracking_code=? WHERE id=?", (code, tid))
        con.commit(); con.close()
        create_internal_notification("Nouveau ticket support", f"{code} — {title}", url_for("support_ticket_detail", ticket_id=tid), role="all" if not province else None, province=province or None)
        flash(f"Votre problème a été signalé. Numéro de ticket : {code}", "success")
        return redirect(url_for("report_problem"))
    return render_template("report_problem.html", member=member)


@app.route("/telecharger-client-windows")
def download_windows_client():
    package = os.path.join(BASE_DIR, "static", "downloads", "FOBAK_Client_Windows.zip")
    if not os.path.exists(package):
        abort(404)
    return send_file(package, as_attachment=True, download_name="FOBAK_Client_Windows.zip")


@app.route("/hors-connexion")
def offline_page():
    return render_template("offline.html")


@app.route("/")
def index():
    con = db()
    activities = con.execute("SELECT * FROM activities WHERE COALESCE(status,'approved')='approved' ORDER BY published_at DESC LIMIT 6").fetchall()
    videos = con.execute("SELECT * FROM videos ORDER BY created_at DESC LIMIT 4").fetchall()
    carousel = con.execute("SELECT * FROM carousel_images WHERE active=1 ORDER BY created_at DESC LIMIT 8").fetchall()
    docs = con.execute("SELECT * FROM documents WHERE public=1 ORDER BY created_at DESC LIMIT 4").fetchall()
    projects = con.execute("SELECT * FROM projects ORDER BY created_at DESC LIMIT 4").fetchall()
    con.close()
    return render_template("index.html", activities=activities, videos=videos, carousel=carousel, docs=docs, projects=projects, youtube_embed=youtube_embed)


@app.route("/fiche-vierge")
def fiche_vierge():
    pdf_path = generate_adhesion_form_pdf(blank=True)
    return send_file(pdf_path, as_attachment=True, download_name="fiche_adhesion_vierge.pdf")


def start_user_session(user, method="password"):
    session.clear()
    session["user_id"] = user["id"]
    session["role"] = user["role"]
    session["session_token"] = secrets.token_urlsafe(24)
    session["voice_welcome_pending"] = 1
    session["last_activity_ts"] = int(datetime.now().timestamp())
    con = db()
    con.execute("UPDATE active_sessions SET active=0, logout_at=? WHERE user_id=? AND active=1", (now(), user["id"]))
    con.execute("INSERT INTO active_sessions(session_token,user_id,login_at,last_seen,ip_address,user_agent,active) VALUES(?,?,?,?,?,?,1)", (session["session_token"], user["id"], now(), now(), request.remote_addr or "", (request.user_agent.string or "")[:500]))
    con.execute("UPDATE users SET last_login=?, failed_login_count=0, locked_until=NULL WHERE id=?", (now(), user["id"]))
    con.commit(); con.close()
    log_action(user["id"], f"Connexion réussie ({method})", "user", user["id"])
    notify_national_login(user)


def webauthn_rp_id():
    return os.environ.get("WEBAUTHN_RP_ID", (request.host or "app.fondationbakitani.org").split(":")[0])


def webauthn_origin():
    return os.environ.get("WEBAUTHN_ORIGIN", request.host_url.rstrip("/"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ident = request.form.get("identifier", "").strip()
        password = request.form.get("password", "")
        con = db()
        user = con.execute("SELECT * FROM users WHERE active=1 AND deleted_at IS NULL AND (email=? OR phone=?)", (ident, ident)).fetchone()
        if user and is_account_locked(user):
            con.close()
            flash("Compte temporairement verrouillé après plusieurs tentatives. Réessayez plus tard ou contactez l'administration.", "danger")
            return render_template("login.html")
        if user and check_password_hash(user["password_hash"], password):
            con.close()
            start_user_session(user, "mot de passe")
            if should_force_password_change(user):
                flash("Connexion réussie. Changez votre mot de passe initial avant de continuer.", "warning")
                return redirect(url_for("change_password"))
            flash("Connexion réussie.", "success")
            return redirect(url_for("dashboard"))
        if user:
            attempts = int(user["failed_login_count"] or 0) + 1
            locked_until = None
            if attempts >= 5:
                locked_until = (datetime.now() + timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S")
                attempts = 0
            con.execute("UPDATE users SET failed_login_count=?, locked_until=? WHERE id=?", (attempts, locked_until, user["id"]))
            con.commit()
        con.close()
        flash("Identifiants incorrects.", "danger")
    return render_template("login.html")


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    user = current_user()
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        message = validate_new_password(new_password, confirm_password)
        if message:
            flash(message, "danger")
            return redirect(url_for("change_password"))
        if not check_password_hash(user["password_hash"], current_password):
            flash("Mot de passe actuel incorrect.", "danger")
            return redirect(url_for("change_password"))
        con = db()
        con.execute("""UPDATE users SET password_hash=?, force_password_change=0, failed_login_count=0, locked_until=NULL, password_changed_at=? WHERE id=?""",
                    (generate_password_hash(new_password), now(), user["id"]))
        con.commit(); con.close()
        log_action(user["id"], "Changement mot de passe production", "user", user["id"])
        flash("Mot de passe changé. Votre compte est prêt pour la production.", "success")
        return redirect(url_for("dashboard"))
    return render_template("change_password.html")


@app.route("/logout")
def logout():
    token = session.get("session_token")
    if token:
        con = db()
        con.execute("UPDATE active_sessions SET active=0, logout_at=?, last_seen=? WHERE session_token=?", (now(), now(), token))
        con.commit(); con.close()
    session.clear()
    flash("Vous êtes déconnecté.", "info")
    return redirect(url_for("index"))


@app.route("/security", methods=["GET", "POST"])
@login_required
def security_settings():
    user = current_user()
    con = db()
    if request.method == "POST":
        action = request.form.get("action", "set_pin")
        if action == "set_pin":
            pin = request.form.get("pin", "").strip()
            confirm = request.form.get("pin_confirm", "").strip()
            if not check_password_hash(user["password_hash"], request.form.get("current_password", "")):
                flash("Mot de passe actuel incorrect.", "danger")
            elif not pin.isdigit() or len(pin) not in (4, 6):
                flash("Le code PIN doit contenir exactement 4 ou 6 chiffres.", "danger")
            elif pin != confirm:
                flash("La confirmation du PIN ne correspond pas.", "danger")
            else:
                con.execute("UPDATE users SET pin_hash=?, pin_length=?, pin_failed_count=0, pin_locked_until=NULL WHERE id=?", (generate_password_hash(pin), len(pin), user["id"]))
                con.commit(); flash("Code PIN enregistré.", "success")
        elif action == "remove_pin":
            if check_password_hash(user["password_hash"], request.form.get("current_password", "")):
                con.execute("UPDATE users SET pin_hash=NULL, pin_length=NULL, pin_failed_count=0, pin_locked_until=NULL WHERE id=?", (user["id"],))
                con.commit(); flash("Code PIN supprimé.", "success")
            else:
                flash("Mot de passe actuel incorrect.", "danger")
    credentials = con.execute("SELECT * FROM passkey_credentials WHERE user_id=? AND active=1 ORDER BY created_at DESC", (user["id"],)).fetchall()
    refreshed = con.execute("SELECT * FROM users WHERE id=?", (user["id"],)).fetchone()
    con.close()
    return render_template("security_settings.html", security_user=refreshed, credentials=credentials, webauthn_available=WEBAUTHN_AVAILABLE)


@app.post("/login/pin")
def login_pin():
    ident = request.form.get("identifier", "").strip()
    pin = request.form.get("pin", "").strip()
    con = db()
    user = con.execute("SELECT * FROM users WHERE active=1 AND deleted_at IS NULL AND (email=? OR phone=?)", (ident, ident)).fetchone()
    if not user or not user["pin_hash"]:
        con.close(); flash("Aucun code PIN n’est configuré pour ce compte.", "danger"); return redirect(url_for("login"))
    if user["pin_locked_until"] and user["pin_locked_until"] > now():
        con.close(); flash("PIN temporairement verrouillé. Utilisez le mot de passe.", "danger"); return redirect(url_for("login"))
    if pin.isdigit() and check_password_hash(user["pin_hash"], pin):
        con.execute("UPDATE users SET pin_failed_count=0, pin_locked_until=NULL WHERE id=?", (user["id"],)); con.commit(); con.close()
        start_user_session(user, "code PIN")
        return redirect(url_for("dashboard"))
    attempts = int(user["pin_failed_count"] or 0) + 1
    locked_until = None
    if attempts >= 5:
        attempts = 0
        locked_until = (datetime.now() + timedelta(minutes=15)).strftime("%Y-%m-%d %H:%M:%S")
    con.execute("UPDATE users SET pin_failed_count=?, pin_locked_until=? WHERE id=?", (attempts, locked_until, user["id"]))
    con.commit(); con.close(); flash("Code PIN incorrect.", "danger"); return redirect(url_for("login"))


@app.post("/api/passkeys/register/options")
@login_required
def passkey_register_options():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"ok": False, "error": "WebAuthn indisponible"}), 503
    user = current_user()
    con = db(); rows = con.execute("SELECT credential_id FROM passkey_credentials WHERE user_id=? AND active=1", (user["id"],)).fetchall(); con.close()
    options = generate_registration_options(
        rp_id=webauthn_rp_id(), rp_name="Fondation Bakitani",
        user_id=str(user["id"]).encode(), user_name=user["email"] or user["phone"] or str(user["id"]),
        user_display_name=user["email"] or user["phone"] or f"Utilisateur {user['id']}",
        exclude_credentials=[PublicKeyCredentialDescriptor(id=base64url_to_bytes(r["credential_id"])) for r in rows],
        authenticator_selection=AuthenticatorSelectionCriteria(authenticator_attachment=AuthenticatorAttachment.PLATFORM, resident_key=ResidentKeyRequirement.PREFERRED, user_verification=UserVerificationRequirement.REQUIRED),
    )
    payload = json.loads(options_to_json(options))
    session["webauthn_registration_challenge"] = payload["challenge"]
    return jsonify(payload)


@app.post("/api/passkeys/register/verify")
@login_required
def passkey_register_verify():
    challenge = session.pop("webauthn_registration_challenge", None)
    if not WEBAUTHN_AVAILABLE or not challenge:
        return jsonify({"ok": False, "error": "Challenge expiré"}), 400
    credential = request.get_json(force=True)
    verification = verify_registration_response(credential=credential, expected_challenge=base64url_to_bytes(challenge), expected_rp_id=webauthn_rp_id(), expected_origin=webauthn_origin(), require_user_verification=True)
    public_key = base64.urlsafe_b64encode(verification.credential_public_key).decode().rstrip("=")
    con = db()
    con.execute("INSERT OR REPLACE INTO passkey_credentials(user_id,credential_id,public_key,sign_count,transports,label,created_at,active) VALUES(?,?,?,?,?,?,?,1)", (session["user_id"], credential["id"], public_key, verification.sign_count, json.dumps(credential.get("response", {}).get("transports", [])), "Empreinte / visage", now()))
    con.execute("UPDATE users SET passkey_enabled=1 WHERE id=?", (session["user_id"],))
    con.commit(); con.close()
    return jsonify({"ok": True})


@app.post("/api/passkeys/login/options")
def passkey_login_options():
    if not WEBAUTHN_AVAILABLE:
        return jsonify({"ok": False, "error": "WebAuthn indisponible"}), 503
    ident = (request.get_json(silent=True) or {}).get("identifier", "").strip()
    con = db(); user = con.execute("SELECT * FROM users WHERE active=1 AND deleted_at IS NULL AND (email=? OR phone=?)", (ident, ident)).fetchone()
    if not user:
        con.close(); return jsonify({"ok": False, "error": "Compte introuvable"}), 404
    rows = con.execute("SELECT credential_id FROM passkey_credentials WHERE user_id=? AND active=1", (user["id"],)).fetchall(); con.close()
    if not rows:
        return jsonify({"ok": False, "error": "Aucune biométrie enregistrée"}), 404
    options = generate_authentication_options(rp_id=webauthn_rp_id(), allow_credentials=[PublicKeyCredentialDescriptor(id=base64url_to_bytes(r["credential_id"])) for r in rows], user_verification=UserVerificationRequirement.REQUIRED)
    payload = json.loads(options_to_json(options)); session["webauthn_auth_challenge"] = payload["challenge"]; session["webauthn_auth_user_id"] = user["id"]
    return jsonify(payload)


@app.post("/api/passkeys/login/verify")
def passkey_login_verify():
    challenge = session.pop("webauthn_auth_challenge", None); user_id = session.pop("webauthn_auth_user_id", None)
    credential = request.get_json(force=True); cid = credential.get("id")
    if not WEBAUTHN_AVAILABLE or not challenge or not user_id or not cid:
        return jsonify({"ok": False, "error": "Challenge expiré"}), 400
    con = db(); row = con.execute("SELECT * FROM passkey_credentials WHERE user_id=? AND credential_id=? AND active=1", (user_id, cid)).fetchone(); user = con.execute("SELECT * FROM users WHERE id=? AND active=1", (user_id,)).fetchone()
    if not row or not user:
        con.close(); return jsonify({"ok": False, "error": "Clé inconnue"}), 404
    verification = verify_authentication_response(credential=credential, expected_challenge=base64url_to_bytes(challenge), expected_rp_id=webauthn_rp_id(), expected_origin=webauthn_origin(), credential_public_key=base64url_to_bytes(row["public_key"]), credential_current_sign_count=row["sign_count"], require_user_verification=True)
    con.execute("UPDATE passkey_credentials SET sign_count=?, last_used_at=? WHERE id=?", (verification.new_sign_count, now(), row["id"])); con.commit(); con.close()
    start_user_session(user, "empreinte / visage")
    return jsonify({"ok": True, "redirect": url_for("dashboard")})


@app.post("/security/passkeys/<int:credential_id>/delete")
@login_required
def delete_passkey(credential_id):
    user = current_user(); con = db(); con.execute("UPDATE passkey_credentials SET active=0 WHERE id=? AND user_id=?", (credential_id, user["id"])); con.commit(); con.close(); flash("Méthode biométrique supprimée.", "success"); return redirect(url_for("security_settings"))


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        new_password = request.form.get("new_password", "")
        con = db()
        user = con.execute("SELECT * FROM users WHERE email=? AND phone=? AND active=1", (email, phone)).fetchone()
        if user and len(new_password) >= 6:
            con.execute("UPDATE users SET password_hash=?, force_password_change=0, failed_login_count=0, locked_until=NULL, password_changed_at=? WHERE id=?", (generate_password_hash(new_password), now(), user["id"]))
            con.commit()
            con.close()
            flash("Mot de passe modifié. Connectez-vous avec le nouveau mot de passe.", "success")
            return redirect(url_for("login"))
        con.close()
        flash("Vérifiez l'e-mail, le téléphone et utilisez au moins 6 caractères.", "danger")
    return render_template("reset_password.html")


@app.route("/devenir-membre", methods=["GET", "POST"])
def become_member():
    if request.method == "POST":
        photo_path = save_data_url_image(request.form.get("photo_capture", ""), "photos") or save_upload(request.files.get("photo"), "photos")
        phone_prefix = request.form.get("phone_country_code", "+243")
        phone_value = normalized_phone(phone_prefix, request.form.get("phone", ""))
        studies_done = ", ".join(request.form.getlist("studies_done"))
        custom_fields = collect_custom_field_values()
        con = db()
        con.execute('''INSERT INTO member_applications(first_name,last_name,gender,email,phone,nationality,province,territory,commune,localite,physical_address,birth_date,birth_place,marital_status,profession,education,studies_done,experience,motivation,photo_path,custom_fields,status,created_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            request.form.get("first_name", "").strip(),
            request.form.get("last_name", "").strip(),
            request.form.get("gender", ""),
            request.form.get("email", "").strip(),
            phone_value,
            request.form.get("nationality", "Congolaise"),
            request.form.get("province", ""),
            request.form.get("territory", ""),
            request.form.get("commune", ""),
            request.form.get("localite", ""),
            request.form.get("physical_address", ""),
            request.form.get("birth_date", ""),
            request.form.get("birth_place", ""),
            request.form.get("marital_status", ""),
            request.form.get("profession", ""),
            request.form.get("education", ""),
            studies_done,
            request.form.get("experience", ""),
            request.form.get("motivation", ""),
            photo_path,
            custom_fields,
            "pending",
            now()
        ))
        con.commit()
        con.close()
        flash("Votre demande d'adhésion est envoyée. L'administration peut l'accepter ou la rejeter.", "success")
        return redirect(url_for("login"))
    return render_template("become_member.html")



@app.route("/lang/<code>")
def set_language(code):
    if code in AVAILABLE_LANGUAGES:
        session["lang"] = code
        user = current_user()
        if user:
            con = db(); con.execute("UPDATE users SET preferred_language=? WHERE id=?", (code, user["id"])); con.commit(); con.close()
    return redirect(request.referrer or url_for("index"))


@app.route("/mon-profil")
@login_required
def my_profile():
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE user_id=? AND deleted_at IS NULL ORDER BY is_administrative DESC, id DESC LIMIT 1", (user["id"],)).fetchone()
    payments = []
    notes = []
    internal_notes = con.execute("""SELECT * FROM internal_notifications
        WHERE (user_id=? OR user_id IS NULL)
          AND (role IS NULL OR role=? OR role='all')
          AND (province IS NULL OR province='' OR province=?)
        ORDER BY created_at DESC LIMIT 10""", (user["id"], user["role"], user["province"] or "")).fetchall()
    if member:
        payments = con.execute("SELECT * FROM payments WHERE member_id=? ORDER BY created_at DESC", (member["id"],)).fetchall()
        notes = con.execute("SELECT * FROM notifications WHERE target_scope IN ('all','members') OR province=? ORDER BY created_at DESC LIMIT 10", (member["province"],)).fetchall()
    total_paid=sum(float(p['amount'] or 0) for p in payments if (p['status'] or '').lower() in ('paid','payé','paye','confirmed','confirmé'))
    unread_notes=sum(1 for n in internal_notes if not n['read_at'])
    completeness=0
    if member:
        fields=['first_name','last_name','gender','email','phone','province','localite','physical_address','birth_date','profession','photo_path']
        completeness=round(100*sum(1 for f in fields if member[f])/len(fields))
    con.close()
    return render_template("member_dashboard.html", member=member, payments=payments, notes=notes, internal_notes=internal_notes, profile_title="Mon profil", total_paid=total_paid, unread_notes=unread_notes, completeness=completeness)


@app.route("/recherche")
@login_required
def global_search():
    user = current_user()
    q = request.args.get("q", "").strip()
    like = f"%{q}%"
    results = {"members": [], "users": [], "payments": [], "activities": [], "tickets": []}
    if q:
        con = db()
        where, params = member_scope_query(user)
        results["members"] = con.execute(f"""SELECT * FROM members {where} AND (first_name LIKE ? OR last_name LIKE ? OR code LIKE ? OR phone LIKE ? OR email LIKE ?) ORDER BY joined_at DESC LIMIT 20""", params + [like, like, like, like, like]).fetchall()
        if user["role"] in NATIONAL_ROLES:
            results["users"] = con.execute("SELECT * FROM users WHERE deleted_at IS NULL AND (first_name LIKE ? OR last_name LIKE ? OR email LIKE ? OR phone LIKE ? OR role LIKE ?) LIMIT 20", (like, like, like, like, like)).fetchall()
        pay_rows, _ = get_payment_rows(user, {"status":"", "method":"", "currency":"", "contribution_type":""})
        results["payments"] = [p for p in pay_rows if q.lower() in (str(p["reference"] or "") + str(p["first_name"] or "") + str(p["last_name"] or "") + str(p["code"] or "")).lower()][:20]
        if user["role"] in NATIONAL_ROLES:
            results["activities"] = con.execute("SELECT * FROM activities WHERE title LIKE ? OR body LIKE ? OR province LIKE ? ORDER BY published_at DESC LIMIT 20", (like, like, like)).fetchall()
        else:
            results["activities"] = con.execute("SELECT * FROM activities WHERE province=? AND (title LIKE ? OR body LIKE ?) ORDER BY published_at DESC LIMIT 20", (user["province"], like, like)).fetchall()
        t_where, t_params = support_scope_sql(user)
        results["tickets"] = con.execute(f"SELECT * FROM support_tickets {t_where} AND (title LIKE ? OR message LIKE ? OR full_name LIKE ? OR tracking_code LIKE ?) ORDER BY created_at DESC LIMIT 20", t_params + [like, like, like, like]).fetchall()
        con.close()
    return render_template("search_results.html", q=q, results=results)


@app.route("/aide-intelligente", methods=["GET", "POST"])
def intelligent_help():
    question = request.form.get("question", "").strip() if request.method == "POST" else request.args.get("q", "").strip()
    answer = ""
    if question:
        ql = question.lower()
        if any(k in ql for k in ["membre", "adhésion", "adhesion", "ajouter"]):
            answer = "Pour ajouter un membre : ouvrez Administration > Membres > Ajouter membre localement, remplissez la fiche, puis enregistrez. La carte et la fiche PDF sont générées automatiquement."
        elif any(k in ql for k in ["carte", "qr", "laisser-passer"]):
            answer = "Pour imprimer une carte : ouvrez la liste des membres, cliquez sur Carte + QR. Le QR code permet de vérifier publiquement la validité du membre."
        elif any(k in ql for k in ["contribution", "paiement", "cotisation"]):
            answer = "Pour gérer les cotisations : ouvrez Cotisations, choisissez le membre, le type, le montant et le mode de paiement. Vous pouvez filtrer, imprimer et exporter selon vos droits."
        elif any(k in ql for k in ["mot de passe", "connexion", "login"]):
            answer = "Pour changer le mot de passe : cliquez sur Mon profil ou l’icône cadenas en haut. Les nouveaux comptes doivent changer le mot de passe initial."
        elif any(k in ql for k in ["langue", "lingala", "anglais", "portugais", "espagnol"]):
            answer = "Pour changer la langue : utilisez le sélecteur de langue en haut ou dans le footer. Les libellés principaux de l’application changent immédiatement."
        elif any(k in ql for k in ["paramètre", "parametre", "logo", "adresse", "structure"]):
            answer = "Les paramètres de structure sont réservés à l’Administrateur général informaticien. Il peut modifier logo, adresse, contacts, signatures, langues et informations officielles."
        else:
            answer = "Je peux vous aider sur les membres, les cartes, les cotisations, la recherche, les tickets, la langue, le profil, les impressions et les paramètres. Reformulez votre question avec l’un de ces mots-clés."
    return render_template("intelligent_help.html", question=question, answer=answer)


@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    if user["role"] in ADMIN_ROLES:
        return redirect(url_for("admin_dashboard"))
    return redirect(url_for("member_dashboard"))


@app.route("/admin")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "national_treasurer", "provincial_president", "provincial_admin", "provincial_secretary", "provincial_treasurer", "local_admin", "registration_agent")
def admin_dashboard():
    user = current_user()
    if demo_mode_enabled():
        nums = demo_numbers()
        demo_activities, demo_notifications, demo_pending, demo_anniversaries = generate_demo_lists()
        return render_template("admin_dashboard.html", total_members=nums["total_members"], active_members=nums["active_members"], pending=demo_pending, activities=demo_activities, payments=[], pending_activities=[], anniversaries=demo_anniversaries, contribution_total=nums["contribution_total"], contribution_currency=nums["contribution_currency"], recent_notifications=demo_notifications, support_tickets=[])
    con = db()
    where, params = member_scope_query(user)
    total_members = con.execute(f"SELECT COUNT(*) AS n FROM members {where}", params).fetchone()["n"]
    active_members = con.execute(f"SELECT COUNT(*) AS n FROM members {where} AND status='active'", params).fetchone()["n"]
    pending_query = "SELECT * FROM member_applications WHERE status='pending'"
    pending_params = []
    if user["role"] in PROVINCIAL_ROLES:
        pending_query += " AND province=?"; pending_params.append(user["province"])
    if user["role"] == "local_admin":
        pending_query += " AND province=? AND localite=?"; pending_params.extend([user["province"], user["localite"]])
    pending = con.execute(pending_query + " ORDER BY created_at DESC LIMIT 8", pending_params).fetchall()
    if user["role"] in NATIONAL_ROLES:
        activities = con.execute("SELECT * FROM activities ORDER BY published_at DESC LIMIT 7").fetchall()
        pending_activities = con.execute("SELECT * FROM activities WHERE status='pending' ORDER BY published_at DESC LIMIT 8").fetchall()
    elif user["role"] in PROVINCIAL_ROLES:
        activities = con.execute("SELECT * FROM activities WHERE province=? ORDER BY published_at DESC LIMIT 7", (user["province"],)).fetchall()
        pending_activities = con.execute("SELECT * FROM activities WHERE province=? AND status='pending' ORDER BY published_at DESC LIMIT 8", (user["province"],)).fetchall()
    else:
        activities = con.execute("SELECT * FROM activities WHERE province=? AND localite=? ORDER BY published_at DESC LIMIT 7", (user["province"], user["localite"])).fetchall()
        pending_activities = con.execute("SELECT * FROM activities WHERE province=? AND localite=? AND status='pending' ORDER BY published_at DESC LIMIT 8", (user["province"], user["localite"])).fetchall()
    payment_where, payment_params = payment_filter_sql(user, request.args)
    payments = con.execute(f"SELECT p.*, m.first_name, m.last_name, m.code FROM payments p LEFT JOIN members m ON m.id=p.member_id {payment_where} ORDER BY p.created_at DESC LIMIT 8", payment_params).fetchall()
    current_month = datetime.now().strftime("%m")
    current_year_value = datetime.now().strftime("%Y")
    month_rows, month_totals = get_payment_rows(user, {"month": current_month, "year": current_year_value})
    contribution_total = sum(float(t["total"] or 0) for t in month_totals)
    contribution_currency = month_totals[0]["currency"] if month_totals else "CDF"
    recent_notifications = con.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT 4").fetchall() if user["role"] in NATIONAL_ROLES else con.execute("SELECT * FROM notifications WHERE province=? OR target_scope='members' ORDER BY created_at DESC LIMIT 4", (user["province"],)).fetchall()
    ticket_where, ticket_params = support_scope_sql(user, "WHERE status!='closed'")
    support_tickets = con.execute(f"SELECT * FROM support_tickets {ticket_where} ORDER BY created_at DESC LIMIT 5", ticket_params).fetchall()
    con.close()
    anniversaries = upcoming_anniversaries(user, 30)[:8]
    return render_template("admin_dashboard.html", total_members=total_members, active_members=active_members, pending=pending, activities=activities, payments=payments, pending_activities=pending_activities, anniversaries=anniversaries, contribution_total=contribution_total, contribution_currency=contribution_currency, recent_notifications=recent_notifications, support_tickets=support_tickets)


@app.route("/admin/applications")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def applications():
    user = current_user()
    con = db()
    status = request.args.get("status", "pending")
    query = "SELECT * FROM member_applications WHERE status=?"
    params = [status]
    if user["role"] in PROVINCIAL_ROLES:
        query += " AND province=?"
        params.append(user["province"])
    if user["role"] == "local_admin":
        query += " AND province=? AND localite=?"
        params.extend([user["province"], user["localite"]])
    query += " ORDER BY created_at DESC"
    rows = con.execute(query, params).fetchall()
    con.close()
    return render_template("applications.html", rows=rows, status=status)


@app.route("/admin/applications/<int:app_id>/accept", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def accept_application(app_id):
    reviewer = current_user()
    con = db()
    a = con.execute("SELECT * FROM member_applications WHERE id=? AND status='pending'", (app_id,)).fetchone()
    if not a:
        con.close(); flash("Demande introuvable ou déjà traitée.", "warning"); return redirect(url_for("applications"))
    if reviewer["role"] in PROVINCIAL_ROLES and a["province"] != reviewer["province"]:
        con.close(); abort(403)
    if reviewer["role"] == "local_admin" and (a["province"] != reviewer["province"] or a["localite"] != reviewer["localite"]):
        con.close(); abort(403)
    try:
        con.execute("INSERT INTO users(email,phone,password_hash,role,province,localite,active,created_at) VALUES(?,?,?,?,?,?,?,?)",
                    (a["email"], a["phone"], generate_password_hash(a["phone"]), "member", a["province"], a["localite"], 1, now()))
        user_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    except sqlite3.IntegrityError:
        existing = con.execute("SELECT * FROM users WHERE email=? OR phone=?", (a["email"], a["phone"])).fetchone()
        user_id = existing["id"]
    con.execute("UPDATE users SET force_password_change=1 WHERE id=?", (user_id,))
    joined = today()
    expires = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
    con.execute('''INSERT INTO members(user_id, code, first_name, last_name, gender, email, phone, nationality, province, territory, commune, localite, physical_address, birth_date, birth_place, marital_status, profession, education, studies_done, experience, photo_path, custom_fields, adhesion_number, joined_at, expires_at, approved_by, created_by)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (user_id, "TEMP", a["first_name"], a["last_name"], a["gender"], a["email"], a["phone"], a["nationality"], a["province"], a["territory"], a["commune"], a["localite"], a["physical_address"], a["birth_date"], a["birth_place"], a["marital_status"], a["profession"], a["education"], a["studies_done"], a["experience"], a["photo_path"], a["custom_fields"], "TEMP", joined, expires, reviewer["id"], reviewer["id"]))
    member_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    code = create_member_code(member_id, a["province"])
    adhesion_number = create_adhesion_number(member_id, joined)
    con.execute("UPDATE members SET code=?, adhesion_number=? WHERE id=?", (code, adhesion_number, member_id))
    con.execute("UPDATE member_applications SET status='accepted', reviewed_at=?, reviewed_by=? WHERE id=?", (now(), reviewer["id"], app_id))
    con.commit()
    member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    con.close()
    generate_member_card_pdf(member)
    generate_adhesion_form_pdf(member)
    flash(f"Adhésion acceptée. Identifiant: {a['email']} ou {a['phone']}. Mot de passe initial: numéro de téléphone.", "success")
    return redirect(url_for("applications"))


@app.route("/admin/applications/<int:app_id>/reject", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def reject_application(app_id):
    user = current_user()
    con = db()
    app_row = con.execute("SELECT * FROM member_applications WHERE id=?", (app_id,)).fetchone()
    if not app_row:
        con.close(); abort(404)
    if user["role"] in PROVINCIAL_ROLES and app_row["province"] != user["province"]:
        con.close(); abort(403)
    if user["role"] == "local_admin" and (app_row["province"] != user["province"] or app_row["localite"] != user["localite"]):
        con.close(); abort(403)
    con.execute("UPDATE member_applications SET status='rejected', reviewed_at=?, reviewed_by=? WHERE id=?", (now(), user["id"], app_id))
    con.commit(); con.close()
    flash("Demande rejetée.", "info")
    return redirect(url_for("applications"))


@app.route("/admin/members")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def members():
    user = current_user()
    province = request.args.get("province", "")
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    status = request.args.get("status", "")
    con = db()
    where, params = member_scope_query(user)
    if province and user["role"] in NATIONAL_ROLES:
        where += " AND province=?"; params.append(province)
    if year:
        where += " AND strftime('%Y', joined_at)=?"; params.append(year)
    if month:
        where += " AND strftime('%m', joined_at)=?"; params.append(month.zfill(2))
    if status:
        where += " AND status=?"; params.append(status)
    rows = con.execute(f"SELECT * FROM members {where} ORDER BY joined_at DESC", params).fetchall()
    con.close()
    return render_template("members.html", rows=rows)


@app.route("/admin/members/export.csv")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def export_members_csv():
    user = current_user()
    con = db()
    where, params = member_scope_query(user)
    rows = con.execute(f"SELECT adhesion_number,code,last_name,first_name,phone,email,province,territory,commune,physical_address,joined_at,expires_at FROM members {where} ORDER BY province,last_name", params).fetchall()
    con.close()
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["N° adhésion", "Code", "Nom", "Prénom", "Téléphone", "E-mail", "Province", "Territoire", "Commune", "Adresse", "Adhésion", "Expiration"])
    for r in rows:
        writer.writerow(list(r))
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=membres_asbl.csv"})



@app.route("/member")
@login_required
def member_dashboard():
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE user_id=? AND deleted_at IS NULL", (user["id"],)).fetchone()
    payments = []
    notes = []
    internal_notes = []
    if member:
        payments = con.execute("SELECT * FROM payments WHERE member_id=? ORDER BY created_at DESC", (member["id"],)).fetchall()
        notes = con.execute("SELECT * FROM notifications WHERE target_scope IN ('all','members') OR province=? ORDER BY created_at DESC LIMIT 10", (member["province"],)).fetchall()
        internal_notes = con.execute("""SELECT * FROM internal_notifications
            WHERE (user_id=? OR user_id IS NULL)
              AND (role IS NULL OR role='member' OR role='all')
              AND (province IS NULL OR province='' OR province=?)
            ORDER BY created_at DESC LIMIT 10""", (user["id"], member["province"] or "")).fetchall()
    con.close()
    return render_template("member_dashboard.html", member=member, payments=payments, notes=notes, internal_notes=internal_notes, profile_title="Mon espace membre")


@app.route("/member/profile", methods=["POST"])
@login_required
def update_profile():
    user = current_user()
    phone = request.form.get("phone", "").strip()
    profession = request.form.get("profession", "").strip()
    physical_address = request.form.get("physical_address", "").strip()
    photo_path = save_upload(request.files.get("photo"), "photos")
    new_password = request.form.get("new_password", "")
    con = db()
    if photo_path:
        con.execute("UPDATE members SET photo_path=? WHERE user_id=?", (photo_path, user["id"]))
    if phone:
        con.execute("UPDATE members SET phone=? WHERE user_id=?", (phone, user["id"]))
        con.execute("UPDATE users SET phone=? WHERE id=?", (phone, user["id"]))
    if profession:
        con.execute("UPDATE members SET profession=? WHERE user_id=?", (profession, user["id"]))
    if physical_address:
        con.execute("UPDATE members SET physical_address=? WHERE user_id=?", (physical_address, user["id"]))
    if new_password and len(new_password) >= 6:
        con.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(new_password), user["id"]))
    con.commit(); con.close()
    flash("Profil mis à jour.", "success")
    return redirect(url_for("my_profile" if user["role"] in ADMIN_ROLES else "member_dashboard"))


@app.route("/card/<int:member_id>")
@login_required
def card(member_id):
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL", (member_id,)).fetchone()
    con.close()
    if not member:
        abort(404)
    if user["role"] == "member":
        abort(403)
    if not can_manage_member(user, member):
        abort(403)
    if not can_output_member_card(user):
        abort(403)
    pdf_path = generate_member_card_pdf(member)
    return send_file(pdf_path, as_attachment=True, download_name=f"carte_{member['code']}.pdf")


@app.route("/card/<int:member_id>/<side>/<fmt>")
@login_required
def card_export(member_id, side, fmt):
    user=current_user(); con=db(); member=con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL",(member_id,)).fetchone(); con.close()
    if not member: abort(404)
    if user["role"]=="member": abort(403)
    if not can_manage_member(user,member): abort(403)
    if not can_output_member_card(user): abort(403)
    assets=generate_member_card_assets(member); key="zip" if fmt=="sources" else f"{side}_{fmt}"
    if key not in assets: abort(404)
    if fmt=="sources": return send_file(assets[key],as_attachment=True,download_name=f"carte_{member['code']}_sources_photoshop.zip")
    return send_file(assets[key],as_attachment=True,download_name=f"carte_{member['code']}_{side}.{fmt}")


@app.route("/admin/members/<int:member_id>/renew-card", methods=["POST"])
@login_required
@role_required("super_admin","president","secretary","national_secretary","provincial_president","provincial_admin","provincial_secretary","local_admin","registration_agent")
def renew_member_card(member_id):
    user=current_user(); con=db(); member=con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL",(member_id,)).fetchone()
    if not member: con.close(); abort(404)
    if not can_manage_member(user,member): con.close(); abort(403)
    if not can_output_member_card(user): con.close(); abort(403)
    old=(member["expires_at"] or "")[:10]
    try: base=max(datetime.now().date(),datetime.strptime(old,"%Y-%m-%d").date())
    except Exception: base=datetime.now().date()
    new=(base+timedelta(days=365)).strftime("%Y-%m-%d")
    con.execute("UPDATE members SET expires_at=?, status='active', updated_at=? WHERE id=?",(new,now(),member_id))
    con.execute("INSERT INTO card_renewals(member_id,old_expiry,new_expiry,renewed_at,renewed_by,notes) VALUES(?,?,?,?,?,?)",(member_id,old,new,now(),user["id"],"Renouvellement annuel"))
    con.commit(); member=con.execute("SELECT * FROM members WHERE id=?",(member_id,)).fetchone(); con.close(); generate_member_card_assets(member)
    flash(f"Carte renouvelée jusqu'au {new}.","success"); return redirect(request.referrer or url_for("members"))


@app.route("/fiche/<int:member_id>")
@login_required
def fiche_adhesion(member_id):
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL", (member_id,)).fetchone()
    con.close()
    if not member:
        abort(404)
    if user["role"] == "member" and member["user_id"] != user["id"]:
        abort(403)
    if user["role"] != "member" and not can_manage_member(user, member):
        abort(403)
    pdf_path = generate_adhesion_form_pdf(member)
    return send_file(pdf_path, as_attachment=True, download_name=f"fiche_adhesion_{member['code']}.pdf")


@app.route("/admin/activities", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def manage_activities():
    user = current_user()
    con = db()
    if request.method == "POST":
        image_path = save_upload(request.files.get("image"), "activities")
        if user["role"] in NATIONAL_ROLES:
            status = "approved"; approved_at = now(); approved_by = user["id"]
        else:
            status = "pending"; approved_at = None; approved_by = None
        province = request.form.get("province", "") if user["role"] in NATIONAL_ROLES else (user["province"] or "")
        localite = request.form.get("localite", "") if user["role"] in NATIONAL_ROLES else (user["localite"] or "")
        if user["role"] in PROVINCIAL_ROLES and province != user["province"]:
            con.close(); abort(403)
        con.execute("""INSERT INTO activities(title,subtitle,body,image_path,image_fit,youtube_url,published_at,author_id,status,province,localite,approved_at,approved_by,category)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (request.form.get("title", ""), request.form.get("subtitle", ""), request.form.get("body", ""), image_path, request.form.get("image_fit", "cover"), request.form.get("youtube_url", ""), now(), user["id"], status, province, localite, approved_at, approved_by, request.form.get("category", "Activité")))
        con.commit()
        if status == "approved":
            flash("Activité publiée immédiatement.", "success")
        else:
            flash("Activité envoyée au niveau national pour validation avant publication.", "warning")
    if user["role"] in NATIONAL_ROLES:
        rows = con.execute("SELECT a.*, u.email AS author_email FROM activities a LEFT JOIN users u ON u.id=a.author_id ORDER BY a.published_at DESC").fetchall()
    elif user["role"] in PROVINCIAL_ROLES:
        rows = con.execute("SELECT a.*, u.email AS author_email FROM activities a LEFT JOIN users u ON u.id=a.author_id WHERE a.province=? ORDER BY a.published_at DESC", (user["province"],)).fetchall()
    else:
        rows = con.execute("SELECT a.*, u.email AS author_email FROM activities a LEFT JOIN users u ON u.id=a.author_id WHERE a.province=? AND a.localite=? ORDER BY a.published_at DESC", (user["province"], user["localite"])).fetchall()
    con.close()
    return render_template("manage_activities.html", rows=rows, youtube_embed=youtube_embed)


@app.route("/admin/activities/<int:activity_id>/approve", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def approve_activity(activity_id):
    user = current_user()
    con = db()
    a = con.execute("SELECT * FROM activities WHERE id=?", (activity_id,)).fetchone()
    if not a:
        con.close(); abort(404)
    con.execute("UPDATE activities SET status='approved', approved_at=?, approved_by=? WHERE id=?", (now(), user["id"], activity_id))
    con.commit(); con.close()
    log_action(user["id"], "Validation activité provinciale", "activity", activity_id)
    flash("Activité validée et publiée sur la page publique.", "success")
    return redirect(url_for("manage_activities"))


@app.route("/admin/activities/<int:activity_id>/reject", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def reject_activity(activity_id):
    user = current_user()
    con = db()
    a = con.execute("SELECT * FROM activities WHERE id=?", (activity_id,)).fetchone()
    if not a:
        con.close(); abort(404)
    con.execute("UPDATE activities SET status='rejected' WHERE id=?", (activity_id,))
    con.commit(); con.close()
    log_action(user["id"], "Rejet activité provinciale", "activity", activity_id)
    flash("Activité rejetée. Elle ne sera pas visible au public.", "warning")
    return redirect(url_for("manage_activities"))


@app.route("/admin/videos", methods=["POST"])
@login_required
@role_required("super_admin")
def add_video():
    con = db()
    con.execute("INSERT INTO videos(title,youtube_url,created_at) VALUES(?,?,?)", (request.form.get("title", ""), request.form.get("youtube_url", ""), now()))
    con.commit(); con.close()
    flash("Vidéo ajoutée.", "success")
    return redirect(url_for("settings_page"))


@app.route("/admin/carousel", methods=["POST"])
@login_required
@role_required("super_admin")
def add_carousel():
    image_path = save_upload(request.files.get("image"), "carousel")
    if image_path:
        con = db()
        con.execute("INSERT INTO carousel_images(title,message,image_path,image_fit,active,created_at) VALUES(?,?,?,?,1,?)", (request.form.get("title", ""), request.form.get("message", ""), image_path, request.form.get("image_fit", "cover"), now()))
        con.commit(); con.close()
        flash("Image ajoutée au carrousel avec son message.", "success")
    return redirect(url_for("settings_page"))


@app.route("/admin/carousel/<int:item_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin")
def delete_carousel(item_id):
    con = db()
    con.execute("DELETE FROM carousel_images WHERE id=?", (item_id,))
    con.commit(); con.close()
    flash("Image du carrousel supprimée.", "warning")
    return redirect(url_for("settings_page"))


@app.route("/admin/carousel/<int:item_id>/toggle", methods=["POST"])
@login_required
@role_required("super_admin")
def toggle_carousel(item_id):
    con = db()
    row = con.execute("SELECT * FROM carousel_images WHERE id=?", (item_id,)).fetchone()
    if not row:
        con.close(); abort(404)
    con.execute("UPDATE carousel_images SET active=? WHERE id=?", (0 if row["active"] else 1, item_id))
    con.commit(); con.close()
    flash("Statut de l’image modifié.", "success")
    return redirect(url_for("settings_page"))


@app.route("/admin/payments", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def payments():
    user = current_user()
    con = db()
    if request.method == "POST":
        member_id = request.form.get("member_id")
        member = con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL", (member_id,)).fetchone()
        if not can_manage_member(user, member):
            con.close(); abort(403)
        con.execute("""INSERT INTO payments(member_id,amount,currency,method,reference,status,paid_at,created_at,contribution_type,created_by)
                       VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (member_id, request.form.get("amount"), request.form.get("currency", "CDF"), request.form.get("method", "Mobile Money"), request.form.get("reference", ""), request.form.get("status", "paid"), request.form.get("paid_at") or today(), now(), request.form.get("contribution_type", "Cotisation"), user["id"]))
        con.commit(); flash("Cotisation enregistrée.", "success")
    where, params = member_scope_query(user)
    members_rows = con.execute(f"SELECT id,code,last_name,first_name,province FROM members {where} ORDER BY last_name", params).fetchall()
    con.close()
    rows, totals = get_payment_rows(user, request.args)
    return render_template("payments.html", rows=rows, members_rows=members_rows, totals=totals, filters=dict(request.args), query_string=request.query_string.decode())


@app.route("/admin/payments/export.csv")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def export_payments_csv():
    user = current_user()
    rows, totals = get_payment_rows(user, request.args)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Code", "Nom", "Téléphone", "Province", "Type", "Montant", "Devise", "Méthode", "Référence", "Statut", "Date paiement", "Date saisie"])
    for p in rows:
        writer.writerow([p["code"], f"{p['last_name']} {p['first_name']}", p["phone"], p["province"], p["contribution_type"], p["amount"], p["currency"], p["method"], p["reference"], p["status"], p["paid_at"], p["created_at"]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": "attachment; filename=cotisations_fobak.csv"})


@app.route("/admin/payments/print")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def print_payments():
    user = current_user()
    rows, totals = get_payment_rows(user, request.args)
    return render_template("print_payments.html", rows=rows, totals=totals, filters=request.args, printed_by=user, now=now())


@app.route("/admin/notifications", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def notifications():
    user = current_user()
    con = db()
    if request.method == "POST":
        scope = request.form.get("target_scope", "members")
        province = request.form.get("province", "")
        if user["role"] not in NATIONAL_ROLES:
            province = user["province"] or ""
            scope = "province"
        message = request.form.get("message", "")
        con.execute("INSERT INTO notifications(target_scope,province,message,created_at,sent) VALUES(?,?,?,?,0)", (scope, province, message, now()))
        if scope == "all":
            con.execute("INSERT INTO internal_notifications(role,province,title,message,link,created_at) VALUES(?,?,?,?,?,?)", ("all", None, "Notification officielle", message, url_for("notifications"), now()))
        elif scope == "province":
            con.execute("INSERT INTO internal_notifications(role,province,title,message,link,created_at) VALUES(?,?,?,?,?,?)", (None, province, "Notification provinciale", message, url_for("notifications"), now()))
        else:
            con.execute("INSERT INTO internal_notifications(role,province,title,message,link,created_at) VALUES(?,?,?,?,?,?)", ("member", province or None, "Notification aux membres", message, url_for("member_dashboard"), now()))
        con.commit(); flash("Notification interne créée. Pour l'envoi SMS/e-mail réel, connectez un service externe.", "success")
    if user["role"] in NATIONAL_ROLES:
        rows = con.execute("SELECT * FROM notifications ORDER BY created_at DESC").fetchall()
    else:
        rows = con.execute("SELECT * FROM notifications WHERE province=? OR target_scope='all' ORDER BY created_at DESC", (user["province"],)).fetchall()
    internal_rows = con.execute("""SELECT * FROM internal_notifications
        WHERE (user_id=? OR user_id IS NULL)
          AND (role IS NULL OR role=? OR role='all')
          AND (province IS NULL OR province='' OR province=? OR ? IN ('super_admin','president','secretary','national_secretary'))
        ORDER BY created_at DESC LIMIT 100""", (user['id'],user['role'],user['province'] or '',user['role'])).fetchall()
    selected_id=request.args.get('selected',type=int)
    selected=next((n for n in internal_rows if n['id']==selected_id),None)
    con.close()
    return render_template("notifications.html", rows=rows, internal_rows=internal_rows, selected=selected)



@app.route("/notifications/<int:notification_id>/open")
@login_required
def open_internal_notification(notification_id):
    """Marque une notification comme lue puis ouvre sa destination interne."""
    user = current_user(); con = db()
    row = con.execute("""SELECT * FROM internal_notifications WHERE id=?
        AND (user_id=? OR user_id IS NULL)
        AND (role IS NULL OR role=? OR role='all')
        AND (province IS NULL OR province='' OR province=? OR ? IN ('super_admin','president','secretary','national_secretary'))""",
        (notification_id, user['id'], user['role'], user['province'] or '', user['role'])).fetchone()
    if not row:
        con.close(); abort(404)
    con.execute("UPDATE internal_notifications SET read_at=COALESCE(read_at,?) WHERE id=?", (now(), notification_id))
    con.commit(); con.close()
    link=(row['link'] or '').strip()
    # Sécurité : seules les URL internes relatives sont acceptées.
    if link.startswith('/') and not link.startswith('//'):
        return redirect(link)
    return redirect(url_for('notifications', selected=notification_id))

@app.route("/notifications/read-all", methods=["POST"])
@login_required
def mark_all_notifications_read():
    user=current_user(); con=db()
    con.execute("""UPDATE internal_notifications SET read_at=? WHERE read_at IS NULL
        AND (user_id=? OR user_id IS NULL)
        AND (role IS NULL OR role=? OR role='all')
        AND (province IS NULL OR province='' OR province=? OR ? IN ('super_admin','president','secretary','national_secretary'))""",
        (now(), user['id'], user['role'], user['province'] or '', user['role']))
    con.commit(); con.close(); flash('Toutes les notifications ont été marquées comme lues.','success')
    return redirect(url_for('notifications'))


@app.route("/verification/<code>")
def verify_member(code):
    con = db()
    member = con.execute("SELECT * FROM members WHERE code=? AND deleted_at IS NULL", (code,)).fetchone()
    con.close()
    return render_template("verification.html", member=member, code=code, today=today())


@app.route("/admin/members/new", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def new_member():
    user = current_user()
    if request.method == "POST":
        province = request.form.get("province", "").strip()
        localite = request.form.get("localite", "").strip()
        if user["role"] in PROVINCIAL_ROLES and province != user["province"]:
            abort(403)
        if user["role"] == "local_admin" and (province != user["province"] or localite != user["localite"]):
            abort(403)
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip()
        phone_prefix = request.form.get("phone_country_code", "+243")
        phone = normalized_phone(phone_prefix, request.form.get("phone", ""))
        if not first_name or not last_name or not phone:
            flash("Nom, prénom et téléphone sont obligatoires.", "danger")
            return redirect(url_for("new_member"))
        if not email:
            email = f"{phone}@asbl.local"
        studies_done = ", ".join(request.form.getlist("studies_done"))
        custom_fields = collect_custom_field_values()
        photo_path = save_data_url_image(request.form.get("photo_capture", ""), "photos") or save_upload(request.files.get("photo"), "photos")
        joined = request.form.get("joined_at") or today()
        expires = request.form.get("expires_at") or (datetime.strptime(joined, "%Y-%m-%d") + timedelta(days=365)).strftime("%Y-%m-%d")
        con = db()
        try:
            con.execute("INSERT INTO users(email,phone,password_hash,role,province,localite,active,created_at) VALUES(?,?,?,?,?,?,?,?)", (email, phone, generate_password_hash(phone), "member", province, localite, 1, now()))
            user_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        except sqlite3.IntegrityError:
            existing = con.execute("SELECT * FROM users WHERE email=? OR phone=?", (email, phone)).fetchone()
            user_id = existing["id"]
        con.execute("UPDATE users SET force_password_change=1 WHERE id=?", (user_id,))
        sql = """INSERT INTO members(user_id, code, first_name, last_name, gender, email, phone, nationality, province, territory, commune, localite, physical_address, birth_date, birth_place, marital_status, profession, education, studies_done, experience, photo_path, custom_fields, adhesion_number, joined_at, expires_at, approved_by, created_by, status, updated_at, is_administrative, role_label) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
        con.execute(sql, (user_id, "TEMP", first_name, last_name, request.form.get("gender", ""), email, phone, request.form.get("nationality", "Congolaise"), province, request.form.get("territory", ""), request.form.get("commune", ""), localite, request.form.get("physical_address", ""), request.form.get("birth_date", ""), request.form.get("birth_place", ""), request.form.get("marital_status", ""), request.form.get("profession", ""), request.form.get("education", ""), studies_done, request.form.get("experience", ""), photo_path, custom_fields, "TEMP", joined, expires, user["id"], user["id"], "active", now(), 0, "Membre"))
        member_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        code = create_member_code(member_id, province)
        adhesion_number = create_adhesion_number(member_id, joined)
        con.execute("UPDATE members SET code=?, adhesion_number=? WHERE id=?", (code, adhesion_number, member_id))
        con.commit()
        member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
        con.close()
        generate_member_card_pdf(member)
        generate_adhesion_form_pdf(member)
        log_action(user["id"], "Création locale membre", "member", member_id, code)
        flash(f"Membre ajouté. Identifiant: {email} ou {phone}. Mot de passe initial: {phone}", "success")
        return redirect(url_for("members"))
    default_province = user["province"] if user["role"] not in NATIONAL_ROLES else ""
    default_localite = user["localite"] if user["role"] == "local_admin" else ""
    return render_template("member_form.html", default_province=default_province, default_localite=default_localite)


@app.route("/admin/members/print")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def print_members():
    user = current_user()
    province = request.args.get("province", "")
    year = request.args.get("year", "")
    month = request.args.get("month", "")
    status = request.args.get("status", "")
    con = db()
    where, params = member_scope_query(user)
    if province and user["role"] in NATIONAL_ROLES:
        where += " AND province=?"; params.append(province)
    if year:
        where += " AND strftime('%Y', joined_at)=?"; params.append(year)
    if month:
        where += " AND strftime('%m', joined_at)=?"; params.append(month.zfill(2))
    if status:
        where += " AND status=?"; params.append(status)
    rows = con.execute(f"SELECT * FROM members {where} ORDER BY province,last_name,first_name", params).fetchall()
    con.close()
    return render_template("print_members.html", rows=rows, printed_by=user, filters={"province": province, "year": year, "month": month, "status": status}, now=now())


@app.route("/admin/members/<int:member_id>/toggle", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin")
def toggle_member(member_id):
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    if not can_manage_member(user, member):
        con.close(); abort(403)
    new_status = "inactive" if (member["status"] or "active") == "active" else "active"
    active = 1 if new_status == "active" else 0
    con.execute("UPDATE members SET status=?, updated_at=? WHERE id=?", (new_status, now(), member_id))
    if member["user_id"]:
        con.execute("UPDATE users SET active=? WHERE id=?", (active, member["user_id"]))
    con.commit(); con.close()
    log_action(user["id"], f"Changement statut membre: {new_status}", "member", member_id)
    flash("Statut du membre modifié.", "success")
    return redirect(url_for("members"))


@app.route("/admin/members/<int:member_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary")
def delete_member(member_id):
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    if not can_manage_member(user, member):
        con.close(); abort(403)
    con.execute("UPDATE members SET deleted_at=?, status='archived', updated_at=? WHERE id=?", (now(), now(), member_id))
    if member["user_id"]:
        con.execute("UPDATE users SET active=0 WHERE id=?", (member["user_id"],))
    con.commit(); con.close()
    log_action(user["id"], "Archivage membre", "member", member_id)
    flash("Membre archivé. Il ne s'affiche plus dans la liste active.", "warning")
    return redirect(url_for("members"))


@app.route("/admin/members/<int:member_id>/delete-permanent", methods=["POST"])
@login_required
@role_required("super_admin", "secretary", "national_secretary")
def delete_member_permanent(member_id):
    user = current_user()
    con = db()
    member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
    if not can_manage_member(user, member):
        con.close(); abort(403)
    user_id = member["user_id"] if member else None
    con.execute("DELETE FROM payments WHERE member_id=?", (member_id,))
    con.execute("DELETE FROM members WHERE id=?", (member_id,))
    if user_id:
        con.execute("UPDATE users SET deleted_at=?, active=0 WHERE id=?", (now(), user_id))
    con.commit(); con.close()
    log_action(user["id"], "Suppression définitive membre", "member", member_id)
    flash("Membre effacé définitivement.", "danger")
    return redirect(url_for("members"))


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def users_page():
    user = current_user()
    con = db()
    can_create_users = user["role"] in {"president", "secretary", "national_secretary"}
    if request.method == "POST":
        if not can_create_users:
            con.close(); abort(403)
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        role = request.form.get("role", "member")
        province = request.form.get("province", "")
        localite = request.form.get("localite", "")
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        password = request.form.get("password", "") or phone or "12345678"
        photo_path = save_upload(request.files.get("photo"), "photos")
        if not first_name or not last_name or not phone:
            flash("Prénom, nom et téléphone sont obligatoires pour créer un utilisateur-personne.", "danger")
            con.close(); return redirect(url_for("users_page"))
        if not email:
            email = f"{phone}@fondation.local"
        try:
            con.execute("""INSERT INTO users(email,phone,password_hash,role,province,localite,active,created_at,force_password_change,first_name,last_name,photo_path,created_by,preferred_language)
                           VALUES(?,?,?,?,?,?,1,?,1,?,?,?,?,?)""", (email, phone, generate_password_hash(password), role, province, localite, now(), first_name, last_name, photo_path, user["id"], request.form.get("preferred_language", "fr")))
            new_user_id = con.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            member_id = create_person_member_for_user(con, new_user_id, first_name, last_name, email, phone, role, province, localite, photo_path, user["id"])
            con.commit()
            member = con.execute("SELECT * FROM members WHERE id=?", (member_id,)).fetchone()
            generate_member_card_pdf(member)
            generate_adhesion_form_pdf(member)
            flash("Utilisateur créé avec son profil complet et sa carte.", "success")
            log_action(user["id"], "Création utilisateur-personne", "user", new_user_id, role)
        except sqlite3.IntegrityError:
            flash("Cet e-mail ou téléphone existe déjà.", "danger")
    q = request.args.get("q", "").strip()
    role_filter = request.args.get("role", "").strip()
    status_filter = request.args.get("status", "").strip()
    connected_filter = request.args.get("connected", "").strip()
    where = ["u.deleted_at IS NULL"]
    params = []
    if q:
        where.append("(u.first_name LIKE ? OR u.last_name LIKE ? OR u.email LIKE ? OR u.phone LIKE ?)")
        params.extend([f"%{q}%"] * 4)
    if role_filter:
        where.append("u.role=?"); params.append(role_filter)
    if status_filter in {"active","inactive"}:
        where.append("u.active=?"); params.append(1 if status_filter == "active" else 0)
    if connected_filter == "yes":
        where.append("EXISTS(SELECT 1 FROM active_sessions s WHERE s.user_id=u.id AND s.active=1 AND datetime(s.last_seen)>=datetime('now','-15 minutes'))")
    sql = f"""
        SELECT u.*, m.id AS member_profile_id, m.code AS member_code, m.first_name AS member_first_name, m.last_name AS member_last_name, m.photo_path AS member_photo_path,
               CASE WHEN EXISTS(SELECT 1 FROM active_sessions s WHERE s.user_id=u.id AND s.active=1 AND datetime(s.last_seen)>=datetime('now','-15 minutes')) THEN 1 ELSE 0 END AS is_connected,
               (SELECT MAX(s.last_seen) FROM active_sessions s WHERE s.user_id=u.id) AS last_seen,
               (SELECT MAX(s.login_at) FROM active_sessions s WHERE s.user_id=u.id) AS login_at
        FROM users u
        LEFT JOIN members m ON m.user_id=u.id AND m.deleted_at IS NULL
        WHERE {' AND '.join(where)}
        GROUP BY u.id
        ORDER BY is_connected DESC, u.role, u.province, u.email
    """
    rows = con.execute(sql, params).fetchall()
    con.close()
    return render_template("users.html", rows=rows, can_create_users=can_create_users)


@app.route("/admin/users/export.csv")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def export_users_csv():
    import csv, io
    con = db()
    rows = con.execute("""SELECT u.*, CASE WHEN EXISTS(SELECT 1 FROM active_sessions s WHERE s.user_id=u.id AND s.active=1 AND datetime(s.last_seen)>=datetime('now','-15 minutes')) THEN 'Oui' ELSE 'Non' END AS connected FROM users u WHERE u.deleted_at IS NULL ORDER BY u.role,u.last_name,u.first_name""").fetchall()
    con.close()
    output = io.StringIO(); writer = csv.writer(output, delimiter=';')
    writer.writerow(["Prénom","Nom","E-mail","Téléphone","Rôle","Province","Localité","Actif","Connecté","Dernière connexion"])
    for r in rows:
        writer.writerow([r['first_name'] or '',r['last_name'] or '',r['email'] or '',r['phone'] or '',ROLE_LABELS.get(r['role'],r['role']),r['province'] or '',r['localite'] or '',"Oui" if r['active'] else "Non",r['connected'],r['last_login'] or ''])
    data='﻿'+output.getvalue()
    return Response(data, mimetype='text/csv; charset=utf-8', headers={'Content-Disposition':'attachment; filename=utilisateurs_fobak.csv'})


@app.route("/admin/users/print")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def print_users():
    con=db()
    rows=con.execute("""SELECT u.*, CASE WHEN EXISTS(SELECT 1 FROM active_sessions s WHERE s.user_id=u.id AND s.active=1 AND datetime(s.last_seen)>=datetime('now','-15 minutes')) THEN 1 ELSE 0 END AS is_connected FROM users u WHERE u.deleted_at IS NULL ORDER BY is_connected DESC,u.role,u.last_name,u.first_name""").fetchall()
    con.close()
    return render_template('print_users.html', rows=rows)


@app.route("/admin/users/<int:user_id>/update", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def update_user(user_id):
    admin = current_user()
    role = request.form.get("role", "member")
    province = request.form.get("province", "")
    localite = request.form.get("localite", "")
    active = 1 if request.form.get("active") == "1" else 0
    first_name = request.form.get("first_name", "").strip()
    last_name = request.form.get("last_name", "").strip()
    phone = request.form.get("phone", "").strip()
    email = request.form.get("email", "").strip()
    photo_path = save_upload(request.files.get("photo"), "photos")
    con = db()
    con.execute("""UPDATE users SET role=?, province=?, localite=?, active=?, first_name=?, last_name=?, phone=?, email=?, preferred_language=? WHERE id=?""", (role, province, localite, active, first_name, last_name, phone, email, request.form.get("preferred_language", "fr"), user_id))
    if photo_path:
        con.execute("UPDATE users SET photo_path=? WHERE id=?", (photo_path, user_id))
    password = request.form.get("password", "")
    if password:
        con.execute("UPDATE users SET password_hash=?, force_password_change=1, password_changed_at=NULL WHERE id=?", (generate_password_hash(password), user_id))
    create_person_member_for_user(con, user_id, first_name, last_name, email, phone, role, province, localite, photo_path, admin["id"])
    if active == 0:
        con.execute("UPDATE members SET status='inactive' WHERE user_id=?", (user_id,))
    con.commit(); con.close()
    log_action(admin["id"], "Modification utilisateur-personne", "user", user_id)
    flash("Utilisateur et profil membre mis à jour.", "success")
    return redirect(url_for("users_page"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "secretary", "national_secretary")
def delete_user(user_id):
    admin = current_user()
    if admin["id"] == user_id:
        flash("Vous ne pouvez pas supprimer votre propre compte.", "warning")
        return redirect(url_for("users_page"))
    con = db()
    con.execute("UPDATE users SET deleted_at=?, active=0 WHERE id=?", (now(), user_id))
    con.execute("UPDATE members SET status='archived', deleted_at=COALESCE(deleted_at, ?), updated_at=? WHERE user_id=?", (now(), now(), user_id))
    con.commit(); con.close()
    log_action(admin["id"], "Suppression utilisateur et archivage profil", "user", user_id)
    flash("Utilisateur désactivé et profil archivé.", "warning")
    return redirect(url_for("users_page"))


@app.route("/admin/activities/<int:activity_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def delete_activity(activity_id):
    user = current_user()
    con = db()
    a = con.execute("SELECT * FROM activities WHERE id=?", (activity_id,)).fetchone()
    if not a:
        con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES:
        if user["role"] in PROVINCIAL_ROLES and a["province"] != user["province"]:
            con.close(); abort(403)
        if user["role"] == "local_admin" and (a["province"] != user["province"] or a["localite"] != user["localite"]):
            con.close(); abort(403)
    con.execute("DELETE FROM activities WHERE id=?", (activity_id,))
    con.commit(); con.close()
    log_action(user["id"], "Suppression activité", "activity", activity_id)
    flash("Activité supprimée du site public.", "warning")
    return redirect(url_for("manage_activities"))


@app.route("/admin/adhesion-fields/add", methods=["POST"])
@login_required
@role_required("super_admin")
def add_adhesion_field():
    label = request.form.get("label", "").strip()
    if not label:
        flash("Le libellé de la rubrique est obligatoire.", "danger")
        return redirect(url_for("settings_page"))
    con = db()
    con.execute("""INSERT INTO adhesion_fields(section,label,field_type,options,required,active,sort_order,created_at)
                   VALUES(?,?,?,?,?,?,?,?)""", (
        request.form.get("section", "Autres informations"), label,
        request.form.get("field_type", "text"), request.form.get("options", ""),
        1 if request.form.get("required") == "1" else 0,
        1, int(request.form.get("sort_order") or 100), now()
    ))
    con.commit(); con.close()
    flash("Nouvelle rubrique ajoutée à la fiche d'adhésion.", "success")
    return redirect(url_for("settings_page"))


@app.route("/admin/adhesion-fields/<int:field_id>/update", methods=["POST"])
@login_required
@role_required("super_admin")
def update_adhesion_field(field_id):
    con = db()
    con.execute("""UPDATE adhesion_fields SET section=?, label=?, field_type=?, options=?, required=?, active=?, sort_order=? WHERE id=?""", (
        request.form.get("section", "Autres informations"), request.form.get("label", "").strip(),
        request.form.get("field_type", "text"), request.form.get("options", ""),
        1 if request.form.get("required") == "1" else 0,
        1 if request.form.get("active") == "1" else 0,
        int(request.form.get("sort_order") or 100), field_id
    ))
    con.commit(); con.close()
    flash("Rubrique mise à jour.", "success")
    return redirect(url_for("settings_page"))


@app.route("/admin/adhesion-fields/<int:field_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin")
def delete_adhesion_field(field_id):
    con = db()
    con.execute("DELETE FROM adhesion_fields WHERE id=?", (field_id,))
    con.commit(); con.close()
    flash("Rubrique supprimée. Les anciennes données déjà enregistrées ne seront plus affichées.", "warning")
    return redirect(url_for("settings_page"))


@app.route("/admin/anniversaires")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def anniversaries_page():
    user = current_user()
    days = int(request.args.get("days", 60) or 60)
    rows = upcoming_anniversaries(user, days)
    return render_template("anniversaries.html", rows=rows, days=days)


@app.route("/admin/projects", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def projects_page():
    user = current_user()
    con = db()
    if request.method == "POST":
        province = request.form.get("province", "") if user["role"] in NATIONAL_ROLES else (user["province"] or "")
        if user["role"] not in NATIONAL_ROLES and province != user["province"]:
            con.close(); abort(403)
        con.execute("INSERT INTO projects(title,province,status,description,budget,created_at,author_id) VALUES(?,?,?,?,?,?,?)",
                    (request.form.get("title", ""), province, request.form.get("status", "en cours"), request.form.get("description", ""), request.form.get("budget") or 0, now(), user["id"]))
        con.commit(); flash("Projet enregistré.", "success")
    if user["role"] in NATIONAL_ROLES:
        rows = con.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()
    else:
        rows = con.execute("SELECT * FROM projects WHERE province=? ORDER BY created_at DESC", (user["province"],)).fetchall()
    con.close()
    return render_template("projects.html", rows=rows)


@app.route("/admin/projects/<int:project_id>/update", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def update_project(project_id):
    user=current_user(); con=db(); row=con.execute("SELECT * FROM projects WHERE id=?",(project_id,)).fetchone()
    if not row: con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES and (row["province"] or "") != (user["province"] or ""): con.close(); abort(403)
    province=request.form.get("province","") if user["role"] in NATIONAL_ROLES else (user["province"] or "")
    con.execute("UPDATE projects SET title=?,province=?,status=?,description=?,budget=? WHERE id=?",(request.form.get("title",""),province,request.form.get("status","en cours"),request.form.get("description",""),request.form.get("budget") or 0,project_id))
    con.commit(); con.close(); log_action(user["id"],"Modification projet","project",project_id); flash("Projet modifié.","success")
    return redirect(url_for("projects_page"))

@app.route("/admin/projects/<int:project_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin")
def delete_project(project_id):
    user=current_user(); con=db(); row=con.execute("SELECT * FROM projects WHERE id=?",(project_id,)).fetchone()
    if not row: con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES and (row["province"] or "") != (user["province"] or ""): con.close(); abort(403)
    con.execute("DELETE FROM projects WHERE id=?",(project_id,)); con.commit(); con.close(); log_action(user["id"],"Suppression projet","project",project_id); flash("Projet supprimé.","warning")
    return redirect(url_for("projects_page"))

@app.route("/admin/documents", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def documents_page():
    user = current_user()
    con = db()
    if request.method == "POST":
        file_path = save_upload(request.files.get("file"), "activities")
        province = request.form.get("province", "") if user["role"] in NATIONAL_ROLES else (user["province"] or "")
        if user["role"] not in NATIONAL_ROLES and province != user["province"]:
            con.close(); abort(403)
        con.execute("INSERT INTO documents(title,description,file_path,public,province,created_at,author_id) VALUES(?,?,?,?,?,?,?)",
                    (request.form.get("title", ""), request.form.get("description", ""), file_path, 1 if request.form.get("public") == "1" else 0, province, now(), user["id"]))
        con.commit(); flash("Document enregistré.", "success")
    if user["role"] in NATIONAL_ROLES:
        rows = con.execute("SELECT * FROM documents ORDER BY created_at DESC").fetchall()
    else:
        rows = con.execute("SELECT * FROM documents WHERE province=? ORDER BY created_at DESC", (user["province"],)).fetchall()
    con.close()
    return render_template("documents.html", rows=rows)


@app.route("/admin/documents/<int:document_id>/update", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def update_document(document_id):
    user=current_user(); con=db(); row=con.execute("SELECT * FROM documents WHERE id=?",(document_id,)).fetchone()
    if not row: con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES and (row["province"] or "") != (user["province"] or ""): con.close(); abort(403)
    province=request.form.get("province","") if user["role"] in NATIONAL_ROLES else (user["province"] or "")
    new_file=save_upload(request.files.get("file"),"activities")
    con.execute("UPDATE documents SET title=?,description=?,public=?,province=? WHERE id=?",(request.form.get("title",""),request.form.get("description",""),1 if request.form.get("public")=="1" else 0,province,document_id))
    if new_file: con.execute("UPDATE documents SET file_path=? WHERE id=?",(new_file,document_id))
    con.commit(); con.close(); log_action(user["id"],"Modification document","document",document_id); flash("Document modifié.","success")
    return redirect(url_for("documents_page"))

@app.route("/admin/documents/<int:document_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin")
def delete_document(document_id):
    user=current_user(); con=db(); row=con.execute("SELECT * FROM documents WHERE id=?",(document_id,)).fetchone()
    if not row: con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES and (row["province"] or "") != (user["province"] or ""): con.close(); abort(403)
    con.execute("DELETE FROM documents WHERE id=?",(document_id,)); con.commit(); con.close(); log_action(user["id"],"Suppression document","document",document_id); flash("Document supprimé.","warning")
    return redirect(url_for("documents_page"))

@app.route("/admin/support")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def support_tickets_page():
    user = current_user()
    status = request.args.get("status", "")
    category = request.args.get("category", "")
    where, params = support_scope_sql(user)
    if status:
        where += " AND status=?"; params.append(status)
    if category:
        where += " AND category=?"; params.append(category)
    con = db()
    rows = con.execute(f"SELECT * FROM support_tickets {where} ORDER BY created_at DESC", params).fetchall()
    con.close()
    return render_template("support_tickets.html", rows=rows, filters=request.args)


@app.route("/admin/support/<int:ticket_id>", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary", "provincial_president", "provincial_admin", "provincial_secretary", "local_admin", "registration_agent")
def support_ticket_detail(ticket_id):
    user = current_user()
    con = db()
    ticket = con.execute("SELECT * FROM support_tickets WHERE id=?", (ticket_id,)).fetchone()
    if not ticket:
        con.close(); abort(404)
    if user["role"] not in NATIONAL_ROLES:
        if ticket["province"] and ticket["province"] != user["province"]:
            con.close(); abort(403)
        if user["role"] == "local_admin" and ticket["localite"] and ticket["localite"] != user["localite"]:
            con.close(); abort(403)
    if request.method == "POST":
        reply = request.form.get("reply", "").strip()
        new_status = request.form.get("status", ticket["status"] or "new")
        if reply:
            con.execute("INSERT INTO support_ticket_messages(ticket_id,user_id,author_name,message,created_at) VALUES(?,?,?,?,?)", (ticket_id, user["id"], user["email"] or ROLE_LABELS.get(user["role"], user["role"]), reply, now()))
        closed_at = now() if new_status == "closed" else ticket["closed_at"]
        con.execute("UPDATE support_tickets SET status=?, priority=?, assigned_to=?, updated_at=?, closed_at=? WHERE id=?", (new_status, request.form.get("priority", ticket["priority"] or "normal"), request.form.get("assigned_to") or user["id"], now(), closed_at, ticket_id))
        if ticket["user_id"]:
            con.execute("INSERT INTO internal_notifications(user_id,title,message,link,created_at) VALUES(?,?,?,?,?)", (ticket["user_id"], "Réponse au ticket support", f"{ticket['tracking_code']} : {new_status}", url_for("report_problem"), now()))
        con.commit(); flash("Ticket mis à jour.", "success")
        log_action(user["id"], "Mise à jour ticket support", "support_ticket", ticket_id)
        return redirect(url_for("support_ticket_detail", ticket_id=ticket_id))
    messages = con.execute("SELECT * FROM support_ticket_messages WHERE ticket_id=? ORDER BY created_at ASC", (ticket_id,)).fetchall()
    admins = con.execute("SELECT id,email,role,province FROM users WHERE active=1 AND deleted_at IS NULL AND role IN ('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent') ORDER BY role,email").fetchall()
    con.close()
    return render_template("support_ticket_detail.html", ticket=ticket, messages=messages, admins=admins)


@app.route("/admin/cloture-exercice", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def year_closure():
    user = current_user()
    con = db()
    if request.method == "POST":
        label = request.form.get("label", str(datetime.now().year)).strip()
        start_date = request.form.get("start_date", f"{label}-01-01")
        end_date = request.form.get("end_date", f"{label}-12-31")
        note = request.form.get("note", "")
        con.execute("""INSERT OR REPLACE INTO fiscal_years(label,start_date,end_date,status,closed_at,closed_by,note)
                       VALUES(?,?,?,?,?,?,?)""", (label, start_date, end_date, "closed", now(), user["id"], note))
        con.execute("UPDATE settings SET value=? WHERE key='current_exercise_label'", (str(int(label)+1) if label.isdigit() else label + " +1",))
        con.commit()
        create_internal_notification("Clôture d'exercice", f"L'exercice {label} a été clôturé sans suppression des historiques.", url_for("year_closure"), role="all")
        flash("Exercice clôturé. Les historiques sont conservés et une nouvelle année peut commencer.", "success")
        log_action(user["id"], "Clôture exercice", "fiscal_year", None, label)
    years = con.execute("SELECT * FROM fiscal_years ORDER BY label DESC").fetchall()
    member_count = con.execute("SELECT COUNT(*) AS n FROM members WHERE deleted_at IS NULL").fetchone()["n"]
    contribution_count = con.execute("SELECT COUNT(*) AS n FROM payments").fetchone()["n"]
    activity_count = con.execute("SELECT COUNT(*) AS n FROM activities").fetchone()["n"]
    con.close()
    return render_template("year_closure.html", years=years, member_count=member_count, contribution_count=contribution_count, activity_count=activity_count)


@app.route("/admin/demo-mode", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def toggle_demo_mode():
    enabled = "1" if request.form.get("enabled") == "1" else "0"
    con = db()
    con.execute("UPDATE settings SET value=? WHERE key='demo_mode_enabled'", (enabled,))
    con.commit(); con.close()
    session["demo_mode"] = 1 if enabled == "1" else 0
    flash("Mode démonstration activé." if enabled == "1" else "Mode démonstration désactivé.", "info")
    return redirect(url_for("admin_dashboard"))


def release_readiness_status():
    settings = get_settings()
    con = db()
    try:
        carousel_count = con.execute("SELECT COUNT(*) AS n FROM carousel_images WHERE active=1").fetchone()["n"]
        weak_accounts = con.execute("SELECT COUNT(*) AS n FROM users WHERE deleted_at IS NULL AND force_password_change=1").fetchone()["n"]
        active_users = con.execute("SELECT COUNT(*) AS n FROM users WHERE deleted_at IS NULL AND active=1").fetchone()["n"]
    finally:
        con.close()
    backup_dir = os.path.join(BASE_DIR, "backups")
    recent_backup = False
    if os.path.isdir(backup_dir):
        for name in os.listdir(backup_dir):
            full = os.path.join(backup_dir, name)
            if name.endswith(".zip") and (datetime.now().timestamp() - os.path.getmtime(full)) <= 7*86400:
                recent_backup = True
                break
    secret_ok = app.secret_key != "change-this-secret-key-before-online-hosting" and len(app.secret_key) >= 24
    checks = [
        {"label":"Base de données accessible", "ok":os.path.exists(DB_PATH) and os.access(DB_PATH, os.W_OK), "detail":DB_PATH},
        {"label":"Dossier des fichiers accessible", "ok":os.path.isdir(UPLOAD_ROOT) and os.access(UPLOAD_ROOT, os.W_OK), "detail":UPLOAD_ROOT},
        {"label":"Clé de sécurité personnalisée", "ok":secret_ok, "detail":"Définissez SECRET_KEY avant une publication Internet." if not secret_ok else "Clé personnalisée détectée."},
        {"label":"Sauvegarde récente", "ok":recent_backup, "detail":"Une sauvegarde de moins de 7 jours est recommandée."},
        {"label":"Carrousel d’accueil", "ok":carousel_count > 0, "detail":f"{carousel_count} image(s) active(s)."},
        {"label":"Comptes initiaux sécurisés", "ok":weak_accounts == 0, "detail":f"{weak_accounts} compte(s) doivent encore changer leur mot de passe."},
        {"label":"Mode démonstration désactivé", "ok":not demo_mode_enabled(), "detail":"Désactivez le mode démo avant livraison."},
        {"label":"Logo officiel configuré", "ok":bool(settings.get("logo_path")), "detail":settings.get("logo_path") or "Aucun logo."},
    ]
    score = round(sum(1 for c in checks if c["ok"]) * 100 / len(checks)) if checks else 0
    return {"checks":checks, "score":score, "carousel_count":carousel_count, "active_users":active_users, "weak_accounts":weak_accounts}


def ensure_daily_backup():
    try:
        marker_path = os.path.join(BASE_DIR, "backups", ".last_auto_backup")
        today_key = datetime.now().strftime("%Y-%m-%d")
        previous = Path(marker_path).read_text(encoding="utf-8").strip() if os.path.exists(marker_path) else ""
        if previous != today_key and os.path.exists(DB_PATH):
            create_backup_archive("Sauvegarde automatique quotidienne")
            Path(marker_path).write_text(today_key, encoding="utf-8")
            logging.info("Sauvegarde automatique quotidienne créée")
    except Exception:
        logging.exception("Échec de la sauvegarde automatique quotidienne")


@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), geolocation=(self), microphone=()")
    if request.is_secure:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


@app.route("/admin/production/readiness")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def release_readiness():
    return render_template("release_readiness.html", readiness=release_readiness_status())


@app.route("/admin/production")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def production_center_admin():
    status = production_health_status()
    con = db()
    audit_logs = con.execute("""
        SELECT a.*, u.email, u.phone, u.role
        FROM audit_logs a
        LEFT JOIN users u ON u.id=a.user_id
        ORDER BY a.created_at DESC, a.id DESC
        LIMIT 60
    """).fetchall()
    locked_users = con.execute("SELECT * FROM users WHERE deleted_at IS NULL AND locked_until IS NOT NULL ORDER BY locked_until DESC").fetchall()
    con.close()
    return render_template("production_center.html", status=status, audit_logs=audit_logs, locked_users=locked_users)


@app.route("/admin/production/migration", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def run_production_migration():
    user = current_user()
    init_db()
    log_action(user["id"], "Vérification/migration manuelle", "database", None)
    flash("Migration vérifiée : les tables et colonnes nécessaires sont présentes.", "success")
    return redirect(url_for("production_center_admin"))


@app.route("/admin/production/backup")
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def download_backup():
    user = current_user()
    archive = create_backup_archive("Téléchargement manuel par l'administration")
    log_action(user["id"], "Téléchargement sauvegarde", "backup", None, os.path.basename(archive))
    return send_file(archive, as_attachment=True, download_name=os.path.basename(archive))


@app.route("/admin/production/unlock-user/<int:user_id>", methods=["POST"])
@login_required
@role_required("super_admin", "secretary", "national_secretary")
def unlock_user(user_id):
    admin = current_user()
    con = db()
    target = con.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not target:
        con.close(); abort(404)
    con.execute("UPDATE users SET failed_login_count=0, locked_until=NULL WHERE id=?", (user_id,))
    con.commit(); con.close()
    log_action(admin["id"], "Déverrouillage compte", "user", user_id)
    flash("Compte déverrouillé.", "success")
    return redirect(url_for("production_center_admin"))


@app.route("/admin/production/force-password/<int:user_id>", methods=["POST"])
@login_required
@role_required("super_admin", "secretary", "national_secretary")
def force_user_password_change(user_id):
    admin = current_user()
    con = db()
    con.execute("UPDATE users SET force_password_change=1 WHERE id=?", (user_id,))
    con.commit(); con.close()
    log_action(admin["id"], "Changement mot de passe imposé", "user", user_id)
    flash("Le compte devra changer son mot de passe à la prochaine connexion.", "success")
    return redirect(url_for("users_page"))



def payment_receipt_pdf(payment_id, user):
    con = db()
    row = con.execute("""
        SELECT p.*, m.code, m.first_name, m.last_name, m.phone, m.email, m.province, m.territory, m.commune, m.photo_path
        FROM payments p LEFT JOIN members m ON m.id=p.member_id WHERE p.id=?
    """, (payment_id,)).fetchone()
    con.close()
    if not row:
        abort(404)
    if user and user["role"] not in NATIONAL_ROLES:
        if user["role"] in PROVINCIAL_ROLES and row["province"] != user["province"]:
            abort(403)
        if user["role"] == "local_admin" and row["province"] != user["province"]:
            abort(403)
    settings = get_settings()
    cards_dir = os.path.join(UPLOAD_ROOT, "cards")
    os.makedirs(cards_dir, exist_ok=True)
    path = os.path.join(cards_dir, f"recu_cotisation_{payment_id}.pdf")
    c = canvas.Canvas(path, pagesize=A5 if 'A5' in globals() else A4)
    page_w, page_h = (A5 if 'A5' in globals() else A4)
    margin = 12*mm
    c.setFillColor(colors.white); c.rect(0,0,page_w,page_h,fill=1,stroke=0)
    c.setStrokeColor(colors.HexColor('#0B3A66')); c.setLineWidth(1.1); c.roundRect(margin, margin, page_w-2*margin, page_h-2*margin, 3*mm, fill=0, stroke=1)
    logo_rel = settings.get('logo_path','')
    draw_logo_rect(c, logo_rel, margin+2*mm, page_h-margin-18*mm, 34*mm, 14*mm, 'LOGO')
    draw_rdc_flag(c, page_w-margin-34*mm, page_h-margin-19*mm, 30*mm, 18*mm)
    c.setFillColor(colors.HexColor('#0B3A66')); c.setFont('Helvetica-Bold', 11)
    c.drawCentredString(page_w/2, page_h-margin-7*mm, settings.get('structure_name','FONDATION')[:44].upper())
    c.setFont('Helvetica', 7); c.drawCentredString(page_w/2, page_h-margin-12*mm, settings.get('structure_header','')[:80])
    c.setFont('Helvetica-Bold', 15); c.setFillColor(colors.HexColor('#CE1021'))
    c.drawCentredString(page_w/2, page_h-margin-31*mm, 'REÇU OFFICIEL DE COTISATION')
    c.setStrokeColor(colors.HexColor('#CE1021')); c.line(page_w/2-44*mm, page_h-margin-34*mm, page_w/2+44*mm, page_h-margin-34*mm)
    # Photo du membre sur le reçu officiel
    photo_rel = row['photo_path'] or ''
    photo_abs = os.path.join(BASE_DIR, 'static', photo_rel) if photo_rel else ''
    photo_x, photo_y, photo_w, photo_h = page_w-margin-34*mm, page_h-margin-67*mm, 25*mm, 29*mm
    c.setStrokeColor(colors.HexColor('#8A8A8A')); c.roundRect(photo_x, photo_y, photo_w, photo_h, 2*mm, fill=0, stroke=1)
    if photo_abs and os.path.exists(photo_abs):
        try: c.drawImage(ImageReader(photo_abs), photo_x+1*mm, photo_y+1*mm, photo_w-2*mm, photo_h-2*mm, preserveAspectRatio=True, anchor='c', mask='auto')
        except Exception: pass
    else:
        c.setFont('Helvetica', 6); c.drawCentredString(photo_x+photo_w/2, photo_y+photo_h/2, 'PHOTO MEMBRE')
    y = page_h-margin-46*mm
    items = [
        ('N° reçu', f'RC-{datetime.now().year}-{payment_id:06d}'),
        ('Date', (row['paid_at'] or row['created_at'] or '')[:10]),
        ('Membre', f"{row['last_name'] or ''} {row['first_name'] or ''}"),
        ('Code membre', row['code'] or ''),
        ('Province', row['province'] or ''),
        ('Type', row['contribution_type'] or 'Cotisation'),
        ('Montant', f"{row['amount']} {row['currency'] or ''}"),
        ('Méthode', row['method'] or ''),
        ('Référence', row['reference'] or ''),
        ('Statut', row['status'] or ''),
    ]
    for lab, val in items:
        c.setFillColor(colors.black); c.setFont('Helvetica-Bold', 8.5); c.drawString(margin+8*mm, y, lab + ' :')
        c.setFont('Helvetica', 8.5); c.drawString(margin+44*mm, y, str(val)[:55])
        y -= 7*mm
    qr_code = qr.QrCodeWidget(verification_url(row['code'] or str(payment_id)))
    bounds = qr_code.getBounds(); qr_w = bounds[2]-bounds[0]; qr_h = bounds[3]-bounds[1]; qr_size = 23*mm
    d = Drawing(qr_size, qr_size, transform=[qr_size/qr_w,0,0,qr_size/qr_h,0,0]); d.add(qr_code)
    renderPDF.draw(d, c, page_w-margin-35*mm, margin+31*mm)
    c.setFont('Helvetica', 6); c.drawCentredString(page_w-margin-23*mm, margin+28*mm, 'QR vérification')
    sign_y = margin+18*mm
    for rel, cx, label in [(settings.get('president_signature_path',''), page_w/2-32*mm, 'Président National'), (settings.get('secretary_signature_path',''), page_w/2+32*mm, 'Bureau National')]:
        sig_abs = os.path.join(BASE_DIR, 'static', rel) if rel else ''
        if sig_abs and os.path.exists(sig_abs):
            try: c.drawImage(ImageReader(sig_abs), cx-15*mm, sign_y+2*mm, 30*mm, 11*mm, preserveAspectRatio=True, mask='auto')
            except Exception: pass
        c.setStrokeColor(colors.grey); c.line(cx-22*mm, sign_y, cx+22*mm, sign_y)
        c.setFont('Helvetica-Bold', 6.5); c.drawCentredString(cx, sign_y-4*mm, label)
    c.setFillColor(colors.HexColor('#0B3A66')); c.setFont('Helvetica', 6.3)
    c.drawCentredString(page_w/2, margin+5*mm, (settings.get('headquarters','') + ' — ' + settings.get('contact_phones',''))[:120])
    c.showPage(); c.save()
    return path


@app.route('/admin/payments/<int:payment_id>/receipt')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent')
def payment_receipt(payment_id):
    user = current_user()
    path = payment_receipt_pdf(payment_id, user)
    log_action(user['id'], 'Impression reçu cotisation', 'payment', payment_id)
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))


@app.route('/qr/member/<code>.png')
def member_qr_png(code):
    con=db(); member=con.execute("SELECT code FROM members WHERE code=? AND deleted_at IS NULL",(code,)).fetchone(); con.close()
    if not member: abort(404)
    img=qrcode.make(verification_url(code)).convert('RGB')
    bio=BytesIO(); img.save(bio,format='PNG'); bio.seek(0)
    return send_file(bio,mimetype='image/png',download_name=f'qr_{code}.png')


@app.route('/admin/payments/<int:payment_id>/print/<paper>')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer','national_treasurer','local_admin','registration_agent')
def payment_receipt_print(payment_id, paper):
    user=current_user(); con=db()
    row=con.execute("""SELECT p.*,m.code,m.first_name,m.last_name,m.phone,m.email,m.province,m.photo_path
                       FROM payments p LEFT JOIN members m ON m.id=p.member_id WHERE p.id=?""",(payment_id,)).fetchone(); con.close()
    if not row: abort(404)
    if user['role'] not in NATIONAL_ROLES and row['province'] != (user['province'] or ''): abort(403)
    if paper not in ('80mm','58mm','a4','a5'): paper='80mm'
    log_action(user['id'],f'Impression reçu {paper}','payment',payment_id)
    return render_template('payment_receipt_print.html', payment=row, paper=paper, receipt_no=f"RC-{datetime.now().year}-{payment_id:06d}", printed_by=user, now=now())

@app.route('/payer-cotisation', methods=['GET','POST'])
def public_pay_contribution():
    settings=get_settings()
    if request.method=='POST':
        code=request.form.get('member_code','').strip(); phone=request.form.get('phone','').strip()
        con=db(); member=con.execute("SELECT * FROM members WHERE deleted_at IS NULL AND code=? AND REPLACE(REPLACE(phone,' ',''),'+','')=REPLACE(REPLACE(?,' ',''),'+','')",(code,phone)).fetchone()
        if not member:
            con.close(); flash('Code membre ou numéro de téléphone non reconnu.','danger'); return redirect(url_for('public_pay_contribution'))
        amount=request.form.get('amount','').strip(); reference=request.form.get('reference','').strip()
        if not amount or not reference:
            con.close(); flash('Le montant et la référence de transaction sont obligatoires.','danger'); return redirect(url_for('public_pay_contribution'))
        cur=con.cursor(); cur.execute("""INSERT INTO payments(member_id,amount,currency,method,reference,status,paid_at,created_at,contribution_type,created_by)
                    VALUES(?,?,?,?,?,?,?,?,?,NULL)""",(member['id'],amount,request.form.get('currency','CDF'),'M-Pesa / Vodacom',reference,'pending',request.form.get('paid_at') or today(),now(),request.form.get('contribution_type','Cotisation')))
        pid=cur.lastrowid; con.commit(); con.close()
        create_internal_notification('Cotisation en ligne à vérifier',f"{member['code']} — {amount} {request.form.get('currency','CDF')} — réf. {reference}",url_for('payments'),role='all' if not member['province'] else None,province=member['province'] or None)
        flash(f'Déclaration reçue sous le numéro PC-{pid:06d}. Elle sera validée après contrôle de la transaction.','success')
        return redirect(url_for('public_pay_contribution'))
    return render_template('public_payment.html', settings=settings)


@app.route('/admin/members/export.xlsx')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent')
def export_members_xlsx():
    try:
        from openpyxl import Workbook
    except Exception:
        flash('Le module openpyxl n’est pas installé. Utilisez pip install openpyxl ou exportez en CSV.', 'danger')
        return redirect(url_for('members'))
    user = current_user(); rows = scoped_members_rows(user)
    wb = Workbook(); ws = wb.active; ws.title = 'Membres'
    headers = ['Code','N° adhésion','Nom','Prénom','Téléphone','Email','Province','Territoire','Commune','Localité','Statut','Date adhésion','Expiration','Rôle/Fonction']
    ws.append(headers)
    for m in rows:
        ws.append([m['code'], m['adhesion_number'], m['last_name'], m['first_name'], m['phone'], m['email'], m['province'], m['territory'], m['commune'], m['localite'], m['status'], (m['joined_at'] or '')[:10], (m['expires_at'] or '')[:10], m['role_label'] or m['profession']])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    log_action(user['id'], 'Export Excel membres', 'members', None)
    return send_file(bio, as_attachment=True, download_name='membres_fondation.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/payments/export.xlsx')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent')
def export_payments_xlsx():
    try:
        from openpyxl import Workbook
    except Exception:
        flash('Le module openpyxl n’est pas installé. Utilisez pip install openpyxl ou exportez en CSV.', 'danger')
        return redirect(url_for('payments'))
    user = current_user(); rows, totals = get_payment_rows(user, request.args)
    wb = Workbook(); ws = wb.active; ws.title = 'Cotisations'
    ws.append(['Date','Code membre','Nom','Province','Type','Montant','Devise','Méthode','Référence','Statut'])
    for pmt in rows:
        ws.append([(pmt['paid_at'] or pmt['created_at'] or '')[:10], pmt['code'], f"{pmt['last_name']} {pmt['first_name']}", pmt['province'], pmt['contribution_type'], pmt['amount'], pmt['currency'], pmt['method'], pmt['reference'], pmt['status']])
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    log_action(user['id'], 'Export Excel cotisations', 'payments', None)
    return send_file(bio, as_attachment=True, download_name='cotisations_fobak.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/members/import', methods=['GET','POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def import_members():
    user = current_user()
    report = []
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename:
            flash('Veuillez choisir un fichier CSV ou Excel.', 'danger')
            return redirect(url_for('import_members'))
        rows = []
        filename = f.filename.lower()
        if filename.endswith('.csv'):
            text = f.stream.read().decode('utf-8-sig', errors='ignore')
            rows = list(csv.DictReader(StringIO(text)))
        elif filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except Exception:
                flash('openpyxl est nécessaire pour importer Excel. Installez-le avec pip install openpyxl.', 'danger')
                return redirect(url_for('import_members'))
            wb = load_workbook(f, data_only=True); ws = wb.active
            headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))})
        else:
            flash('Format non accepté. Utilisez .csv ou .xlsx.', 'danger')
            return redirect(url_for('import_members'))
        con = db(); created = 0
        for idx, r in enumerate(rows, start=2):
            def pick(*names):
                for n in names:
                    if n in r and r[n] not in (None, ''):
                        return str(r[n]).strip()
                return ''
            first = pick('prenom','prénom','first_name','nom')
            last = pick('postnom','post-nom','last_name','nom complet','nom')
            if not first and not last:
                report.append(f'Ligne {idx}: ignorée, nom manquant.'); continue
            phone = pick('telephone','téléphone','phone') or f'import{datetime.now().strftime("%H%M%S")}{idx}'
            email = pick('email','e-mail') or f'{phone}@fondation.local'
            province = pick('province') or ('National' if user['role'] in NATIONAL_ROLES else user['province'])
            if user['role'] not in NATIONAL_ROLES and province != user['province']:
                report.append(f'Ligne {idx}: province non autorisée.'); continue
            localite = pick('localite','localité')
            cur = con.execute("INSERT INTO users(email,phone,password_hash,role,province,localite,active,created_at,force_password_change,first_name,last_name,created_by) VALUES(?,?,?,?,?,?,1,?,1,?,?,?)", (email, phone, generate_password_hash(phone), 'member', province, localite, now(), first, last, user['id']))
            uid = cur.lastrowid
            joined = pick('date adhesion','date adhésion','adhesion','adhésion') or today()
            expires = (datetime.now() + timedelta(days=365)).strftime('%Y-%m-%d')
            con.execute("""INSERT INTO members(user_id, code, first_name, last_name, gender, email, phone, nationality, province, territory, commune, localite, physical_address, birth_date, birth_place, marital_status, profession, education, studies_done, experience, photo_path, custom_fields, adhesion_number, joined_at, expires_at, approved_by, created_by, status, updated_at, is_administrative, role_label)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (uid, 'TEMP', first, last, pick('sexe','genre'), email, phone, pick('nationalite','nationalité') or 'Congolaise', province, pick('territoire'), pick('commune'), localite, pick('adresse','adresse physique'), pick('date naissance'), pick('lieu naissance'), pick('etat civil','état civil'), pick('profession','fonction'), pick('niveau etude','études','education'), pick('etudes faites','études faites'), pick('experience','expérience'), '', '{}', 'TEMP', joined, expires, user['id'], user['id'], 'active', now(), 0, pick('role','rôle','fonction')))
            mid = con.execute('SELECT last_insert_rowid() AS id').fetchone()['id']
            code = create_member_code(mid, province); adhesion = create_adhesion_number(mid, joined)
            con.execute('UPDATE members SET code=?, adhesion_number=? WHERE id=?', (code, adhesion, mid))
            created += 1
        con.commit(); con.close()
        log_action(user['id'], 'Import membres', 'members', None, f'{created} membre(s) importé(s)')
        flash(f'Import terminé : {created} membre(s) ajouté(s).', 'success')
    return render_template('import_members.html', report=report)


@app.route('/admin/production/restore', methods=['POST'])
@login_required
@role_required('super_admin')
def restore_backup():
    user = current_user()
    f = request.files.get('backup_file')
    if not f or not f.filename.endswith('.zip'):
        flash('Veuillez fournir une archive .zip de sauvegarde.', 'danger')
        return redirect(url_for('production_center_admin'))
    current = create_backup_archive('Sauvegarde automatique avant restauration')
    temp_path = os.path.join(BASE_DIR, 'backups', secure_filename(f'restore_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'))
    f.save(temp_path)
    try:
        with zipfile.ZipFile(temp_path, 'r') as zf:
            names = zf.namelist()
            if 'asbl.db' not in names:
                raise ValueError('Archive invalide : asbl.db introuvable')
            zf.extract('asbl.db', BASE_DIR)
            for name in names:
                norm = name.replace('\\','/')
                if norm.startswith('static/uploads/') and not norm.endswith('/'):
                    zf.extract(name, BASE_DIR)
        init_db()
        log_action(user['id'], 'Restauration sauvegarde', 'backup', None, os.path.basename(temp_path))
        flash('Sauvegarde restaurée. Une copie de sécurité de l’ancienne base a été gardée.', 'success')
    except Exception as e:
        shutil.copy(current, os.path.join(BASE_DIR, 'backups', 'restauration_echouee_copie.zip'))
        flash('Restauration impossible : ' + str(e), 'danger')
    return redirect(url_for('production_center_admin'))


@app.route('/admin/roles', methods=['GET','POST'])
@login_required
@role_required('super_admin')
def roles_permissions_page():
    modules = PERMISSION_MODULES
    actions = PERMISSION_ACTIONS
    con=db()
    if request.method=='POST':
        for role_key, role_label in ROLE_LABELS.items():
            new_label=request.form.get(f'label_{role_key}',role_label).strip() or role_label
            for module in modules:
                for action in actions:
                    pkey=f'{module}.{action}'
                    allowed=1 if request.form.get(f'{role_key}__{module}__{action}')=='1' else 0
                    con.execute('DELETE FROM role_permissions WHERE role_key=? AND permission_key=?',(role_key,pkey))
                    con.execute('INSERT INTO role_permissions(role_key,role_label,permission_key,allowed,updated_at) VALUES(?,?,?,?,?)',(role_key,new_label,pkey,allowed,now()))
        con.commit(); flash('Autorisations complètes enregistrées.','success'); log_action(current_user()['id'],'Modification des autorisations par module','roles',None)
    rows=con.execute('SELECT * FROM role_permissions').fetchall(); con.close()
    data={r:{m:{a:0 for a in actions} for m in modules} for r in ROLE_LABELS}; labels=dict(ROLE_LABELS)
    for row in rows:
        key=row['permission_key']
        if '.' in key:
            module,action=key.split('.',1)
            if module in modules and action in actions: data.setdefault(row['role_key'],{}).setdefault(module,{})[action]=row['allowed']
        labels[row['role_key']]=row['role_label']
    return render_template('roles_permissions.html',modules=modules,actions=actions,data=data,labels=labels)

def treasury_scope_sql(user, base='WHERE 1=1'):
    where, params = base, []
    if user and user['role'] in PROVINCIAL_ROLES:
        where += ' AND province=?'; params.append(user['province'] or '')
    elif user and user['role'] == 'national_treasurer':
        # Le trésorier national consulte toutes les provinces mais ne modifie que la caisse nationale.
        pass
    if user and user['role'] == 'local_admin':
        where += ' AND province=? AND localite=?'; params.extend([user['province'] or '', user['localite'] or ''])
    if user and user['role'] in NATIONAL_ROLES and request.args.get('province'):
        where += ' AND province=?'; params.append(request.args.get('province'))
    for key in ['type','currency','category']:
        val = request.args.get(key, '')
        if val:
            where += f' AND {key}=?'; params.append(val)
    if request.args.get('date_start'):
        where += ' AND date(entry_date)>=date(?)'; params.append(request.args.get('date_start'))
    if request.args.get('date_end'):
        where += ' AND date(entry_date)<=date(?)'; params.append(request.args.get('date_end'))
    return where, params


def treasury_decision_data(con, where, params):
    evolution = con.execute(f"""
        SELECT substr(entry_date,1,7) AS month, currency,
               SUM(CASE WHEN type='entrée' THEN amount ELSE 0 END) AS income,
               SUM(CASE WHEN type='sortie' THEN amount ELSE 0 END) AS expense
        FROM treasury_entries t {where}
        GROUP BY substr(entry_date,1,7), currency
        ORDER BY month DESC LIMIT 24
    """, params).fetchall()
    balances = con.execute(f"""
        SELECT currency,
               SUM(CASE WHEN type='entrée' THEN amount ELSE -amount END) AS balance,
               SUM(CASE WHEN type='entrée' THEN amount ELSE 0 END) AS income,
               SUM(CASE WHEN type='sortie' THEN amount ELSE 0 END) AS expense
        FROM treasury_entries t {where}
        GROUP BY currency
    """, params).fetchall()
    return evolution, balances


@app.route('/admin/tresorerie', methods=['GET','POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary','national_treasurer','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def treasury_page():
    user = current_user(); con = db()
    can_edit = is_treasury_editor(user)
    if request.method == 'POST':
        if not can_edit:
            con.close(); abort(403)
        if user['role'] == 'provincial_treasurer':
            province = user['province'] or ''; localite = user['localite'] or ''
        elif user['role'] == 'national_treasurer':
            province = ''; localite = ''
        else:
            province = request.form.get('province',''); localite = request.form.get('localite','')
        cur = con.execute("""INSERT INTO treasury_entries(type,category,description,amount,currency,method,reference,province,localite,entry_date,created_by,created_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", (request.form.get('type','entrée'), request.form.get('category',''), request.form.get('description',''), request.form.get('amount') or 0, request.form.get('currency','CDF'), request.form.get('method',''), request.form.get('reference',''), province, localite, request.form.get('entry_date') or today(), user['id'], now()))
        con.commit(); flash('Opération de trésorerie enregistrée.', 'success')
        log_action(user['id'], 'Ajout trésorerie', 'treasury', cur.lastrowid, f"{request.form.get('amount') or 0} {request.form.get('currency','CDF')}", url_for('treasury_page'))
    where, params = treasury_scope_sql(user)
    rows = con.execute(f'SELECT t.*, u.email FROM treasury_entries t LEFT JOIN users u ON u.id=t.created_by {where} ORDER BY entry_date DESC, id DESC', params).fetchall()
    totals = con.execute(f'SELECT type,currency,SUM(amount) AS total,COUNT(*) AS n FROM treasury_entries t {where} GROUP BY type,currency', params).fetchall()
    evolution, balances = treasury_decision_data(con, where, params)
    con.close()
    return render_template('treasury.html', rows=rows, totals=totals, evolution=evolution, balances=balances, filters=request.args, can_edit=can_edit)


@app.route('/admin/tresorerie/<int:entry_id>/modifier', methods=['POST'])
@login_required
@role_required('super_admin','national_treasurer','provincial_treasurer')
def treasury_edit(entry_id):
    user=current_user(); con=db(); row=con.execute('SELECT * FROM treasury_entries WHERE id=?',(entry_id,)).fetchone()
    if not row: con.close(); abort(404)
    if not can_edit_treasury_entry(user,row): con.close(); abort(403)
    con.execute('UPDATE treasury_entries SET type=?,category=?,description=?,amount=?,currency=?,method=?,reference=?,entry_date=? WHERE id=?',
                (request.form.get('type',row['type']),request.form.get('category',row['category']),request.form.get('description',row['description']),request.form.get('amount',row['amount']),request.form.get('currency',row['currency']),request.form.get('method',row['method']),request.form.get('reference',row['reference']),request.form.get('entry_date',row['entry_date']),entry_id))
    con.commit(); con.close(); log_action(user['id'],'Modification trésorerie','treasury',entry_id,link=url_for('treasury_page')); flash('Opération mise à jour.','success')
    return redirect(url_for('treasury_page'))


@app.route('/admin/tresorerie/export.csv')
@login_required
@role_required('super_admin','president','secretary','national_secretary','national_treasurer','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def treasury_export_csv():
    user=current_user(); con=db(); where, params = treasury_scope_sql(user)
    rows=con.execute(f'SELECT * FROM treasury_entries t {where} ORDER BY entry_date DESC,id DESC', params).fetchall(); con.close()
    output=StringIO(); writer=csv.writer(output); writer.writerow(['Date','Type','Catégorie','Description','Montant','Devise','Méthode','Référence','Province'])
    for r in rows: writer.writerow([r['entry_date'], r['type'], r['category'], r['description'], r['amount'], r['currency'], r['method'], r['reference'], r['province']])
    log_action(user['id'],'Export trésorerie','treasury',None,link=url_for('treasury_page'))
    return Response(output.getvalue(), mimetype='text/csv', headers={'Content-Disposition':'attachment; filename=tresorerie.csv'})


@app.route('/admin/tresorerie/print')
@login_required
@role_required('super_admin','president','secretary','national_secretary','national_treasurer','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def treasury_print():
    user=current_user(); con=db(); where, params = treasury_scope_sql(user)
    rows=con.execute(f'SELECT * FROM treasury_entries t {where} ORDER BY entry_date DESC,id DESC', params).fetchall()
    evolution, balances = treasury_decision_data(con, where, params); con.close()
    log_action(user['id'],'Impression trésorerie','treasury',None,link=url_for('treasury_page'))
    return render_template('print_treasury.html', rows=rows, filters=request.args, evolution=evolution, balances=balances, scope_province=(user['province'] if user['role'] in PROVINCIAL_ROLES else request.args.get('province','')))


def _report_scope(user, requested_province=""):
    if user["role"] in PROVINCIAL_ROLES:
        return user["province"] or ""
    if user["role"] in NATIONAL_ROLES:
        return requested_province or ""
    return user["province"] or ""


def build_monthly_report_data(con, province, report_month):
    start = report_month + "-01"
    try:
        dt = datetime.strptime(start, "%Y-%m-%d")
        if dt.month == 12:
            next_start = f"{dt.year+1}-01-01"
        else:
            next_start = f"{dt.year}-{dt.month+1:02d}-01"
    except Exception:
        report_month = datetime.now().strftime("%Y-%m")
        start = report_month + "-01"
        dt = datetime.strptime(start, "%Y-%m-%d")
        next_start = f"{dt.year+1}-01-01" if dt.month == 12 else f"{dt.year}-{dt.month+1:02d}-01"
    scope_sql = " AND province=?" if province else ""
    scope_params = [province] if province else []
    members_new = con.execute(f"SELECT COUNT(*) n FROM members WHERE deleted_at IS NULL AND date(joined_at)>=date(?) AND date(joined_at)<date(?) {scope_sql}", [start,next_start]+scope_params).fetchone()["n"]
    members_total = con.execute(f"SELECT COUNT(*) n FROM members WHERE deleted_at IS NULL {scope_sql}", scope_params).fetchone()["n"]
    members_active = con.execute(f"SELECT COUNT(*) n FROM members WHERE deleted_at IS NULL AND status='active' {scope_sql}", scope_params).fetchone()["n"]
    pay_sql = " AND m.province=?" if province else ""
    payments = con.execute(f"""SELECT p.currency, SUM(COALESCE(p.amount,0)) total, COUNT(*) n FROM payments p JOIN members m ON m.id=p.member_id WHERE date(COALESCE(p.paid_at,p.created_at))>=date(?) AND date(COALESCE(p.paid_at,p.created_at))<date(?) {pay_sql} GROUP BY p.currency""", [start,next_start]+scope_params).fetchall()
    tr_sql = " AND province=?" if province else ""
    treasury = con.execute(f"""SELECT currency, SUM(CASE WHEN type='entrée' THEN amount ELSE 0 END) income, SUM(CASE WHEN type='sortie' THEN amount ELSE 0 END) expense FROM treasury_entries WHERE date(entry_date)>=date(?) AND date(entry_date)<date(?) {tr_sql} GROUP BY currency""", [start,next_start]+scope_params).fetchall()
    act_sql = " AND province=?" if province else ""
    activities = con.execute(f"SELECT COUNT(*) n FROM activities WHERE date(published_at)>=date(?) AND date(published_at)<date(?) {act_sql}", [start,next_start]+scope_params).fetchone()["n"]
    office = con.execute("SELECT * FROM provincial_offices WHERE province=?", (province,)).fetchone() if province else None
    return dict(report_month=report_month, province=province, members_new=members_new, members_total=members_total, members_active=members_active, payments=payments, treasury=treasury, activities=activities, office=office)


@app.route('/admin/rapports-mensuels')
@login_required
@role_required('super_admin','president','secretary','national_secretary','national_treasurer','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def monthly_reports_page():
    user=current_user(); month=request.args.get('month') or datetime.now().strftime('%Y-%m'); province=_report_scope(user, request.args.get('province',''))
    con=db(); data=build_monthly_report_data(con,province,month)
    con.execute("INSERT OR IGNORE INTO monthly_reports(province,report_month,status,generated_at,generated_by) VALUES(?,?,?,?,?)", (province or 'National',month,'generated',now(),user['id']))
    reports=con.execute("SELECT * FROM monthly_reports WHERE (?='' OR province=?) ORDER BY report_month DESC, generated_at DESC LIMIT 24", (province,province)).fetchall()
    con.commit(); con.close()
    log_action(user['id'],'Génération rapport mensuel','monthly_report',None,f"{province or 'National'} {month}",url_for('monthly_reports_page',month=month,province=province))
    return render_template('monthly_reports.html', data=data, reports=reports, month=month, scope_province=province)


@app.route('/admin/rapports-mensuels/imprimer')
@login_required
@role_required('super_admin','president','secretary','national_secretary','national_treasurer','provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def monthly_report_print():
    user=current_user(); month=request.args.get('month') or datetime.now().strftime('%Y-%m'); province=_report_scope(user, request.args.get('province',''))
    con=db(); data=build_monthly_report_data(con,province,month); con.close()
    log_action(user['id'],'Impression rapport mensuel','monthly_report',None,f"{province or 'National'} {month}")
    return render_template('print_monthly_report.html', data=data)


@app.route('/admin/rapports-mensuels/soumettre', methods=['POST'])
@login_required
@role_required('provincial_president','provincial_admin','provincial_secretary','provincial_treasurer')
def monthly_report_submit():
    user=current_user(); month=request.form.get('month') or datetime.now().strftime('%Y-%m'); province=user['province'] or ''
    con=db(); con.execute("INSERT OR IGNORE INTO monthly_reports(province,report_month,status,generated_at,generated_by) VALUES(?,?,?,?,?)",(province,month,'generated',now(),user['id']))
    con.execute("UPDATE monthly_reports SET status='submitted',submitted_at=?,submitted_by=? WHERE province=? AND report_month=?",(now(),user['id'],province,month)); con.commit(); con.close()
    create_internal_notification('Rapport provincial reçu',f'{province} — rapport {month}',url_for('monthly_reports_page',month=month,province=province),role='all')
    log_action(user['id'],'Soumission rapport provincial','monthly_report',None,f'{province} {month}')
    flash('Rapport mensuel transmis au Bureau National.','success'); return redirect(url_for('monthly_reports_page',month=month))


@app.route('/admin/bureau-provincial', methods=['GET','POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary')
def provincial_office_settings():
    user=current_user(); province=user['province'] if user['role'] in PROVINCIAL_ROLES else request.values.get('province','')
    if not province: province='Kinshasa'
    con=db()
    if request.method=='POST':
        president_sig=save_upload(request.files.get('president_signature'),'signatures')
        secretary_sig=save_upload(request.files.get('secretary_signature'),'signatures')
        old=con.execute('SELECT * FROM provincial_offices WHERE province=?',(province,)).fetchone()
        president_sig=president_sig or (old['president_signature_path'] if old else '')
        secretary_sig=secretary_sig or (old['secretary_signature_path'] if old else '')
        con.execute("""INSERT INTO provincial_offices(province,president_name,president_function,president_signature_path,secretary_name,secretary_function,secretary_signature_path,office_address,office_phones,updated_at,updated_by)
        VALUES(?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(province) DO UPDATE SET president_name=excluded.president_name,president_function=excluded.president_function,president_signature_path=excluded.president_signature_path,secretary_name=excluded.secretary_name,secretary_function=excluded.secretary_function,secretary_signature_path=excluded.secretary_signature_path,office_address=excluded.office_address,office_phones=excluded.office_phones,updated_at=excluded.updated_at,updated_by=excluded.updated_by""",
        (province,request.form.get('president_name',''),request.form.get('president_function','Président provincial'),president_sig,request.form.get('secretary_name',''),request.form.get('secretary_function','Secrétaire provincial'),secretary_sig,request.form.get('office_address',''),request.form.get('office_phones',''),now(),user['id']))
        con.commit(); flash('Informations du bureau provincial enregistrées.','success'); log_action(user['id'],'Mise à jour bureau provincial','province',None,province)
    office=con.execute('SELECT * FROM provincial_offices WHERE province=?',(province,)).fetchone(); con.close()
    return render_template('provincial_office.html',office=office,selected_province=province)


@app.route('/guide-rapide')
def quick_guide():
    return render_template('quick_guide.html')


@app.route('/admin/alertes')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent')
def alerts_page():
    user=current_user(); con=db()
    where, params = member_scope_query(user, 'WHERE deleted_at IS NULL')
    try:
        expiring = con.execute(
            f"SELECT * FROM members {where} AND expires_at IS NOT NULL AND date(expires_at)<=date('now','+90 day') ORDER BY date(expires_at) ASC LIMIT 50",
            params,
        ).fetchall()
        if user['role'] in NATIONAL_ROLES:
            pending_activities = con.execute(
                "SELECT * FROM activities WHERE COALESCE(status,'approved')='pending' ORDER BY published_at DESC LIMIT 50"
            ).fetchall()
        else:
            pending_activities = con.execute(
                "SELECT * FROM activities WHERE COALESCE(status,'approved')='pending' AND province=? ORDER BY published_at DESC LIMIT 50",
                (user['province'],),
            ).fetchall()
        tickets_where, tickets_params = support_scope_sql(user, "WHERE status NOT IN ('closed','résolu','resolu')")
        tickets = con.execute(
            f'SELECT * FROM support_tickets {tickets_where} ORDER BY created_at DESC LIMIT 50',
            tickets_params,
        ).fetchall()
        years = upcoming_anniversaries(user, 30)
    finally:
        con.close()
    return render_template('alerts.html', expiring=expiring, pending_activities=pending_activities, tickets=tickets, years=years, today=today())


@app.route('/admin/members/<int:member_id>/historique')
@login_required
@role_required('super_admin','president','secretary','national_secretary','provincial_president','provincial_admin','provincial_secretary','local_admin','registration_agent')
def member_history(member_id):
    user=current_user(); con=db(); member=con.execute('SELECT * FROM members WHERE id=?', (member_id,)).fetchone()
    if not member: con.close(); abort(404)
    if not can_manage_member(user, member): con.close(); abort(403)
    payments_rows=con.execute('SELECT * FROM payments WHERE member_id=? ORDER BY created_at DESC', (member_id,)).fetchall()
    actions=con.execute("SELECT * FROM audit_logs WHERE target_id=? AND target_type IN ('member','members','payment') ORDER BY created_at DESC LIMIT 80", (member_id,)).fetchall()
    con.close(); return render_template('member_history.html', member=member, payments=payments_rows, actions=actions)


@app.route('/admin/assistant-demarrage', methods=['GET','POST'])
@login_required
@role_required('super_admin')
def setup_wizard():
    con=db()
    if request.method == 'POST':
        keys=['structure_name','structure_motto','structure_header','headquarters','contact_phones','president_name','secretary_name','default_language','public_base_url']
        for key in keys:
            con.execute('UPDATE settings SET value=? WHERE key=?', (request.form.get(key,''), key))
        logo_paths=save_logo_upload(request.files.get('logo'))
        for key, value in logo_paths.items():
            con.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
        con.execute('UPDATE app_setup SET completed=1, completed_at=?, completed_by=? WHERE id=1', (now(), current_user()['id']))
        con.commit(); flash('Première configuration enregistrée.', 'success'); return redirect(url_for('admin_dashboard'))
    status=con.execute('SELECT * FROM app_setup WHERE id=1').fetchone(); con.close()
    return render_template('setup_wizard.html', setup=status)


@app.route("/statuts")
def public_statutes():
    con = db()
    document = con.execute("SELECT * FROM official_documents WHERE document_type='statuts' AND active=1 AND deleted_at IS NULL AND public=1 ORDER BY id DESC LIMIT 1").fetchone()
    sections = con.execute("SELECT * FROM statute_sections WHERE active=1 AND public=1 ORDER BY sort_order,id").fetchall()
    con.close()
    return render_template("statutes.html", document=document, sections=sections, admin_mode=False)


@app.route("/admin/statuts", methods=["GET", "POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def manage_statutes():
    user = current_user()
    con = db()
    if request.method == "POST":
        title = request.form.get("title", "Statuts et Règlement intérieur FOBAK").strip()
        file_path = save_upload(request.files.get("file"), "official_docs")
        if not file_path:
            flash("Veuillez sélectionner un document PDF ou image.", "warning")
        else:
            con.execute("UPDATE official_documents SET active=0 WHERE document_type='statuts' AND active=1")
            con.execute("INSERT INTO official_documents(document_type,title,description,file_path,version_label,adoption_date,effective_date,public,active,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                        ("statuts", title, request.form.get("description", ""), file_path, request.form.get("version_label", ""), request.form.get("adoption_date", ""), request.form.get("effective_date", ""), 1 if request.form.get("public") == "1" else 0, 1, now(), user["id"]))
            con.commit(); log_action(user["id"], "Importation d'une nouvelle version des statuts", "official_document", None); flash("Nouvelle version des statuts importée.", "success")
    document = con.execute("SELECT * FROM official_documents WHERE document_type='statuts' AND active=1 AND deleted_at IS NULL ORDER BY id DESC LIMIT 1").fetchone()
    versions = con.execute("SELECT * FROM official_documents WHERE document_type='statuts' ORDER BY id DESC").fetchall()
    sections = con.execute("SELECT * FROM statute_sections ORDER BY sort_order,id").fetchall()
    con.close()
    return render_template("statutes.html", document=document, versions=versions, sections=sections, admin_mode=True)


@app.route("/admin/statuts/sections/<int:section_id>/update", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def update_statute_section(section_id):
    user=current_user(); con=db()
    con.execute("UPDATE statute_sections SET title=?,content=?,sort_order=?,public=?,active=?,updated_at=?,updated_by=? WHERE id=?",
                (request.form.get("title", ""), request.form.get("content", ""), request.form.get("sort_order") or 100, 1 if request.form.get("public") == "1" else 0, 1 if request.form.get("active") == "1" else 0, now(), user["id"], section_id))
    con.commit(); con.close(); log_action(user["id"], "Modification rubrique statutaire", "statute_section", section_id); flash("Rubrique mise à jour.", "success")
    return redirect(url_for("manage_statutes"))


@app.route("/admin/statuts/sections/add", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def add_statute_section():
    user=current_user(); con=db()
    skey="custom_"+uuid.uuid4().hex[:12]
    con.execute("INSERT INTO statute_sections(section_key,title,content,sort_order,public,active,updated_at,updated_by) VALUES(?,?,?,?,?,?,?,?)",
                (skey, request.form.get("title", "Nouvelle rubrique"), request.form.get("content", ""), request.form.get("sort_order") or 100, 1, 1, now(), user["id"]))
    con.commit(); con.close(); flash("Rubrique ajoutée.", "success")
    return redirect(url_for("manage_statutes"))


@app.route("/admin/statuts/sections/<int:section_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def delete_statute_section(section_id):
    user=current_user(); con=db(); con.execute("DELETE FROM statute_sections WHERE id=?", (section_id,)); con.commit(); con.close(); log_action(user["id"], "Suppression rubrique statutaire", "statute_section", section_id); flash("Rubrique supprimée.", "warning")
    return redirect(url_for("manage_statutes"))


@app.route("/admin/statuts/document/<int:document_id>/activate", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def activate_statute_document(document_id):
    con=db(); con.execute("UPDATE official_documents SET active=0 WHERE document_type='statuts'"); con.execute("UPDATE official_documents SET active=1,deleted_at=NULL WHERE id=?", (document_id,)); con.commit(); con.close(); flash("Version restaurée et activée.", "success"); return redirect(url_for("manage_statutes"))


@app.route("/admin/statuts/document/<int:document_id>/delete", methods=["POST"])
@login_required
@role_required("super_admin", "president", "secretary", "national_secretary")
def delete_statute_document(document_id):
    con=db(); con.execute("UPDATE official_documents SET active=0,deleted_at=? WHERE id=?", (now(), document_id)); con.commit(); con.close(); flash("Version retirée de l'application.", "warning"); return redirect(url_for("manage_statutes"))


@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
@role_required("super_admin")
def settings_page():
    con = db()
    if request.method == "POST":
        keys = ["structure_name", "structure_motto", "structure_header", "structure_foundation", "structure_legal", "secretariat_label", "headquarters", "contact_phones", "history", "mission", "vision", "values", "objectives", "advantages", "partners", "president_name", "secretary_name", "facebook", "youtube", "whatsapp", "instagram", "tiktok", "x_twitter", "linkedin", "payment_info", "card_notice", "public_base_url", "public_communiques", "dashboard_message", "footer_note", "initiator", "stability_center_text", "privacy_policy", "terms_of_use", "support_intro", "mobile_app_name", "structure_address", "default_language", "global_search_placeholder", "ai_help_intro", "statute_intro", "windows_client_download_url", "windows_server_download_url", "android_download_url", "download_section_enabled"]
        for key in keys:
            con.execute("UPDATE settings SET value=? WHERE key=?", (request.form.get(key, ""), key))
        logo_paths = save_logo_upload(request.files.get("logo"))
        psig = save_upload(request.files.get("president_signature"), "signatures")
        ssig = save_upload(request.files.get("secretary_signature"), "signatures")
        stamp = save_upload(request.files.get("official_stamp"), "signatures")
        for key, value in logo_paths.items():
            con.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
        if psig: con.execute("UPDATE settings SET value=? WHERE key='president_signature_path'", (psig,))
        if ssig: con.execute("UPDATE settings SET value=? WHERE key='secretary_signature_path'", (ssig,))
        if stamp: con.execute("UPDATE settings SET value=? WHERE key='official_stamp_path'", (stamp,))
        con.execute("UPDATE settings SET value=? WHERE key='stamp_application_mode'", (request.form.get("stamp_application_mode", "validated"),))
        con.commit(); flash("Paramètres de la structure mis à jour.", "success")
    videos = con.execute("SELECT * FROM videos ORDER BY created_at DESC").fetchall()
    carousel = con.execute("SELECT * FROM carousel_images ORDER BY created_at DESC").fetchall()
    fields = con.execute("SELECT * FROM adhesion_fields ORDER BY sort_order, id").fetchall()
    con.close()
    return render_template("settings.html", videos=videos, carousel=carousel, fields=fields)



# ===== V35 : droits par module, services et santé =====
PERMISSION_MODULES = {
    'dashboard':'Tableau de bord','membres':'Membres','adhesions':'Adhésions','cotisations':'Cotisations',
    'tresorerie':'Trésorerie','alertes':'Alertes','activites':'Activités','projets':'Projets','documents':'Documents',
    'statuts':'Statuts et règlement','notifications':'Notifications','support':'Support / tickets','anniversaires':'Anniversaires',
    'utilisateurs':'Utilisateurs','roles':'Rôles et permissions','cloture':'Clôture d’exercice','production':'Production et sécurité',
    'synchronisation':'Synchronisation','corbeille':'Corbeille','audit':'Journal d’audit','diagnostic':'Diagnostic général',
    'assistant_demarrage':'Assistant de démarrage','parametres':'Paramètres','rapports':'Rapports','services':'Services de la Fondation',
    'structures_services':'Structures des services','roles_services':'Rôles des services','affectations_services':'Affectations des services',
    'centres_sante':'Centres de santé','personnel_sante':'Personnel sanitaire','patients':'Patients','rendez_vous':'Rendez-vous',
    'files_attente':'File d’attente','triage':'Triage','consultations':'Consultations','dossiers_medicaux':'Dossiers médicaux',
    'laboratoire':'Laboratoire','pharmacie':'Pharmacie','stocks':'Stocks','hospitalisation':'Hospitalisation','maternite':'Maternité',
    'vaccination':'Vaccination','facturation':'Facturation','caisse_sante':'Caisse sanitaire','aide_sociale':'Prise en charge sociale',
    'equipements':'Équipements','fournisseurs':'Fournisseurs','references':'Références et transferts','rapports_sante':'Rapports sanitaires','statistiques':'Statistiques',
    'qualite_donnees':'Qualité des données','modeles_fiches':'Modèles de fiches','profil':'Profil','aide':'Aide intelligente',
    'guide':'Guide rapide','site_public':'Site public','numerotation_cadres':'Numérotation des hauts cadres','cartes_sanitaires':'Cartes du personnel sanitaire'
}
PERMISSION_ACTIONS = {'voir':'Voir','ajouter':'Ajouter','modifier':'Modifier','supprimer':'Supprimer','valider':'Valider','imprimer':'Imprimer','exporter':'Exporter','parametrer':'Paramétrer'}

def module_allowed(module, action='voir', default=False):
    return role_permission_allowed(current_user(), f'{module}.{action}', default)

def module_permission_required(module, action='voir'):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not module_allowed(module, action, False):
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

def health_scope_clause(user, alias=''):
    prefix=(alias+'.') if alias else ''
    if user and user['role'] in PROVINCIAL_ROLES:
        return f' WHERE {prefix}province=?', [user['province'] or '']
    return ' WHERE 1=1', []

def _next_code(prefix, table):
    con=db(); n=con.execute(f'SELECT COUNT(*) AS n FROM {table}').fetchone()['n']+1; con.close()
    return f'{prefix}-{datetime.now().year}-{n:06d}'

@app.route('/admin/services', methods=['GET','POST'])
@login_required
@module_permission_required('services','voir')
def foundation_services_page():
    con=db()
    if request.method=='POST':
        if not module_allowed('services','ajouter'): abort(403)
        name=request.form.get('name','').strip(); code=request.form.get('code','').strip().upper()
        if not name or not code:
            flash('Le code et le nom sont obligatoires.','warning')
        else:
            try:
                con.execute('''INSERT INTO foundation_services(code,name,description,scope,icon,color,manager_member_id,active,created_at,created_by)
                               VALUES(?,?,?,?,?,?,?,1,?,?)''',
                            (code,name,request.form.get('description',''),request.form.get('scope','national'),request.form.get('icon','🏛️'),request.form.get('color','#0f766e'),request.form.get('manager_member_id') or None,now(),current_user()['id']))
                con.commit(); flash('Service ajouté et architecture réservée.','success')
            except sqlite3.IntegrityError: flash('Ce code de service existe déjà.','danger')
    rows=con.execute('''SELECT s.*, (SELECT COUNT(*) FROM service_structures st WHERE st.service_id=s.id) structure_count,
                        (SELECT COUNT(*) FROM service_roles r WHERE r.service_id=s.id) role_count,
                        (SELECT COUNT(*) FROM service_modules m WHERE m.service_id=s.id) module_count
                        FROM foundation_services s ORDER BY s.name''').fetchall()
    members=con.execute("SELECT id, first_name, last_name, code AS member_code FROM members WHERE status IN ('actif','active') ORDER BY last_name,first_name").fetchall()
    con.close()
    return render_template('foundation_services.html', rows=rows, members=members)

@app.route('/admin/services/<int:service_id>', methods=['GET','POST'])
@login_required
@module_permission_required('services','voir')
def foundation_service_detail(service_id):
    con=db(); service=con.execute('SELECT * FROM foundation_services WHERE id=?',(service_id,)).fetchone()
    if not service: con.close(); abort(404)
    if request.method=='POST':
        action=request.form.get('action','')
        if action=='structure':
            if not module_allowed('services','ajouter'): abort(403)
            code=request.form.get('code','').strip().upper(); name=request.form.get('name','').strip()
            if not code or not name: flash('Code et nom de la structure obligatoires.','warning')
            else:
                try:
                    con.execute('''INSERT INTO service_structures(service_id,parent_id,code,name,structure_type,province,territory,commune,address,phone,email,manager_member_id,active,created_at,created_by)
                                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,1,?,?)''',(service_id,request.form.get('parent_id') or None,code,name,request.form.get('structure_type','antenne'),request.form.get('province',''),request.form.get('territory',''),request.form.get('commune',''),request.form.get('address',''),request.form.get('phone',''),request.form.get('email',''),request.form.get('manager_member_id') or None,now(),current_user()['id']))
                    con.commit(); flash('Structure ajoutée au service.','success')
                except sqlite3.IntegrityError: flash('Ce code de structure existe déjà.','danger')
        elif action=='role':
            if not module_allowed('services','parametrer'): abort(403)
            name=request.form.get('name','').strip()
            if name:
                try:
                    con.execute('INSERT INTO service_roles(service_id,name,description,level,active,created_at,created_by) VALUES(?,?,?,?,1,?,?)',(service_id,name,request.form.get('description',''),request.form.get('level','structure'),now(),current_user()['id']))
                    con.commit(); flash('Rôle ajouté au service.','success')
                except sqlite3.IntegrityError: flash('Ce rôle existe déjà pour ce service.','warning')
        elif action=='module':
            if not module_allowed('services','parametrer'): abort(403)
            code=request.form.get('code','').strip().upper(); name=request.form.get('name','').strip()
            if code and name:
                try:
                    con.execute('INSERT INTO service_modules(service_id,code,name,description,active,created_at,created_by) VALUES(?,?,?,?,1,?,?)',(service_id,code,name,request.form.get('description',''),now(),current_user()['id']))
                    con.commit(); flash('Module réservé pour ce service.','success')
                except sqlite3.IntegrityError: flash('Ce module existe déjà.','warning')
        elif action=='assignment':
            if not module_allowed('services','ajouter'): abort(403)
            member_id=request.form.get('member_id'); role_id=request.form.get('role_id'); structure_id=request.form.get('structure_id') or None
            member=con.execute("SELECT id FROM members WHERE id=? AND status IN ('actif','active')",(member_id,)).fetchone()
            if not member: flash("L'affectation exige un membre actif de la Fondation.",'danger')
            elif role_id:
                try:
                    con.execute('''INSERT INTO service_staff_assignments(member_id,service_id,structure_id,role_id,start_date,status,notes,created_at,created_by)
                                   VALUES(?,?,?,?,?,'active',?,?,?)''',(member_id,service_id,structure_id,role_id,request.form.get('start_date',''),request.form.get('notes',''),now(),current_user()['id']))
                    con.commit(); flash('Membre affecté au service.','success')
                except sqlite3.IntegrityError: flash('Cette affectation existe déjà.','warning')
    structures=con.execute('''SELECT st.*, m.first_name manager_first_name,m.last_name manager_last_name
                              FROM service_structures st LEFT JOIN members m ON m.id=st.manager_member_id
                              WHERE st.service_id=? ORDER BY st.province,st.name''',(service_id,)).fetchall()
    roles=con.execute('SELECT * FROM service_roles WHERE service_id=? ORDER BY level,name',(service_id,)).fetchall()
    modules=con.execute('SELECT * FROM service_modules WHERE service_id=? ORDER BY name',(service_id,)).fetchall()
    assignments=con.execute('''SELECT a.*,mb.first_name,mb.last_name,mb.code AS member_code,r.name role_name,st.name structure_name
                               FROM service_staff_assignments a JOIN members mb ON mb.id=a.member_id
                               JOIN service_roles r ON r.id=a.role_id LEFT JOIN service_structures st ON st.id=a.structure_id
                               WHERE a.service_id=? ORDER BY mb.last_name,mb.first_name''',(service_id,)).fetchall()
    members=con.execute("SELECT id,first_name,last_name,code AS member_code FROM members WHERE status IN ('actif','active') ORDER BY last_name,first_name").fetchall()
    con.close()
    return render_template('foundation_service_detail.html',service=service,structures=structures,roles=roles,modules=modules,assignments=assignments,members=members)

@app.post('/admin/services/<int:service_id>/toggle')
@login_required
@module_permission_required('services','modifier')
def foundation_service_toggle(service_id):
    con=db(); row=con.execute('SELECT active FROM foundation_services WHERE id=?',(service_id,)).fetchone()
    if not row: con.close(); abort(404)
    con.execute('UPDATE foundation_services SET active=?,updated_at=?,updated_by=? WHERE id=?',(0 if row['active'] else 1,now(),current_user()['id'],service_id)); con.commit(); con.close()
    flash('Statut du service mis à jour.','success'); return redirect(url_for('foundation_services_page'))

@app.route('/health')
@login_required
@module_permission_required('centres_sante','voir')
def health_dashboard():
    con=db(); user=current_user(); where,params=health_scope_clause(user)
    centers=con.execute('SELECT COUNT(*) n FROM health_centers'+where,params).fetchone()['n']
    if user['role'] in PROVINCIAL_ROLES:
        center_ids=[r['id'] for r in con.execute('SELECT id FROM health_centers WHERE province=?',(user['province'],)).fetchall()]
        if center_ids:
            marks=','.join('?'*len(center_ids)); patients=con.execute(f'SELECT COUNT(*) n FROM patients WHERE center_id IN ({marks})',center_ids).fetchone()['n']; consultations=con.execute(f'SELECT COUNT(*) n FROM consultations WHERE center_id IN ({marks})',center_ids).fetchone()['n']
        else: patients=consultations=0
    else:
        patients=con.execute('SELECT COUNT(*) n FROM patients').fetchone()['n']; consultations=con.execute('SELECT COUNT(*) n FROM consultations').fetchone()['n']
    staff=con.execute("SELECT COUNT(*) n FROM health_staff WHERE status='active'").fetchone()['n']; con.close()
    return render_template('health_dashboard.html', centers=centers, patients=patients, consultations=consultations, staff=staff)

@app.route('/health/centers', methods=['GET','POST'])
@login_required
@module_permission_required('centres_sante','voir')
def health_centers_page():
    con=db(); user=current_user()
    if request.method=='POST':
        if not module_allowed('centres_sante','ajouter'): abort(403)
        province=request.form.get('province','').strip(); name=request.form.get('name','').strip()
        if user['role'] in PROVINCIAL_ROLES: province=user['province'] or ''
        if not province or not name: flash('Nom et province obligatoires.','warning')
        else:
            prefix=''.join([w[:3].upper() for w in province.split('-')])[:4] or 'RDC'
            code=request.form.get('code','').strip().upper() or f'FOBAK-CS-{prefix}-{secrets.randbelow(900)+100}'
            try:
                center_logo = save_upload(request.files.get('logo'), 'health_centers')
                con.execute('''INSERT INTO health_centers(
                    code,name,province,territory,commune,address,phone,email,opening_date,
                    header_title,header_subtitle,header_address,header_contact,logo_path,footer_note,
                    active,created_at,created_by
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?)''', (
                    code,name,province,request.form.get('territory',''),request.form.get('commune',''),
                    request.form.get('address',''),request.form.get('phone',''),request.form.get('email',''),request.form.get('opening_date',''),
                    request.form.get('header_title','').strip(),request.form.get('header_subtitle','').strip(),
                    request.form.get('header_address','').strip(),request.form.get('header_contact','').strip(),
                    center_logo,request.form.get('footer_note','').strip(),now(),user['id']))
                con.commit(); flash('Centre de santé ajouté.','success')
            except sqlite3.IntegrityError:
                con.rollback(); flash('Le code du centre existe déjà.','danger')
            except Exception as exc:
                con.rollback(); logging.exception('Erreur création centre de santé: %s', exc); flash('Impossible d’enregistrer le centre. Vérifiez le logo et les champs puis réessayez.','danger')
    where,params=health_scope_clause(user)
    query='SELECT * FROM health_centers'+where+(' AND deleted_at IS NULL' if 'WHERE' in where else ' WHERE deleted_at IS NULL')+' ORDER BY province,name'; rows=con.execute(query,params).fetchall(); con.close()
    return render_template('health_centers.html',rows=rows)

@app.route('/health/roles', methods=['GET','POST'])
@login_required
@module_permission_required('personnel_sante','voir')
def health_roles_page():
    con=db()
    if request.method=='POST':
        if not module_allowed('personnel_sante','parametrer'): abort(403)
        name=request.form.get('name','').strip()
        if name:
            try: con.execute('INSERT INTO health_roles(name,description,level,active,created_at,created_by) VALUES(?,?,?,?,?,?)',(name,request.form.get('description',''),request.form.get('level','centre'),1,now(),current_user()['id'])); con.commit(); flash('Rôle sanitaire ajouté.','success')
            except sqlite3.IntegrityError: flash('Ce rôle existe déjà.','warning')
    rows=con.execute('SELECT * FROM health_roles WHERE deleted_at IS NULL ORDER BY level,name').fetchall(); con.close()
    return render_template('health_roles.html',rows=rows)

@app.route('/health/staff', methods=['GET','POST'])
@login_required
@module_permission_required('personnel_sante','voir')
def health_staff_page():
    con=db(); user=current_user()
    if request.method=='POST':
        if not module_allowed('personnel_sante','ajouter'): abort(403)
        member_id=request.form.get('member_id',type=int); center_id=request.form.get('center_id',type=int); role_id=request.form.get('role_id',type=int)
        member=con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL AND status='active'",(member_id,)).fetchone()
        center=con.execute('SELECT * FROM health_centers WHERE id=? AND active=1',(center_id,)).fetchone()
        if not member: flash("Affectation refusée : la personne doit d'abord être un membre actif de la Fondation.",'danger')
        elif not center or (user['role'] in PROVINCIAL_ROLES and center['province']!=(user['province'] or '')): abort(403)
        elif not role_id: flash('Choisissez un rôle sanitaire.','warning')
        else:
            try:
                con.execute('INSERT INTO health_staff(member_id,center_id,role_id,specialty,professional_number,diploma,engagement_date,status,created_at,created_by) VALUES(?,?,?,?,?,?,?,\'active\',?,?)',(member_id,center_id,role_id,request.form.get('specialty',''),request.form.get('professional_number',''),request.form.get('diploma',''),request.form.get('engagement_date',''),now(),user['id']))
                con.commit(); flash('Membre affecté au centre de santé.','success')
            except sqlite3.IntegrityError: flash('Cette affectation existe déjà.','warning')
    where,params=health_scope_clause(user,'c')
    rows=con.execute('''SELECT hs.*,m.first_name,m.last_name,m.code member_code,c.name center_name,c.province,hr.name role_name FROM health_staff hs JOIN members m ON m.id=hs.member_id JOIN health_centers c ON c.id=hs.center_id JOIN health_roles hr ON hr.id=hs.role_id'''+where+' ORDER BY c.province,c.name,m.last_name',params).fetchall()
    centers=con.execute('SELECT * FROM health_centers'+health_scope_clause(user)[0]+' ORDER BY name',health_scope_clause(user)[1]).fetchall()
    members=con.execute("SELECT id,code,first_name,last_name,province FROM members WHERE deleted_at IS NULL AND status='active' ORDER BY last_name,first_name LIMIT 1000").fetchall()
    roles=con.execute('SELECT * FROM health_roles WHERE active=1 ORDER BY name').fetchall(); con.close()
    return render_template('health_staff.html',rows=rows,centers=centers,members=members,roles=roles)

@app.route('/health/patients', methods=['GET','POST'])
@login_required
@module_permission_required('patients','voir')
def patients_page():
    con=db(); user=current_user()
    if request.method=='POST':
        if not module_allowed('patients','ajouter'): abort(403)
        center_id=request.form.get('center_id',type=int); center=con.execute('SELECT * FROM health_centers WHERE id=?',(center_id,)).fetchone()
        if not center or (user['role'] in PROVINCIAL_ROLES and center['province']!=(user['province'] or '')): abort(403)
        code=_next_code('PAT-'+(center['province'][:3].upper()),'patients')
        con.execute('''INSERT INTO patients(patient_code,center_id,first_name,last_name,gender,birth_date,phone,address,province,emergency_contact,blood_group,allergies,medical_history,created_at,created_by,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',(code,center_id,request.form.get('first_name','').strip(),request.form.get('last_name','').strip(),request.form.get('gender',''),request.form.get('birth_date',''),request.form.get('phone',''),request.form.get('address',''),center['province'],request.form.get('emergency_contact',''),request.form.get('blood_group',''),request.form.get('allergies',''),request.form.get('medical_history',''),now(),user['id'],now()))
        con.commit(); flash('Patient enregistré.','success')
    if user['role'] in PROVINCIAL_ROLES:
        rows=con.execute('''SELECT p.*,c.name center_name FROM patients p JOIN health_centers c ON c.id=p.center_id WHERE c.province=? AND p.deleted_at IS NULL ORDER BY p.id DESC''',(user['province'],)).fetchall(); centers=con.execute('SELECT * FROM health_centers WHERE province=? AND active=1 ORDER BY name',(user['province'],)).fetchall()
    else:
        rows=con.execute('''SELECT p.*,c.name center_name FROM patients p JOIN health_centers c ON c.id=p.center_id WHERE p.deleted_at IS NULL ORDER BY p.id DESC''').fetchall(); centers=con.execute('SELECT * FROM health_centers WHERE active=1 ORDER BY province,name').fetchall()
    con.close(); return render_template('patients.html',rows=rows,centers=centers)

@app.route('/health/consultations', methods=['GET','POST'])
@login_required
@module_permission_required('consultations','voir')
def consultations_page():
    con=db(); user=current_user()
    if request.method=='POST':
        if not module_allowed('consultations','ajouter'): abort(403)
        patient_id=request.form.get('patient_id',type=int); patient=con.execute('SELECT p.*,c.province FROM patients p JOIN health_centers c ON c.id=p.center_id WHERE p.id=?',(patient_id,)).fetchone()
        if not patient or (user['role'] in PROVINCIAL_ROLES and patient['province']!=(user['province'] or '')): abort(403)
        code=_next_code('CONS','consultations')
        cols=['reason','complaints','disease_history','temperature','blood_pressure','heart_rate','respiratory_rate','weight','height','clinical_exam','diagnosis','requested_tests','treatment','orientation','next_appointment','notes']
        vals=[request.form.get(c,'') for c in cols]
        con.execute('''INSERT INTO consultations(consultation_code,patient_id,center_id,staff_id,consultation_date,reason,complaints,disease_history,temperature,blood_pressure,heart_rate,respiratory_rate,weight,height,clinical_exam,diagnosis,requested_tests,treatment,orientation,next_appointment,notes,status,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'draft',?,?)''',[code,patient_id,patient['center_id'],request.form.get('staff_id',type=int),request.form.get('consultation_date') or today(),*vals,now(),user['id']])
        con.commit(); flash('Consultation enregistrée.','success')
    if user['role'] in PROVINCIAL_ROLES:
        rows=con.execute('''SELECT co.*,p.patient_code,p.first_name,p.last_name,c.name center_name FROM consultations co JOIN patients p ON p.id=co.patient_id JOIN health_centers c ON c.id=co.center_id WHERE c.province=? AND co.deleted_at IS NULL ORDER BY co.id DESC''',(user['province'],)).fetchall(); patients=con.execute('''SELECT p.* FROM patients p JOIN health_centers c ON c.id=p.center_id WHERE c.province=? ORDER BY p.last_name''',(user['province'],)).fetchall(); staff=con.execute('''SELECT hs.id,m.first_name,m.last_name,hr.name role_name FROM health_staff hs JOIN members m ON m.id=hs.member_id JOIN health_roles hr ON hr.id=hs.role_id JOIN health_centers c ON c.id=hs.center_id WHERE c.province=? AND hs.status='active' ORDER BY m.last_name''',(user['province'],)).fetchall()
    else:
        rows=con.execute('''SELECT co.*,p.patient_code,p.first_name,p.last_name,c.name center_name FROM consultations co JOIN patients p ON p.id=co.patient_id JOIN health_centers c ON c.id=co.center_id WHERE co.deleted_at IS NULL ORDER BY co.id DESC''').fetchall(); patients=con.execute('SELECT * FROM patients ORDER BY last_name').fetchall(); staff=con.execute('''SELECT hs.id,m.first_name,m.last_name,hr.name role_name FROM health_staff hs JOIN members m ON m.id=hs.member_id JOIN health_roles hr ON hr.id=hs.role_id WHERE hs.status='active' AND hs.deleted_at IS NULL ORDER BY m.last_name''').fetchall()
    con.close(); return render_template('consultations.html',rows=rows,patients=patients,staff=staff)

@app.route('/health/forms', methods=['GET','POST'])
@login_required
@module_permission_required('modeles_fiches','voir')
def medical_forms_page():
    con=db(); user=current_user()
    if request.method=='POST':
        if not module_allowed('modeles_fiches','ajouter') and not module_allowed('modeles_fiches','parametrer'): abort(403)
        uploaded=request.files.get('source_file'); source=''; status='ready'; fields=[]
        raw=request.form.get('fields_json','').strip()
        if raw:
            try: fields=json.loads(raw)
            except Exception: flash('Le JSON des champs est invalide.','danger'); fields=[]
        if uploaded and uploaded.filename:
            source=save_upload(uploaded,'medical_forms') or ''
            ext=uploaded.filename.rsplit('.',1)[-1].lower() if '.' in uploaded.filename else ''
            if ext=='json' and source:
                try:
                    fields=json.loads(Path(os.path.join(STATIC_DIR,source)).read_text(encoding='utf-8'))
                    status='ready'
                except Exception: status='review'
            elif ext in {'pdf','doc','docx','jpg','jpeg','png'}: status='review'
        con.execute('''INSERT INTO medical_form_templates(name,form_type,scope,province,center_id,fields_json,source_file,import_status,active,version,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,1,1,?,?)''',(request.form.get('name','').strip(),request.form.get('form_type','patient'),request.form.get('scope','national'),request.form.get('province',''),request.form.get('center_id',type=int),json.dumps(fields,ensure_ascii=False),source,status,now(),user['id']))
        con.commit(); flash('Modèle importé. Vérifiez les champs avant utilisation.' if status=='review' else 'Modèle enregistré.','success')
    rows=con.execute('SELECT mf.*,hc.name center_name FROM medical_form_templates mf LEFT JOIN health_centers hc ON hc.id=mf.center_id WHERE mf.deleted_at IS NULL ORDER BY mf.id DESC').fetchall(); centers=con.execute('SELECT * FROM health_centers WHERE deleted_at IS NULL ORDER BY province,name').fetchall(); con.close()
    return render_template('medical_forms.html',rows=rows,centers=centers)




def _health_scope_clause(alias='hc'):
    user=current_user()
    if user and user['role'] in PROVINCIAL_ROLES:
        return f" WHERE {alias}.province=?", [user['province'] or '']
    return '', []

@app.route('/health/operations', methods=['GET','POST'])
@login_required
@module_permission_required('consultations','voir')
def health_operations_page():
    con=db(); user=current_user(); action=request.form.get('action','') if request.method=='POST' else ''
    if request.method=='POST':
        center_id=request.form.get('center_id',type=int)
        center=con.execute('SELECT * FROM health_centers WHERE id=?',(center_id,)).fetchone() if center_id else None
        if not center or (user['role'] in PROVINCIAL_ROLES and center['province']!=(user['province'] or '')): abort(403)
        patient_id=request.form.get('patient_id',type=int)
        if action=='appointment':
            code=_next_code('RDV','health_appointments'); con.execute('INSERT INTO health_appointments(code,patient_id,center_id,service_name,professional_id,appointment_date,appointment_time,reason,status,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(code,patient_id,center_id,request.form.get('service_name',''),request.form.get('professional_id',type=int),request.form.get('appointment_date') or today(),request.form.get('appointment_time',''),request.form.get('reason',''),'scheduled',now(),user['id']))
            flash('Rendez-vous enregistré.','success')
        elif action=='visit':
            code=_next_code('VIS','health_visits'); q=con.execute('SELECT COALESCE(MAX(queue_number),0)+1 n FROM health_visits WHERE center_id=? AND visit_date=?',(center_id,today())).fetchone()['n']; con.execute('INSERT INTO health_visits(code,patient_id,center_id,visit_date,arrival_time,service_name,priority,queue_number,status,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(code,patient_id,center_id,today(),request.form.get('arrival_time',''),request.form.get('service_name','Consultation'),request.form.get('priority','normal'),q,'waiting',now(),user['id']))
            flash(f'Visite ouverte, numéro de file {q}.','success')
        elif action=='triage':
            visit_id=request.form.get('visit_id',type=int); visit=con.execute('SELECT * FROM health_visits WHERE id=?',(visit_id,)).fetchone()
            if not visit or visit['center_id']!=center_id: abort(400)
            temp=request.form.get('temperature',type=float); sat=request.form.get('oxygen_saturation',type=float); pain=request.form.get('pain_score',type=int)
            level='urgent' if (temp and temp>=39) or (sat and sat<92) or (pain and pain>=8) or request.form.get('danger_signs','').strip() else 'normal'
            con.execute('INSERT OR REPLACE INTO health_triage(visit_id,patient_id,center_id,temperature,systolic,diastolic,heart_rate,respiratory_rate,oxygen_saturation,weight,height,glucose,pain_score,consciousness,danger_signs,pregnancy_status,triage_level,notes,recorded_at,recorded_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(visit_id,visit['patient_id'],center_id,temp,request.form.get('systolic',type=int),request.form.get('diastolic',type=int),request.form.get('heart_rate',type=int),request.form.get('respiratory_rate',type=int),sat,request.form.get('weight',type=float),request.form.get('height',type=float),request.form.get('glucose',type=float),pain,request.form.get('consciousness',''),request.form.get('danger_signs',''),request.form.get('pregnancy_status',''),level,request.form.get('notes',''),now(),user['id']))
            con.execute("UPDATE health_visits SET status='triaged' WHERE id=?",(visit_id,)); flash('Triage enregistré.','success')
        elif action=='lab':
            code=_next_code('LAB','health_lab_orders'); con.execute('INSERT INTO health_lab_orders(code,patient_id,consultation_id,center_id,test_name,sample_type,priority,status,requested_at,requested_by) VALUES(?,?,?,?,?,?,?,?,?,?)',(code,patient_id,request.form.get('consultation_id',type=int),center_id,request.form.get('test_name',''),request.form.get('sample_type',''),request.form.get('priority','normal'),'ordered',now(),user['id'])); flash('Examen de laboratoire demandé.','success')
        elif action=='lab_result':
            order_id=request.form.get('order_id',type=int); con.execute("UPDATE health_lab_orders SET result_text=?,reference_range=?,status='validated',validated_at=?,validated_by=? WHERE id=? AND center_id=?",(request.form.get('result_text',''),request.form.get('reference_range',''),now(),user['id'],order_id,center_id)); flash('Résultat validé.','success')
        elif action=='product':
            code=request.form.get('product_code','').strip() or _next_code('MED','health_products'); con.execute('INSERT OR IGNORE INTO health_products(center_id,code,name,category,unit,minimum_stock,sale_price,active) VALUES(?,?,?,?,?,?,?,1)',(center_id,code,request.form.get('product_name',''),request.form.get('category','medicament'),request.form.get('unit',''),request.form.get('minimum_stock',type=float) or 0,request.form.get('sale_price',type=float) or 0)); flash('Produit enregistré.','success')
        elif action=='batch':
            con.execute('INSERT INTO health_stock_batches(product_id,batch_number,expiry_date,quantity,purchase_price,supplier_id,received_at,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?)',(request.form.get('product_id',type=int),request.form.get('batch_number',''),request.form.get('expiry_date',''),request.form.get('quantity',type=float) or 0,request.form.get('purchase_price',type=float) or 0,request.form.get('supplier_id',type=int),today(),now(),user['id'])); flash('Lot ajouté au stock.','success')
        elif action=='admission':
            code=_next_code('HOSP','health_admissions'); con.execute('INSERT INTO health_admissions(code,patient_id,center_id,ward,room,bed,admitted_at,diagnosis,responsible_staff_id,status,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(code,patient_id,center_id,request.form.get('ward',''),request.form.get('room',''),request.form.get('bed',''),now(),request.form.get('diagnosis',''),request.form.get('professional_id',type=int),'admitted',user['id'])); flash('Hospitalisation enregistrée.','success')
        elif action=='invoice':
            qty=request.form.get('quantity',type=float) or 1; price=request.form.get('unit_price',type=float) or 0; discount=request.form.get('discount',type=float) or 0; support=request.form.get('social_support',type=float) or 0; total=max(0,qty*price-discount-support); code=_next_code('FAC','health_invoices'); cur=con.execute('INSERT INTO health_invoices(code,patient_id,center_id,invoice_date,subtotal,discount,social_support,total,paid,status,payment_method,notes,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(code,patient_id,center_id,today(),qty*price,discount,support,total,request.form.get('paid',type=float) or 0,'paid' if (request.form.get('paid',type=float) or 0)>=total else 'partial',request.form.get('payment_method',''),request.form.get('notes',''),now(),user['id'])); con.execute('INSERT INTO health_invoice_items(invoice_id,item_type,description,quantity,unit_price,total) VALUES(?,?,?,?,?,?)',(cur.lastrowid,request.form.get('item_type','acte'),request.form.get('description',''),qty,price,qty*price)); flash('Facture créée.','success')
        elif action=='support':
            con.execute('INSERT INTO health_social_support(patient_id,center_id,category,reason,requested_amount,approved_amount,status,valid_until,approved_by,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?)',(patient_id,center_id,request.form.get('category','Vulnérabilité'),request.form.get('reason',''),request.form.get('requested_amount',type=float) or 0,request.form.get('approved_amount',type=float) or 0,'approved' if request.form.get('approved_amount') else 'pending',request.form.get('valid_until',''),user['id'] if request.form.get('approved_amount') else None,now(),user['id'])); flash('Prise en charge sociale enregistrée.','success')
        elif action=='equipment':
            code=request.form.get('inventory_code','').strip() or _next_code('EQP','health_equipment'); con.execute('INSERT INTO health_equipment(center_id,inventory_code,name,category,location,state,acquisition_date,warranty_end,next_maintenance,supplier_id,notes,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',(center_id,code,request.form.get('equipment_name',''),request.form.get('category',''),request.form.get('location',''),request.form.get('state','operational'),request.form.get('acquisition_date',''),request.form.get('warranty_end',''),request.form.get('next_maintenance',''),request.form.get('supplier_id',type=int),request.form.get('notes',''),now(),user['id'])); flash('Équipement enregistré.','success')
        elif action=='supplier':
            con.execute('INSERT INTO health_suppliers(name,phone,email,address,province,active,created_at) VALUES(?,?,?,?,?,1,?)',(request.form.get('supplier_name',''),request.form.get('phone',''),request.form.get('email',''),request.form.get('address',''),center['province'],now())); flash('Fournisseur enregistré.','success')
        elif action=='referral':
            code=_next_code('REF','health_referrals'); con.execute('INSERT INTO health_referrals(code,patient_id,from_center_id,to_center_id,external_destination,urgency,reason,clinical_summary,treatment_given,status,sent_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(code,patient_id,center_id,request.form.get('to_center_id',type=int),request.form.get('external_destination',''),request.form.get('urgency','normal'),request.form.get('reason',''),request.form.get('clinical_summary',''),request.form.get('treatment_given',''),'sent',now(),user['id'])); flash('Référence enregistrée.','success')
        con.commit(); log_action(user['id'],'Opération centre de santé',action,None,center['name']); return redirect(url_for('health_operations_page'))
    where,args=_health_scope_clause('hc')
    centers=con.execute('SELECT hc.* FROM health_centers hc'+where+' ORDER BY hc.province,hc.name',args).fetchall()
    if user['role'] in PROVINCIAL_ROLES:
        patients=con.execute('SELECT p.* FROM patients p JOIN health_centers hc ON hc.id=p.center_id WHERE hc.province=? ORDER BY p.last_name',(user['province'],)).fetchall(); visits=con.execute("SELECT v.*,p.patient_code,p.last_name,p.first_name,hc.name center_name FROM health_visits v JOIN patients p ON p.id=v.patient_id JOIN health_centers hc ON hc.id=v.center_id WHERE hc.province=? ORDER BY v.id DESC LIMIT 100",(user['province'],)).fetchall()
    else:
        patients=con.execute('SELECT * FROM patients ORDER BY last_name').fetchall(); visits=con.execute("SELECT v.*,p.patient_code,p.last_name,p.first_name,hc.name center_name FROM health_visits v JOIN patients p ON p.id=v.patient_id JOIN health_centers hc ON hc.id=v.center_id ORDER BY v.id DESC LIMIT 100").fetchall()
    labs=con.execute('SELECT l.*,p.patient_code,p.last_name,p.first_name FROM health_lab_orders l JOIN patients p ON p.id=l.patient_id ORDER BY l.id DESC LIMIT 100').fetchall(); products=con.execute('SELECT hp.*,COALESCE(SUM(b.quantity),0) stock FROM health_products hp LEFT JOIN health_stock_batches b ON b.product_id=hp.id GROUP BY hp.id ORDER BY hp.name').fetchall(); invoices=con.execute('SELECT i.*,p.patient_code,p.last_name,p.first_name FROM health_invoices i JOIN patients p ON p.id=i.patient_id ORDER BY i.id DESC LIMIT 100').fetchall(); admissions=con.execute('SELECT a.*,p.patient_code,p.last_name,p.first_name FROM health_admissions a JOIN patients p ON p.id=a.patient_id ORDER BY a.id DESC LIMIT 100').fetchall(); suppliers=con.execute('SELECT * FROM health_suppliers WHERE active=1 ORDER BY name').fetchall(); equipment=con.execute('SELECT e.*,hc.name center_name FROM health_equipment e JOIN health_centers hc ON hc.id=e.center_id ORDER BY e.id DESC LIMIT 100').fetchall(); appointments=con.execute('SELECT a.*,p.patient_code,p.last_name,p.first_name FROM health_appointments a JOIN patients p ON p.id=a.patient_id ORDER BY a.appointment_date DESC,a.appointment_time DESC LIMIT 100').fetchall(); referrals=con.execute('SELECT r.*,p.patient_code,p.last_name,p.first_name FROM health_referrals r JOIN patients p ON p.id=r.patient_id ORDER BY r.id DESC LIMIT 100').fetchall()
    con.close(); return render_template('health_operations.html',centers=centers,patients=patients,visits=visits,labs=labs,products=products,invoices=invoices,admissions=admissions,suppliers=suppliers,equipment=equipment,appointments=appointments,referrals=referrals)


def _month_bounds(month):
    try:
        start = datetime.strptime((month or '') + '-01', '%Y-%m-%d')
    except Exception:
        start = datetime.now().replace(day=1)
        month = start.strftime('%Y-%m')
    if start.month == 12:
        end = start.replace(year=start.year + 1, month=1)
    else:
        end = start.replace(month=start.month + 1)
    return month, start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')


def _health_report_scope(user, requested_center=None, requested_province=''):
    center_id = int(requested_center) if str(requested_center or '').isdigit() else None
    province = (requested_province or '').strip()
    if user['role'] in PROVINCIAL_ROLES:
        province = user['province'] or ''
        if center_id:
            con=db(); row=con.execute('SELECT id FROM health_centers WHERE id=? AND province=?',(center_id,province)).fetchone(); con.close()
            if not row: center_id=None
    return center_id, province


def build_health_monthly_report(con, month, center_id=None, province=''):
    month,start,end=_month_bounds(month)
    where=[]; params=[]
    if center_id:
        where.append('hc.id=?'); params.append(center_id)
    elif province:
        where.append('hc.province=?'); params.append(province)
    where_sql=(' WHERE '+ ' AND '.join(where)) if where else ''
    centers=con.execute('SELECT hc.* FROM health_centers hc'+where_sql+' ORDER BY hc.province,hc.name',params).fetchall()
    ids=[r['id'] for r in centers]
    if not ids:
        return {'month':month,'centers':centers,'scope_label':province or 'National','patients':0,'new_patients':0,'consultations':0,'visits':0,'appointments':0,'lab_orders':0,'admissions':0,'referrals':0,'invoice_total':0,'paid_total':0,'debt_total':0,'discount_total':0,'support_total':0,'loss_total':0,'losses':[],'low_stock':0,'expired_batches':0,'equipment_down':0,'staff_active':0,'top_diagnoses':[],'center_breakdown':[]}
    ph=','.join('?'*len(ids))
    def scalar(sql, extra=()):
        row=con.execute(sql, tuple(ids)+tuple(extra)).fetchone(); return (row[0] if row and row[0] is not None else 0)
    patients=scalar(f'SELECT COUNT(*) FROM patients WHERE center_id IN ({ph})')
    new_patients=scalar(f"SELECT COUNT(*) FROM patients WHERE center_id IN ({ph}) AND date(created_at)>=date(?) AND date(created_at)<date(?)",(start,end))
    consultations=scalar(f"SELECT COUNT(*) FROM consultations WHERE center_id IN ({ph}) AND date(consultation_date)>=date(?) AND date(consultation_date)<date(?)",(start,end))
    visits=scalar(f"SELECT COUNT(*) FROM health_visits WHERE center_id IN ({ph}) AND date(visit_date)>=date(?) AND date(visit_date)<date(?)",(start,end))
    appointments=scalar(f"SELECT COUNT(*) FROM health_appointments WHERE center_id IN ({ph}) AND date(appointment_date)>=date(?) AND date(appointment_date)<date(?)",(start,end))
    lab_orders=scalar(f"SELECT COUNT(*) FROM health_lab_orders WHERE center_id IN ({ph}) AND date(requested_at)>=date(?) AND date(requested_at)<date(?)",(start,end))
    admissions=scalar(f"SELECT COUNT(*) FROM health_admissions WHERE center_id IN ({ph}) AND date(admitted_at)>=date(?) AND date(admitted_at)<date(?)",(start,end))
    referrals=scalar(f"SELECT COUNT(*) FROM health_referrals WHERE from_center_id IN ({ph}) AND date(sent_at)>=date(?) AND date(sent_at)<date(?)",(start,end))
    fin=con.execute(f"SELECT COALESCE(SUM(total),0),COALESCE(SUM(paid),0),COALESCE(SUM(total-paid),0),COALESCE(SUM(discount),0),COALESCE(SUM(social_support),0) FROM health_invoices WHERE center_id IN ({ph}) AND date(invoice_date)>=date(?) AND date(invoice_date)<date(?)",tuple(ids)+(start,end)).fetchone()
    losses=con.execute(f"SELECT category,currency,COUNT(*) n,COALESCE(SUM(amount),0) total FROM health_losses WHERE center_id IN ({ph}) AND date(loss_date)>=date(?) AND date(loss_date)<date(?) GROUP BY category,currency ORDER BY total DESC",tuple(ids)+(start,end)).fetchall()
    loss_total=sum(float(r['total'] or 0) for r in losses)
    low_stock=scalar(f"SELECT COUNT(*) FROM (SELECT hp.id,hp.minimum_stock,COALESCE(SUM(b.quantity),0) q FROM health_products hp LEFT JOIN health_stock_batches b ON b.product_id=hp.id WHERE hp.center_id IN ({ph}) GROUP BY hp.id HAVING q<=hp.minimum_stock)")
    expired_batches=scalar(f"SELECT COUNT(*) FROM health_stock_batches b JOIN health_products hp ON hp.id=b.product_id WHERE hp.center_id IN ({ph}) AND b.quantity>0 AND b.expiry_date!='' AND date(b.expiry_date)<date(?)",(end,))
    equipment_down=scalar(f"SELECT COUNT(*) FROM health_equipment WHERE center_id IN ({ph}) AND state NOT IN ('operational','fonctionnel','active')")
    staff_active=scalar(f"SELECT COUNT(*) FROM health_staff WHERE center_id IN ({ph}) AND status='active'")
    top_diagnoses=con.execute(f"SELECT diagnosis,COUNT(*) n FROM consultations WHERE center_id IN ({ph}) AND date(consultation_date)>=date(?) AND date(consultation_date)<date(?) AND diagnosis IS NOT NULL AND trim(diagnosis)!='' GROUP BY diagnosis ORDER BY n DESC LIMIT 10",tuple(ids)+(start,end)).fetchall()
    center_breakdown=[]
    for c in centers:
        cid=c['id']
        row=con.execute("SELECT COUNT(*) consultations FROM consultations WHERE center_id=? AND date(consultation_date)>=date(?) AND date(consultation_date)<date(?)",(cid,start,end)).fetchone()
        f=con.execute("SELECT COALESCE(SUM(total),0) total,COALESCE(SUM(paid),0) paid FROM health_invoices WHERE center_id=? AND date(invoice_date)>=date(?) AND date(invoice_date)<date(?)",(cid,start,end)).fetchone()
        l=con.execute("SELECT COALESCE(SUM(amount),0) total FROM health_losses WHERE center_id=? AND date(loss_date)>=date(?) AND date(loss_date)<date(?)",(cid,start,end)).fetchone()
        center_breakdown.append({'id':cid,'code':c['code'],'name':c['name'],'province':c['province'],'consultations':row['consultations'],'invoiced':f['total'],'paid':f['paid'],'losses':l['total']})
    scope_label=centers[0]['name'] if center_id and centers else (province or 'National')
    return {'month':month,'centers':centers,'scope_label':scope_label,'patients':patients,'new_patients':new_patients,'consultations':consultations,'visits':visits,'appointments':appointments,'lab_orders':lab_orders,'admissions':admissions,'referrals':referrals,'invoice_total':fin[0] or 0,'paid_total':fin[1] or 0,'debt_total':fin[2] or 0,'discount_total':fin[3] or 0,'support_total':fin[4] or 0,'loss_total':loss_total,'losses':losses,'low_stock':low_stock,'expired_batches':expired_batches,'equipment_down':equipment_down,'staff_active':staff_active,'top_diagnoses':top_diagnoses,'center_breakdown':center_breakdown}


def get_health_branding(center=None):
    settings = get_settings()
    title = (center['header_title'] if center and 'header_title' in center.keys() and center['header_title'] else 'CENTRE DE SANTÉ')
    subtitle = (center['header_subtitle'] if center and 'header_subtitle' in center.keys() and center['header_subtitle'] else (center['name'] if center else 'Réseau sanitaire de la Fondation'))
    address = (center['header_address'] if center and 'header_address' in center.keys() and center['header_address'] else (center['address'] if center and 'address' in center.keys() else settings.get('headquarters','')))
    contact = (center['header_contact'] if center and 'header_contact' in center.keys() and center['header_contact'] else ((center['phone'] if center and 'phone' in center.keys() else '') or settings.get('contact_phones','')))
    footer_note = (center['footer_note'] if center and 'footer_note' in center.keys() and center['footer_note'] else 'Document généré automatiquement par FOBAK Manager Pro.')
    left_logo = center['logo_path'] if center and 'logo_path' in center.keys() and center['logo_path'] else None
    return {'title': title, 'subtitle': subtitle, 'address': address, 'contact': contact, 'footer_note': footer_note, 'left_logo': left_logo, 'right_logo': settings.get('logo_path','')}


@app.route('/health/reports', methods=['GET','POST'])
@login_required
@module_permission_required('rapports_sante','voir')
def health_reports_page():
    user=current_user(); month=request.values.get('month') or datetime.now().strftime('%Y-%m')
    center_id,province=_health_report_scope(user,request.values.get('center_id'),request.values.get('province',''))
    con=db()
    if request.method=='POST':
        action=request.form.get('action','save')
        if action=='loss':
            if not module_allowed('rapports_sante','ajouter'): abort(403)
            cid=request.form.get('center_id',type=int)
            con.execute('INSERT INTO health_losses(center_id,loss_date,category,description,quantity,amount,currency,reason,responsible_member_id,status,created_at,created_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)',(cid,request.form.get('loss_date') or today(),request.form.get('category','Autre'),request.form.get('description',''),request.form.get('quantity',type=float) or 0,request.form.get('amount',type=float) or 0,request.form.get('currency','CDF'),request.form.get('reason',''),request.form.get('responsible_member_id',type=int),'declared',now(),user['id']))
            con.commit(); log_action(user['id'],'Déclaration perte centre de santé','health_loss',None,request.form.get('description','')); flash('Perte enregistrée dans le rapport.','success'); con.close(); return redirect(url_for('health_reports_page',month=month,center_id=cid,province=province))
        data=build_health_monthly_report(con,month,center_id,province)
        scope_type='center' if center_id else ('province' if province else 'national')
        db_center=center_id or 0; db_province=province or ''
        snapshot=json.dumps(data,ensure_ascii=False,default=lambda o: dict(o) if hasattr(o,'keys') else str(o))
        vals=(month,scope_type,db_center,db_province,request.form.get('status','draft'),request.form.get('narrative_activities',''),request.form.get('difficulties',''),request.form.get('needs',''),request.form.get('recommendations',''),request.form.get('important_events',''),snapshot,now(),user['id'],now(),user['id'])
        con.execute("INSERT INTO health_monthly_reports(report_month,scope_type,center_id,province,status,narrative_activities,difficulties,needs,recommendations,important_events,snapshot_json,generated_at,generated_by,updated_at,updated_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(report_month,scope_type,center_id,province) DO UPDATE SET status=excluded.status,narrative_activities=excluded.narrative_activities,difficulties=excluded.difficulties,needs=excluded.needs,recommendations=excluded.recommendations,important_events=excluded.important_events,snapshot_json=excluded.snapshot_json,updated_at=excluded.updated_at,updated_by=excluded.updated_by",vals)
        if action=='submit':
            con.execute('UPDATE health_monthly_reports SET status=?,submitted_at=?,submitted_by=? WHERE report_month=? AND scope_type=? AND center_id=? AND province=?',('submitted',now(),user['id'],month,scope_type,db_center,db_province))
            create_internal_notification('Rapport sanitaire soumis',f'{data["scope_label"]} — {month}',url_for('health_reports_page',month=month,center_id=center_id or '',province=province),role='all')
        elif action=='validate':
            if user['role'] not in NATIONAL_ROLES and user['role'] not in PROVINCIAL_ROLES: abort(403)
            con.execute('UPDATE health_monthly_reports SET status=?,validated_at=?,validated_by=? WHERE report_month=? AND scope_type=? AND center_id=? AND province=?',('validated',now(),user['id'],month,scope_type,db_center,db_province))
        con.commit(); flash('Rapport sanitaire enregistré.','success')
    data=build_health_monthly_report(con,month,center_id,province)
    scope_type='center' if center_id else ('province' if province else 'national')
    db_center=center_id or 0; db_province=province or ''
    saved=con.execute('SELECT * FROM health_monthly_reports WHERE report_month=? AND scope_type=? AND center_id=? AND province=?',(month,scope_type,db_center,db_province)).fetchone()
    where,args=_health_scope_clause('hc'); centers=con.execute('SELECT hc.* FROM health_centers hc'+where+' ORDER BY hc.province,hc.name',args).fetchall()
    selected_center = con.execute('SELECT * FROM health_centers WHERE id=?',(center_id,)).fetchone() if center_id else None
    branding = get_health_branding(selected_center)
    report_rows=con.execute("SELECT r.*,hc.name center_name FROM health_monthly_reports r LEFT JOIN health_centers hc ON hc.id=r.center_id ORDER BY r.report_month DESC,r.updated_at DESC LIMIT 60").fetchall()
    members=con.execute("SELECT id,last_name,first_name,code FROM members WHERE deleted_at IS NULL AND status='active' ORDER BY last_name,first_name LIMIT 500").fetchall()
    con.close(); return render_template('health_reports.html',data=data,saved=saved,month=month,center_id=center_id,scope_province=province,centers=centers,report_rows=report_rows,members=members,selected_center=selected_center,branding=branding)


@app.route('/health/reports/print')
@login_required
@module_permission_required('rapports_sante','imprimer')
def health_report_print():
    user=current_user(); month=request.args.get('month') or datetime.now().strftime('%Y-%m'); center_id,province=_health_report_scope(user,request.args.get('center_id'),request.args.get('province',''))
    con=db(); data=build_health_monthly_report(con,month,center_id,province); scope_type='center' if center_id else ('province' if province else 'national'); db_center=center_id or 0; db_province=province or ''; saved=con.execute('SELECT * FROM health_monthly_reports WHERE report_month=? AND scope_type=? AND center_id=? AND province=?',(month,scope_type,db_center,db_province)).fetchone(); selected_center = con.execute('SELECT * FROM health_centers WHERE id=?',(center_id,)).fetchone() if center_id else None; branding = get_health_branding(selected_center); con.close()
    log_action(user['id'],'Impression rapport sanitaire','health_report',None,f'{data["scope_label"]} {month}')
    return render_template('print_health_report.html',data=data,saved=saved,selected_center=selected_center,branding=branding)


@app.route('/health/data-quality')
@login_required
@module_permission_required('statistiques','voir')
def health_data_quality_page():
    con=db(); user=current_user(); where,args=_health_scope_clause('hc')
    centers=con.execute('SELECT hc.* FROM health_centers hc'+where+' ORDER BY hc.province,hc.name',args).fetchall(); report=[]
    for c in centers:
        patients=con.execute('SELECT COUNT(*) n FROM patients WHERE center_id=?',(c['id'],)).fetchone()['n']; incomplete=con.execute("SELECT COUNT(*) n FROM patients WHERE center_id=? AND (birth_date IS NULL OR birth_date='' OR phone IS NULL OR phone='')",(c['id'],)).fetchone()['n']; waiting=con.execute("SELECT COUNT(*) n FROM health_visits WHERE center_id=? AND status IN ('waiting','triaged')",(c['id'],)).fetchone()['n']; low_stock=con.execute('SELECT COUNT(*) n FROM (SELECT hp.id,hp.minimum_stock,COALESCE(SUM(b.quantity),0) q FROM health_products hp LEFT JOIN health_stock_batches b ON b.product_id=hp.id WHERE hp.center_id=? GROUP BY hp.id HAVING q<=hp.minimum_stock)',(c['id'],)).fetchone()['n']; unpaid=con.execute("SELECT COALESCE(SUM(total-paid),0) n FROM health_invoices WHERE center_id=? AND status!='paid'",(c['id'],)).fetchone()['n']; score=100 if patients==0 else max(0,round(100-(incomplete/patients*40)-min(waiting,10)*2-min(low_stock,10)*2)); report.append({'center':c,'patients':patients,'incomplete':incomplete,'waiting':waiting,'low_stock':low_stock,'unpaid':unpaid,'score':score})
    con.close(); return render_template('health_data_quality.html',report=report)


DELETE_REGISTRY = {
    'health_center': ('health_centers', 'centres_sante', 'health_centers_page'),
    'health_role': ('health_roles', 'personnel_sante', 'health_roles_page'),
    'health_staff': ('health_staff', 'personnel_sante', 'health_staff_page'),
    'patient': ('patients', 'patients', 'patients_page'),
    'consultation': ('consultations', 'consultations', 'consultations_page'),
    'medical_form': ('medical_form_templates', 'modeles_fiches', 'medical_forms_page'),
    'service': ('foundation_services', 'services', 'foundation_services_page'),
    'service_structure': ('service_structures', 'structures_services', 'foundation_services_page'),
    'service_role': ('service_roles', 'roles_services', 'foundation_services_page'),
    'service_assignment': ('service_staff_assignments', 'affectations_services', 'foundation_services_page'),
    'service_module': ('service_modules', 'services', 'foundation_services_page'),
    'appointment': ('health_appointments', 'rendez_vous', 'health_operations_page'),
    'visit': ('health_visits', 'files_attente', 'health_operations_page'),
    'lab_order': ('health_lab_orders', 'laboratoire', 'health_operations_page'),
    'product': ('health_products', 'pharmacie', 'health_operations_page'),
    'stock_batch': ('health_stock_batches', 'stocks', 'health_operations_page'),
    'invoice': ('health_invoices', 'facturation', 'health_operations_page'),
    'admission': ('health_admissions', 'hospitalisation', 'health_operations_page'),
    'social_support': ('health_social_support', 'aide_sociale', 'health_operations_page'),
    'equipment': ('health_equipment', 'equipements', 'health_operations_page'),
    'supplier': ('health_suppliers', 'fournisseurs', 'health_operations_page'),
    'referral': ('health_referrals', 'references', 'health_operations_page'),
    'health_loss': ('health_losses', 'rapports_sante', 'health_reports_page'),
    'health_report': ('health_monthly_reports', 'rapports_sante', 'health_reports_page'),
    'health_card_request': ('health_card_requests', 'cartes_sanitaires', 'health_cards_page'),
    'executive_number': ('executive_numbers', 'numerotation_cadres', 'executive_numbers_page'),
}

@app.post('/admin/delete/<entity>/<int:record_id>')
@login_required
def universal_delete(entity, record_id):
    config = DELETE_REGISTRY.get(entity)
    if not config:
        abort(404)
    table, module, endpoint = config
    if not module_allowed(module, 'supprimer'):
        abort(403)
    con = db()
    row = con.execute(f'SELECT id FROM {table} WHERE id=? AND deleted_at IS NULL', (record_id,)).fetchone()
    if not row:
        con.close(); flash('Élément introuvable ou déjà supprimé.', 'warning'); return redirect(request.referrer or url_for(endpoint))
    try:
        con.execute(f'UPDATE {table} SET deleted_at=?, deleted_by=? WHERE id=?', (now(), current_user()['id'], record_id))
        con.commit()
        log_action(current_user()['id'], 'Suppression logique', table, record_id, entity)
        flash('Élément supprimé. Il reste traçable dans le journal.', 'success')
    except Exception as exc:
        con.rollback(); logging.exception('Erreur suppression %s/%s: %s', entity, record_id, exc); flash('La suppression a échoué.', 'danger')
    finally:
        con.close()
    return redirect(request.referrer or url_for(endpoint))

@app.route('/admin/executive-numbers', methods=['GET','POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def executive_numbers_page():
    con=db(); user=current_user()
    if request.method=='POST':
        action=request.form.get('action','reserve')
        if action=='reserve':
            number=request.form.get('number','').strip(); position=request.form.get('position_name','').strip()
            if not number or not position: flash('Numéro et fonction obligatoires.','warning')
            else:
                try:
                    con.execute('INSERT INTO executive_numbers(number,position_name,status,notes,created_at,created_by) VALUES(?,?,?,?,?,?)',(number,position,'reserved',request.form.get('notes',''),now(),user['id']))
                    con.commit(); flash('Numéro institutionnel réservé.','success')
                except sqlite3.IntegrityError: flash('Ce numéro est déjà réservé.','danger')
        elif action=='assign':
            row_id=request.form.get('row_id',type=int); member_id=request.form.get('member_id',type=int)
            row=con.execute('SELECT * FROM executive_numbers WHERE id=? AND deleted_at IS NULL',(row_id,)).fetchone()
            member=con.execute("SELECT * FROM members WHERE id=? AND deleted_at IS NULL AND status='active'",(member_id,)).fetchone()
            if not row or not member: flash('Numéro ou membre invalide.','danger')
            else:
                con.execute("UPDATE executive_numbers SET member_id=?,status='assigned',assigned_at=?,assigned_by=?,updated_at=? WHERE id=?",(member_id,now(),user['id'],now(),row_id))
                con.execute('UPDATE members SET executive_number=?,executive_position=?,updated_at=? WHERE id=?',(row['number'],row['position_name'],now(),member_id))
                con.commit(); log_action(user['id'],'Attribution numéro haut cadre','member',member_id,row['number']); flash('Numéro attribué au haut cadre.','success')
        elif action=='release':
            row_id=request.form.get('row_id',type=int); row=con.execute('SELECT * FROM executive_numbers WHERE id=?',(row_id,)).fetchone()
            if row:
                if row['member_id']: con.execute('UPDATE members SET executive_number=NULL,executive_position=NULL,updated_at=? WHERE id=?',(now(),row['member_id']))
                con.execute("UPDATE executive_numbers SET member_id=NULL,status='reserved',released_at=?,updated_at=? WHERE id=?",(now(),now(),row_id)); con.commit(); flash('Numéro libéré et conservé comme réservé.','success')
    rows=con.execute("SELECT e.*,m.code member_code,m.first_name,m.last_name FROM executive_numbers e LEFT JOIN members m ON m.id=e.member_id WHERE e.deleted_at IS NULL ORDER BY CAST(e.number AS INTEGER),e.number").fetchall()
    members=con.execute("SELECT id,code,first_name,last_name,province FROM members WHERE deleted_at IS NULL AND status='active' ORDER BY last_name,first_name").fetchall(); con.close()
    return render_template('executive_numbers.html',rows=rows,members=members)

@app.route('/health/cards', methods=['GET','POST'])
@login_required
@module_permission_required('cartes_sanitaires','voir')
def health_cards_page():
    con=db(); user=current_user()
    if request.method=='POST':
        action=request.form.get('action','request'); rid=request.form.get('request_id',type=int)
        if action=='request':
            staff_id=request.form.get('staff_id',type=int)
            staff=con.execute("SELECT hs.*,hc.province,hc.id center_id FROM health_staff hs JOIN health_centers hc ON hc.id=hs.center_id WHERE hs.id=? AND hs.deleted_at IS NULL",(staff_id,)).fetchone()
            if not staff: flash('Affectation sanitaire introuvable.','danger')
            else:
                code=_next_code('CARTE-SANTE','health_card_requests')
                con.execute('INSERT INTO health_card_requests(request_code,staff_id,center_id,province,reason,status,requested_at,requested_by,notes) VALUES(?,?,?,?,?,?,?,?,?)',(code,staff_id,staff['center_id'],staff['province'],request.form.get('reason','creation'),'requested',now(),user['id'],request.form.get('notes',''))); con.commit(); flash('Demande de carte transmise.','success')
        else:
            row=con.execute('SELECT * FROM health_card_requests WHERE id=? AND deleted_at IS NULL',(rid,)).fetchone()
            if not row: flash('Demande introuvable.','danger')
            elif action=='province_validate' and (user['role'] in PROVINCIAL_ROLES or user['role'] in NATIONAL_ROLES):
                con.execute("UPDATE health_card_requests SET status='provincial_validated',provincial_review_at=?,provincial_review_by=? WHERE id=?",(now(),user['id'],rid)); con.commit(); flash('Demande validée au niveau provincial.','success')
            elif action=='national_approve' and user['role'] in NATIONAL_ROLES:
                card_no=request.form.get('card_number','').strip() or f"SANTE-{datetime.now().year}-{rid:06d}"
                expires=request.form.get('expires_at') or (datetime.now()+timedelta(days=365)).strftime('%Y-%m-%d')
                con.execute("UPDATE health_card_requests SET status='approved',national_approval_at=?,national_approval_by=?,card_number=?,expires_at=? WHERE id=?",(now(),user['id'],card_no,expires,rid)); con.commit(); flash('Carte approuvée par l’administration centrale.','success')
            elif action=='printed' and user['role'] in NATIONAL_ROLES:
                con.execute("UPDATE health_card_requests SET status='printed',printed_at=?,printed_by=? WHERE id=?",(now(),user['id'],rid)); con.commit(); flash('Impression enregistrée.','success')
            elif action=='delivered' and user['role'] in NATIONAL_ROLES:
                con.execute("UPDATE health_card_requests SET status='delivered',delivered_at=?,delivered_by=? WHERE id=?",(now(),user['id'],rid)); con.commit(); flash('Remise de la carte enregistrée.','success')
    where=''; args=[]
    if user['role'] in PROVINCIAL_ROLES: where=' WHERE hc.province=?'; args=[user['province'] or '']
    staff=con.execute("SELECT hs.id,m.code,m.first_name,m.last_name,hc.name center_name,hc.province FROM health_staff hs JOIN members m ON m.id=hs.member_id JOIN health_centers hc ON hc.id=hs.center_id"+where+' ORDER BY m.last_name,m.first_name',args).fetchall()
    req_where=' WHERE r.deleted_at IS NULL'; req_args=[]
    if user['role'] in PROVINCIAL_ROLES: req_where+=' AND r.province=?'; req_args=[user['province'] or '']
    rows=con.execute("SELECT r.*,m.code member_code,m.first_name,m.last_name,hc.name center_name,hr.name role_name FROM health_card_requests r JOIN health_staff hs ON hs.id=r.staff_id JOIN members m ON m.id=hs.member_id JOIN health_centers hc ON hc.id=r.center_id JOIN health_roles hr ON hr.id=hs.role_id"+req_where+' ORDER BY r.requested_at DESC',req_args).fetchall(); con.close()
    return render_template('health_cards.html',rows=rows,staff=staff)

@app.route('/health/cards/<int:request_id>/print')
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def health_card_print(request_id):
    con=db(); row=con.execute("SELECT r.*,m.code member_code,m.first_name,m.last_name,m.photo_path,hc.name center_name,hc.logo_path center_logo,hr.name role_name,hs.specialty FROM health_card_requests r JOIN health_staff hs ON hs.id=r.staff_id JOIN members m ON m.id=hs.member_id JOIN health_centers hc ON hc.id=r.center_id JOIN health_roles hr ON hr.id=hs.role_id WHERE r.id=? AND r.deleted_at IS NULL",(request_id,)).fetchone(); con.close()
    if not row or row['status'] not in ('approved','printed','delivered'): abort(404)
    return render_template('print_health_card.html',row=row)

@app.route('/health/patients/<int:patient_id>/print')
@login_required
@module_permission_required('patients','imprimer')
def patient_print(patient_id):
    con=db(); row=con.execute("SELECT p.*,hc.name center_name,hc.logo_path center_logo,hc.logo_path logo_path,hc.header_title,hc.header_subtitle,hc.header_address,hc.header_contact,hc.footer_note FROM patients p JOIN health_centers hc ON hc.id=p.center_id WHERE p.id=? AND p.deleted_at IS NULL",(patient_id,)).fetchone(); con.close()
    if not row: abort(404)
    return render_template('print_patient.html',row=row,branding=get_health_branding(row))

@app.route('/health/consultations/<int:consultation_id>/print')
@login_required
@module_permission_required('consultations','imprimer')
def consultation_print(consultation_id):
    con=db(); row=con.execute("SELECT co.*,p.patient_code,p.first_name,p.last_name,hc.name center_name,hc.logo_path center_logo,hc.logo_path logo_path,hc.header_title,hc.header_subtitle,hc.header_address,hc.header_contact,hc.footer_note FROM consultations co JOIN patients p ON p.id=co.patient_id JOIN health_centers hc ON hc.id=co.center_id WHERE co.id=? AND co.deleted_at IS NULL",(consultation_id,)).fetchone(); con.close()
    if not row: abort(404)
    return render_template('print_consultation.html',row=row,branding=get_health_branding(row))

@app.errorhandler(403)
def forbidden(_):
    return render_template("error.html", message="Accès refusé : votre rôle ne permet pas cette action."), 403


@app.errorhandler(404)
def not_found(_):
    return render_template("error.html", message="Page introuvable."), 404


@app.errorhandler(413)
def too_large(_):
    return render_template("error.html", message="Le fichier dépasse la taille autorisée de 128 Mo."), 413


@app.errorhandler(500)
def internal_error(error):
    logging.error("Erreur interne: %s\n%s", error, traceback.format_exc())
    return render_template("error.html", message="Une erreur interne est survenue. Aucun détail technique sensible n’est affiché. Consultez le journal administrateur."), 500


init_db()

@app.route('/admin/synchronisation', methods=['GET','POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def synchronization_center():
    user=current_user(); settings=get_settings()
    if request.method=='POST':
        remote=request.form.get('sync_remote_url','').strip().rstrip('/')
        enabled='1' if request.form.get('sync_enabled')=='1' else '0'
        api_key=request.form.get('sync_api_key','').strip()
        con=db()
        for k,v in [('sync_remote_url',remote),('sync_enabled',enabled),('sync_api_key',api_key)]:
            con.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",(k,v))
        con.commit(); con.close(); log_action(user['id'],'Configuration synchronisation','sync',None,remote)
        flash('Configuration de synchronisation enregistrée.','success')
        return redirect(url_for('synchronization_center'))
    con=db()
    queue=con.execute("SELECT * FROM sync_queue ORDER BY created_at DESC LIMIT 100").fetchall()
    conflicts=con.execute("SELECT * FROM sync_conflicts WHERE status='open' ORDER BY created_at DESC").fetchall()
    counts=con.execute("SELECT status,COUNT(*) n FROM sync_queue GROUP BY status").fetchall()
    con.close()
    remote_ok=False; remote_message='Serveur distant non configuré.'
    remote=settings.get('sync_remote_url','')
    if remote:
        try:
            import urllib.request
            req=urllib.request.Request(remote, method='HEAD')
            with urllib.request.urlopen(req, timeout=4) as r:
                remote_ok=200 <= r.status < 500; remote_message=f'Serveur joignable (HTTP {r.status}).'
        except Exception as e:
            remote_message='Serveur distant indisponible : '+str(e)[:120]
    return render_template('synchronization_center.html',queue=queue,conflicts=conflicts,counts=counts,remote_ok=remote_ok,remote_message=remote_message)

@app.route('/admin/synchronisation/run', methods=['POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def run_synchronization():
    user=current_user(); s=get_settings()
    if s.get('sync_enabled')!='1' or not s.get('sync_remote_url'):
        flash('Configurez et activez d’abord le serveur distant.','warning')
        return redirect(url_for('synchronization_center'))
    con=db(); pending=con.execute("SELECT COUNT(*) n FROM sync_queue WHERE status='pending'").fetchone()['n']; con.close()
    log_action(user['id'],'Synchronisation manuelle','sync',None,f'{pending} élément(s) en attente')
    flash(f'Synchronisation préparée : {pending} élément(s) en attente. La transmission réelle sera activée après configuration de l’API du serveur en ligne.','info')
    return redirect(url_for('synchronization_center'))

@app.route('/admin/corbeille')
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def recycle_bin():
    con=db()
    members=con.execute("SELECT * FROM members WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall()
    users=con.execute("SELECT * FROM users WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall()
    docs=con.execute("SELECT * FROM official_documents WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall()
    con.close()
    return render_template('recycle_bin.html',members=members,users=users,docs=docs)

@app.route('/admin/corbeille/member/<int:item_id>/restore',methods=['POST'])
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def restore_deleted_member(item_id):
    user=current_user(); con=db(); con.execute("UPDATE members SET deleted_at=NULL,status='active',updated_at=? WHERE id=?",(now(),item_id)); con.commit(); con.close(); log_action(user['id'],'Restauration corbeille','member',item_id); flash('Membre restauré.','success'); return redirect(url_for('recycle_bin'))

@app.route('/admin/audit')
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def audit_center():
    q=request.args.get('q','').strip(); action=request.args.get('action','').strip(); status=request.args.get('status','').strip()
    where=['1=1']; params=[]
    if q:
        like=f'%{q}%'; where.append('(u.first_name LIKE ? OR u.last_name LIKE ? OR u.email LIKE ? OR a.action LIKE ? OR a.details LIKE ?)'); params += [like]*5
    if action:
        where.append('a.action=?'); params.append(action)
    if status:
        where.append('a.status=?'); params.append(status)
    con=db()
    rows=con.execute(f"""SELECT a.*,u.email,u.first_name,u.last_name,u.role,u.province
                         FROM audit_logs a LEFT JOIN users u ON u.id=a.user_id
                         WHERE {' AND '.join(where)} ORDER BY a.created_at DESC LIMIT 1000""",params).fetchall()
    actions=con.execute('SELECT DISTINCT action FROM audit_logs ORDER BY action').fetchall()
    stats=con.execute("""SELECT COUNT(*) total, COUNT(DISTINCT user_id) users,
                       SUM(CASE WHEN status='success' OR status IS NULL THEN 1 ELSE 0 END) success_count,
                       SUM(CASE WHEN status='failed' THEN 1 ELSE 0 END) failed_count FROM audit_logs""").fetchone()
    con.close()
    return render_template('audit_center.html',rows=rows,actions=actions,stats=stats,q=q,selected_action=action,selected_status=status)

@app.route('/admin/diagnostic')
@login_required
@role_required('super_admin','president','secretary','national_secretary')
def diagnostic_center():
    return render_template('diagnostic_center.html',status=production_health_status(),readiness=release_readiness_status())


ensure_daily_backup()

if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=5000)
